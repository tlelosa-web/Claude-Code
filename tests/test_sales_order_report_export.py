"""Tests for GET /sales-orders/export-excel and
services/sales_order_report_export.py -- see docs/specs/
sales-order-report-excel-export-2026-07-17.md Sequencing item 20.

The in-memory test DB is shared/cumulative across the whole pytest session
(see test_dashboard.py's `_card_values()` docstring for the same
convention). Row-content assertions here locate a uniquely-named fixture
row by SO Number within the exported sheet rather than assuming a fixed
row index. Aggregate/summary-math assertions call
`build_sales_order_report_workbook()` directly with an explicit, fully
controlled `sales_orders`/`invoices` list instead of querying the shared
DB, so they're immune to cross-test contamination (other test files add
SalesOrder/Invoice rows that would otherwise land in the same
current-month/next-month buckets).
"""
import io
from datetime import date, datetime, timedelta, timezone

import openpyxl

from models import SalesOrder, SOLineItem, WorksOrder, Invoice
from services.sales_order_report_export import build_sales_order_report_workbook, SHEET1_HEADERS, _next_month


def _load_workbook(resp):
    return openpyxl.load_workbook(io.BytesIO(resp.data))


def _sheet1_rows(ws):
    """Yield {header: value} dicts for every data row (row 3 onward),
    stopping at the first row with a blank SO Number cell (column D) --
    i.e. the blank row separating the data from the summary block."""
    row_idx = 3
    while True:
        so_number = ws.cell(row=row_idx, column=4).value
        if not so_number:
            return
        yield {h: ws.cell(row=row_idx, column=i + 1).value for i, h in enumerate(SHEET1_HEADERS)}
        row_idx += 1


def _find_row(ws, so_number):
    for row in _sheet1_rows(ws):
        if row['SO Number'] == so_number:
            return row
    return None


class TestExportRouteBasics:
    def test_export_returns_loadable_xlsx(self, client):
        resp = client.get('/sales-orders/export-excel')
        assert resp.status_code == 200
        assert resp.headers['Content-Type'] == (
            'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        assert 'Sales_Order_Report_' in resp.headers['Content-Disposition']
        wb = _load_workbook(resp)  # raises if not a valid xlsx
        assert 'Sales Order Report' in wb.sheetnames

    def test_sheet1_headers_match_spec_column_list(self, client):
        resp = client.get('/sales-orders/export-excel')
        ws = _load_workbook(resp)['Sales Order Report']
        headers = [ws.cell(row=2, column=i + 1).value for i in range(len(SHEET1_HEADERS))]
        assert headers == SHEET1_HEADERS

    def test_sheet2_present_with_current_month_name(self, client):
        resp = client.get('/sales-orders/export-excel')
        wb = _load_workbook(resp)
        expected_name = f"Monthly INV - {date.today().strftime('%m.%Y')}"
        assert expected_name in wb.sheetnames


class TestExportRowContent:
    _c = 0

    def _mk_so(self, session, **kwargs):
        TestExportRowContent._c += 1
        n = TestExportRowContent._c
        defaults = dict(
            so_number=f"EXPORT-ROW-{n:03d}", customer_name="Export Row Test Co",
            status="Open", payment_status="Cash Sale - Paid", job_numbers="FM9001",
            created_at=datetime.now(timezone.utc),
        )
        defaults.update(kwargs)
        so = SalesOrder(**defaults)
        session.add(so)
        session.flush()
        session.add(SOLineItem(so_id=so.id, description="Line 1", qty=1.0, incl_total=1000.0))
        session.commit()
        return so

    def test_row_shows_job_number_status_and_payment_status(self, client, session):
        so = self._mk_so(session)
        session.add(WorksOrder(wo_number=f"EXPORT-ROW-WO-{so.id}", so_id=so.id,
                                order_type='ASSEMBLY', status='Open'))
        session.commit()
        assert so.report_status == 'Released'

        resp = client.get('/sales-orders/export-excel')
        ws = _load_workbook(resp)['Sales Order Report']
        row = _find_row(ws, so.so_number)

        assert row is not None
        assert row['Job Number'] == 'FM9001'
        assert row['Status'] == 'Released'
        assert row['Payment Status'] == 'Cash Sale - Paid'


class TestExportBalanceDueNotTotalIncl:
    """The single most important assertion in this chunk (Decision 3):
    the exported Total column must equal balance_due, not total_incl."""
    _c = 0

    def _mk_so(self, session):
        TestExportBalanceDueNotTotalIncl._c += 1
        n = TestExportBalanceDueNotTotalIncl._c
        so = SalesOrder(so_number=f"EXPORT-BAL-{n:03d}", customer_name="Export Balance Test Co",
                        status="Draft", payment_status="Cash Sale - Partial", amount_paid=300.0,
                        created_at=datetime.now(timezone.utc))
        session.add(so)
        session.flush()
        session.add(SOLineItem(so_id=so.id, description="Line 1", qty=1.0, incl_total=1000.0))
        session.commit()
        return so

    def test_exported_total_equals_balance_due_not_total_incl(self, client, session):
        so = self._mk_so(session)
        assert so.total_incl == 1000.0
        assert so.balance_due == 700.0

        resp = client.get('/sales-orders/export-excel')
        ws = _load_workbook(resp)['Sales Order Report']
        row = _find_row(ws, so.so_number)

        assert row is not None
        assert row['Total'] == so.balance_due == 700.0
        assert row['Total'] != so.total_incl


class TestExportViewFilter:
    _c = 0

    def _mk_so(self, session, status):
        TestExportViewFilter._c += 1
        n = TestExportViewFilter._c
        so = SalesOrder(so_number=f"EXPORT-VIEW-{n:03d}", customer_name="Export View Test Co",
                        status=status, created_at=datetime.now(timezone.utc))
        session.add(so)
        session.commit()
        return so

    def test_default_open_view_excludes_closed_so(self, client, session):
        so = self._mk_so(session, status='Closed')
        resp = client.get('/sales-orders/export-excel')
        ws = _load_workbook(resp)['Sales Order Report']
        assert _find_row(ws, so.so_number) is None

    def test_all_view_includes_closed_so_with_dash_status(self, client, session):
        so = self._mk_so(session, status='Closed')
        assert so.report_status is None

        resp = client.get('/sales-orders/export-excel?view=all')
        ws = _load_workbook(resp)['Sales Order Report']
        row = _find_row(ws, so.so_number)

        assert row is not None
        assert row['Status'] == '-'


class TestSheet1SummaryMath:
    """Exercises build_sales_order_report_workbook() directly with an
    explicit, controlled sales_orders list -- immune to the shared test DB's
    other fixtures landing in the same current/next-month buckets."""
    _c = 0

    def _mk_so(self, session, incl_total, payment_status='Cash Sale - Paid', amount_paid=0.0,
               delivery_date=None, wo_status=None):
        TestSheet1SummaryMath._c += 1
        n = TestSheet1SummaryMath._c
        so = SalesOrder(so_number=f"SUMMARY-SO-{n:03d}", customer_name="Summary Test Co",
                        status="Open", payment_status=payment_status, amount_paid=amount_paid,
                        delivery_date=delivery_date, created_at=datetime.now(timezone.utc))
        session.add(so)
        session.flush()
        session.add(SOLineItem(so_id=so.id, description="Line", qty=1.0, incl_total=incl_total))
        if wo_status:
            session.add(WorksOrder(wo_number=f"SUMMARY-WO-{n:03d}", so_id=so.id,
                                    order_type='ASSEMBLY', status=wo_status))
        session.commit()
        return so

    def test_total_value_and_breakdown_sum_balance_due(self, session):
        today = date.today()
        nxt_y, nxt_m = _next_month(today.year, today.month)
        next_month_date = date(nxt_y, nxt_m, 15)

        so_cash_cur = self._mk_so(session, 1000.0, payment_status='Cash Sale - Paid', delivery_date=today)
        so_acct_cur = self._mk_so(session, 500.0, payment_status='Account - Pending', delivery_date=today)
        so_cash_nxt = self._mk_so(session, 200.0, payment_status='Cash Sale - Unpaid', delivery_date=next_month_date)
        so_partial = self._mk_so(session, 1000.0, payment_status='Cash Sale - Partial',
                                  amount_paid=300.0, delivery_date=today)

        orders = [so_cash_cur, so_acct_cur, so_cash_nxt, so_partial]
        wb = build_sales_order_report_workbook(orders, invoices=[], currency_symbol='R')
        ws = wb['Sales Order Report']

        last_data_row = 2 + len(orders)
        summary_row = last_data_row + 2
        assert ws.cell(row=summary_row, column=6).value == 'Total Value:'
        expected_total = sum(so.balance_due for so in orders)
        assert ws.cell(row=summary_row, column=7).value == round(expected_total, 2)

        breakdown_header_row = summary_row + 2
        assert ws.cell(row=breakdown_header_row, column=6).value == 'BREAKDOWN'
        r1 = breakdown_header_row + 1  # Cash Sale row
        r2 = breakdown_header_row + 2  # Account row

        assert ws.cell(row=r1, column=7).value == round(so_cash_cur.balance_due + so_partial.balance_due, 2)
        assert ws.cell(row=r2, column=7).value == round(so_acct_cur.balance_due, 2)
        assert ws.cell(row=r1, column=10).value == round(so_cash_nxt.balance_due, 2)
        assert ws.cell(row=r2, column=10).value == 0.0

    def test_due_today_and_ready_dispatch_totals(self, session):
        today = date.today()
        so_due_today = self._mk_so(session, 333.0, delivery_date=today)
        so_ready = self._mk_so(session, 444.0, delivery_date=today + timedelta(days=10), wo_status='Complete')
        so_not_ready = self._mk_so(session, 999.0, delivery_date=today + timedelta(days=10), wo_status='Open')
        assert so_ready.report_status == 'Ready-Dispatch'
        assert so_not_ready.report_status == 'Released'

        orders = [so_due_today, so_ready, so_not_ready]
        wb = build_sales_order_report_workbook(orders, invoices=[], currency_symbol='R')
        ws = wb['Sales Order Report']

        last_data_row = 2 + len(orders)
        summary_row = last_data_row + 2
        breakdown_header_row = summary_row + 2
        r2 = breakdown_header_row + 2
        dt_header_row = r2 + 3

        assert ws.cell(row=dt_header_row + 1, column=4).value == so_due_today.so_number
        total_row = dt_header_row + 2
        assert ws.cell(row=total_row, column=6).value == 'Due Today - Total:'
        assert ws.cell(row=total_row, column=7).value == round(so_due_today.balance_due, 2)

        rd_row = total_row + 2
        assert ws.cell(row=rd_row, column=6).value == 'Ready-Dispatch:'
        assert ws.cell(row=rd_row, column=7).value == round(so_ready.balance_due, 2)


class TestSheet2InvoiceMonthBuckets:
    """Sheet 2 is sourced entirely from the invoices list -- controlled
    fixtures passed directly, no DB query, so no contamination risk from
    other test files' Invoice rows."""

    def test_current_and_next_month_cash_account_buckets(self):
        today = date.today()
        nxt_y, nxt_m = _next_month(today.year, today.month)
        next_month_date = date(nxt_y, nxt_m, 10)

        far_due = date(today.year + 1, today.month, 1)
        if far_due.month == today.month and far_due.year == today.year:
            far_due = far_due.replace(year=far_due.year + 1)

        cash_cur = Invoice(invoice_number='SHEET2-CASH-CUR', invoice_date=today, due_date=today,
                            exclusive=100.0, total_selling=115.0)
        acct_cur = Invoice(invoice_number='SHEET2-ACCT-CUR', invoice_date=today, due_date=far_due,
                            exclusive=200.0, total_selling=230.0)
        cash_nxt = Invoice(invoice_number='SHEET2-CASH-NXT', invoice_date=next_month_date, due_date=next_month_date,
                            exclusive=50.0, total_selling=57.5)

        wb = build_sales_order_report_workbook([], [cash_cur, acct_cur, cash_nxt], currency_symbol='R')
        ws = wb[f"Monthly INV - {today.strftime('%m.%Y')}"]

        # Current month block: row 3 (Cash Sale), row 4 (Account), row 5 (Total)
        assert ws.cell(row=3, column=3).value == 100.0
        assert ws.cell(row=3, column=6).value == 115.0
        assert ws.cell(row=4, column=3).value == 200.0
        assert ws.cell(row=4, column=6).value == 230.0
        assert ws.cell(row=5, column=3).value == 300.0
        assert ws.cell(row=5, column=6).value == 345.0

        # Next month block starts at row 7 (5 + 2 blank-row gap)
        assert ws.cell(row=7, column=3).value == 50.0
        assert ws.cell(row=7, column=6).value == 57.5
        assert ws.cell(row=8, column=3).value == 0.0
        assert ws.cell(row=8, column=6).value == 0.0
