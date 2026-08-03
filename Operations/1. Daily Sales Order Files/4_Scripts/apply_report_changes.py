import openpyxl
import os
import glob
import datetime
import re

# File path
def resolve_sales_report_file(dt):
    reports_dir = "3_Live_Reports"
    month_tag = dt.strftime("%m.%Y")
    preferred = os.path.join(reports_dir, f"Ops Sales Order Report - {month_tag}.xlsx")
    legacy = os.path.join(reports_dir, f"Sales Order Report - {month_tag}.xlsx")

    if os.path.exists(preferred):
        return preferred

    if os.path.exists(legacy):
        try:
            os.rename(legacy, preferred)
            return preferred
        except Exception:
            return legacy

    candidates = []
    for pattern in (
        os.path.join(reports_dir, "Ops Sales Order Report - *.xlsx"),
        os.path.join(reports_dir, "Sales Order Report - *.xlsx"),
    ):
        candidates.extend(glob.glob(pattern))

    if candidates:
        candidates.sort(key=lambda p: os.path.getmtime(p))
        return candidates[-1]

    return preferred

SALES_REPORT_FILE = resolve_sales_report_file(datetime.datetime.now())

# Check if file exists
if not os.path.exists(SALES_REPORT_FILE):
    print(f"ERROR: {SALES_REPORT_FILE} not found.")
    exit(1)

# Open the workbook
print(f"Opening {SALES_REPORT_FILE}...")
try:
    wb = openpyxl.load_workbook(SALES_REPORT_FILE)
except PermissionError:
    print("ERROR: File is open. Please close it and try again.")
    exit(1)

# Get the active sheet or the sheet with today's date
ws = wb.active

# If there are multiple sheets, find the one with the date format (DD.MM.YYYY)
date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
date_sheets = [s for s in wb.sheetnames if date_pattern.match(s)]

if date_sheets:
    # Use the most recent date sheet
    date_sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
    ws = wb[date_sheets[-1]]
    print(f"Working on sheet: {ws.title}")

# Make the three requested changes:

# 1. Set column J width to 21.22
print("Setting column J width to 21.22...")
ws.column_dimensions['J'].width = 23.3

# 2. Move date from K1 to L1
# Read the value/formula from K1
k1_value = ws.cell(row=1, column=11).value
k1_font = ws.cell(row=1, column=11).font.copy() if ws.cell(row=1, column=11).font else None
k1_number_format = ws.cell(row=1, column=11).number_format

# Clear K1
ws.cell(row=1, column=11).value = None

# Write to L1
l1_cell = ws.cell(row=1, column=12)
l1_cell.value = k1_value
if k1_font:
    l1_cell.font = k1_font
l1_cell.number_format = k1_number_format

print("Moved date from K1 to L1...")

# 3. Remove 1 empty row between Breakdown & Summary
# Find "BREAKDOWN" text in the sheet
breakdown_row = None
for row_idx in range(1, ws.max_row + 1):
    cell_value = ws.cell(row=row_idx, column=6).value  # Column F
    if cell_value and str(cell_value).strip().upper() == "BREAKDOWN":
        breakdown_row = row_idx
        break

if breakdown_row:
    print(f"Found BREAKDOWN at row {breakdown_row}")
    # The empty row should be right before BREAKDOWN
    # Check if the row before BREAKDOWN is empty
    empty_row = breakdown_row - 1
    
    # Check if this row is truly empty (at least in columns F and surrounding)
    is_empty = True
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
        cell_val = ws[f'{col}{empty_row}'].value
        if cell_val is not None and str(cell_val).strip() != '':
            is_empty = False
            break
    
    if is_empty:
        print(f"Removing empty row {empty_row}...")
        ws.delete_rows(empty_row, 1)
    else:
        print(f"Warning: Row {empty_row} is not empty. Skipping removal.")
else:
    print("Warning: BREAKDOWN section not found. Skipping row removal.")

# Save the workbook
print("Saving workbook...")
try:
    wb.save(SALES_REPORT_FILE)
    print("Done! All changes applied successfully.")
except PermissionError:
    print("ERROR: Could not save file. It might be open.")
    exit(1)
