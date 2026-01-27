from __future__ import annotations
from pathlib import Path
import shutil
import time
import uuid

import openpyxl

from motor_fla_lookup import lookup_fla_safe


def _norm(s: str) -> str:
    return "".join(str(s).strip().lower().split())


def _read_block_by_labels(ws, label_col: str, value_col: str) -> dict:
    out = {}
    for r in range(1, 250):
        lab = ws[f"{label_col}{r}"].value
        if lab is None:
            continue
        key = _norm(lab)
        val = ws[f"{value_col}{r}"].value
        out[key] = "" if val is None else str(val).strip()
    return out


def _to_float(s: str):
    try:
        return float(str(s).strip())
    except Exception:
        return None


def _to_int(s: str):
    try:
        return int(float(str(s).strip()))
    except Exception:
        return None


def _safe_copy_excel(src: Path) -> Path:
    tmp = src.parent / f"_tmp_excel_{uuid.uuid4().hex}.xlsx"
    last_err = None
    for _ in range(3):
        try:
            shutil.copy2(src, tmp)
            return tmp
        except Exception as e:
            last_err = e
            time.sleep(0.2)
    raise PermissionError(f"Cannot read Excel file: {last_err}")


def read_nameplate_from_excel(xlsx_path: str) -> dict:
    src = Path(xlsx_path)
    if not src.exists():
        raise FileNotFoundError(f"Excel not found: {src}")

    tmp = _safe_copy_excel(src)

    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
        ws = wb["Table 1"]

        left = _read_block_by_labels(ws, "G", "H")
        right = _read_block_by_labels(ws, "J", "K")

        def pick(d: dict, *labels: str, default: str = "") -> str:
            for lab in labels:
                k = _norm(lab)
                if k in d and d[k] != "":
                    return d[k]
            for lab in labels:
                k = _norm(lab)
                if k in d:
                    return d[k]
            return default

        # LEFT
        series = pick(left, "Series")
        class_pitch = pick(left, "Class/Pitch", "Class Pitch", "Pitch", "Class")
        motor_kw_str = pick(left, "Motor")
        voltage_str = pick(left, "Voltage")
        pole_str = pick(left, "Pole", "Poles")
        fla_str = pick(left, "F.L.A", "F.L.A.", "FLA")
        op_temp = pick(left, "Op Temp", default="20")
        serial_no = pick(left, "Serial No")

        # RIGHT
        size = pick(right, "Size")
        op_speed = pick(right, "Op Speed")
        max_speed = op_speed
        phase = pick(right, "Phase")
        frequency = pick(right, "Frequency", default="50")
        connection = pick(right, "Conn", "Connection")
        date_of_manuf = pick(right, "Date of Manuf", "Date of Manufacture")

        relube_interval = "N/A"

        # ---- SAFE AUTO FLA ----
        fla_error = None
        motor_kw = _to_float(motor_kw_str)
        voltage = _to_float(voltage_str)
        poles = _to_int(pole_str)
        fla_val = _to_float(fla_str)

        if fla_val is None and motor_kw and voltage and poles:
            pdf_path = Path(__file__).resolve().parent / "2025 - CTP 022- PB4  Performance Data Rev 0.pdf"
            fla_str, fla_error = lookup_fla_safe(motor_kw, poles, voltage, str(pdf_path))

        return {
            "series": series,
            "class_pitch": class_pitch,
            "motor": motor_kw_str,
            "voltage": voltage_str,
            "fla": fla_str,
            "op_temp": op_temp,
            "serial_no": serial_no,
            "size": size,
            "op_speed": op_speed,
            "max_speed": max_speed,
            "phase": phase,
            "frequency": frequency,
            "connection": connection,
            "date_of_manuf": date_of_manuf,
            "relube_interval": relube_interval,
            # debug-only
            "pole": pole_str,
            "fla_error": fla_error,
        }

    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
