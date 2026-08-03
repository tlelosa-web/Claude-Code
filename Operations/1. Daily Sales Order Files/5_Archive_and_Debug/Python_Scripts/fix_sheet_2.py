import openpyxl

file_path = "Sales Order Report - 03.2026.xlsx"
wb = openpyxl.load_workbook(file_path)
sheet_name = "23.03.2026"

if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    
    # User's manual fixes for bloated CSV rows on today's sheet
    # F26 -> F25
    # F25 -> I24
    # G26 -> J24
    # G27 -> J25
    
    val_f26 = ws["F26"].value
    val_f25 = ws["F25"].value
    val_g26 = ws["G26"].value
    val_g27 = ws["G27"].value
    
    ws["F25"].value = val_f26
    ws["I24"].value = val_f25
    ws["J24"].value = val_g26
    ws["J25"].value = val_g27
    
    ws["F26"].value = None
    ws["G26"].value = None
    ws["G27"].value = None
    
    wb.save(file_path)
    print("One-off cell movements applied.")
else:
    print(f"Sheet {sheet_name} not found.")
