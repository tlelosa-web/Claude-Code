import os
import re
import pandas as pd
import openpyxl
import datetime
import glob

# --- Configuration ---
COMMENTS_FILE = "5_Archive_and_Debug\\Released_Jobs_Comments.xlsx"

ALLOWED_STATUSES = [
    "CASH SALE - UNPAID",
    "CASH SALE - PAID",
    "CASH SALE – PARTIAL",
    "ACCOUNT - OVER LIMIT",
    "ACCOUNT - UP TO DATE",
    "ACCOUNT - PENDING"
]

def resolve_sales_report_file(dt):
    reports_dir = "3_Live_Reports"
    month_tag = dt.strftime("%m.%Y")
    preferred = os.path.join(reports_dir, f"Ops Sales Order Report - {month_tag}.xlsx")
    legacy = os.path.join(reports_dir, f"Sales Order Report - {month_tag}.xlsx")

    if os.path.exists(preferred):
        return preferred

    if os.path.exists(legacy):
        try:
            os.rename(legacy, preferred)
            return preferred
        except Exception:
            return legacy

    # Always use the current month output file when starting a new month.
    return preferred

def map_comment_to_status(comment):
    """
    Applies strict validations to map text box comments 
    to one of the 6 allowed column J statuses.
    """
    if pd.isna(comment) or not str(comment).strip() or comment == "NO TEXT BOXES FOUND" or comment == "ERROR READING PDF":
        return None
    
    text = str(comment).upper()
    
    # Normalizing dashes for consistency
    normalized_text = text.replace('–', '-').replace('—', '-')
    
    # 1. Look for Exact Allowed Strings (highest confidence)
    if "ACCOUNT - OVER LIMIT" in normalized_text: return "ACCOUNT - OVER LIMIT"
    if "ACCOUNT - UP TO DATE" in normalized_text: return "ACCOUNT - UP TO DATE"
    if "ACCOUNT - PENDING" in normalized_text: return "ACCOUNT - PENDING"
    if "CASH SALE - UNPAID" in normalized_text: return "CASH SALE - UNPAID"
    if "CASH SALE - PAID" in normalized_text: return "CASH SALE - PAID"
    if "CASH SALE - PARTIAL" in normalized_text: return "CASH SALE – PARTIAL"
    
    # 2. Look for explicit loose combinations without getting tricked by form fields
    if "ACCOUNT" in normalized_text and "OVER LIMIT" in normalized_text: return "ACCOUNT - OVER LIMIT"
    if "ACCOUNT" in normalized_text and "UP TO DATE" in normalized_text: return "ACCOUNT - UP TO DATE"
    
    if "CASH SALE" in normalized_text and "UNPAID" in normalized_text: return "CASH SALE - UNPAID"
    if "CASH SALE" in normalized_text and "PAID" in normalized_text and "UNPAID" not in normalized_text: return "CASH SALE - PAID"
    if "CASH SALE" in normalized_text and "PARTIAL" in normalized_text: return "CASH SALE – PARTIAL"
    
    # 3. Fallbacks for very short text boxes (e.g. Sales Rep just dropping a sticky note that says "PAID")
    if len(normalized_text) < 100:
        if "PARTIAL" in normalized_text: return "CASH SALE – PARTIAL"
        elif "PAID" in normalized_text and "UNPAID" not in normalized_text: return "CASH SALE - PAID"
        elif "UNPAID" in normalized_text: return "CASH SALE - UNPAID"
    
    return None

def main():
    sales_report_file = resolve_sales_report_file(datetime.datetime.now())

    missing = []
    if not os.path.exists(sales_report_file):
        missing.append(sales_report_file)
    if not os.path.exists(COMMENTS_FILE):
        missing.append(COMMENTS_FILE)

    if missing:
        print("Required files not found:")
        for p in missing:
            print(f"- {p}")
        return

    # 1. Load the comments mapped by order number
    print(f"Loading {COMMENTS_FILE}...")
    df_comments = pd.read_excel(COMMENTS_FILE)
    
    # Build dictionary mappings: filename guessing and strict SO number guessing
    comment_map = {}
    for _, row in df_comments.iterrows():
        guessed_so = str(row.get("Guessed Order Number", "")).strip()
        filename = str(row.get("Filename", ""))
        text_content = str(row.get("Text Box Contents", ""))
        
        status = map_comment_to_status(text_content)
        if status:
            if guessed_so:
                comment_map[guessed_so] = status
            # Also store under filename as backup
            comment_map[filename] = status

    # 2. Process Sales Order Report
    print(f"Opening workbook {sales_report_file}...")
    try:
         wb = openpyxl.load_workbook(sales_report_file)
    except PermissionError:
         print("File is open. Please close Excel.")
         return
         
    # Find today's sheet
    date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    sheets = [s for s in wb.sheetnames if date_pattern.match(s)]
    
    if not sheets:
        print("No valid date sheets found.")
        return
        
    sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
    latest_sheet_name = sheets[-1]
    
    print(f"Updating sheet: {latest_sheet_name}")
    ws = wb[latest_sheet_name]
    
    # Locate columns (Row 2 headers)
    job_col = None
    so_col = None
    pay_col = None
    
    for c_idx in range(1, 20):
        val = ws.cell(row=2, column=c_idx).value
        if val == "Job Number":
            job_col = c_idx
        elif val == "SO Number":
            so_col = c_idx
        elif val == "Payment Status":
            pay_col = c_idx

    if not pay_col:
        print("Could not find 'Payment Status' column.")
        return
        
    updates = 0
    # Process from Row 3 downwards
    for row_idx in range(3, ws.max_row + 1):
        so_val = str(ws.cell(row=row_idx, column=so_col).value).strip() if so_col else ""
        job_val = str(ws.cell(row=row_idx, column=job_col).value).strip() if job_col else ""
        
        match_status = None
        
        # Priority matching against the parsed dictionary
        if job_val and job_val in comment_map:
             match_status = comment_map[job_val]
        elif so_val and so_val in comment_map:
             match_status = comment_map[so_val]
        else:
             # Try substring matching filename just in case
             for key, st in comment_map.items():
                 if so_val and so_val in key:
                     match_status = st
                     break
                 if job_val and job_val != "Pending" and job_val in key:
                     match_status = st
                     break
                     
        if match_status:
             ws.cell(row=row_idx, column=pay_col).value = match_status
             updates += 1

    print(f"Made {updates} updates to Column J.")
    
    # Save the workbook
    print("Saving workbook...")
    wb.save(sales_report_file)
    print("Done!")

if __name__ == "__main__":
    main()
