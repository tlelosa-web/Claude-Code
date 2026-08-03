"""
Item Movement By Category Report Generator
Produces a formatted Excel report with:
  - Tab 1: "Category Summary" — aggregated totals by category
  - Tabs per category: detailed item rows (18 columns, matching Item Movement Summary format)

Columns per category tab (18):
  Item Code | Description | Category | Period (month(s)) | Total QTY Purchased | Supplier | Cost |
  Total QTY Sold | Lead Time (months) | AMU | Min | Re-Order Point (ROP) | Current Stock | On Order | Max |
  Stock Value | Re-Order Qty | Re-Order Value

Re-Order Point (ROP) Formula:
  ROP = ROUNDUP((AMU × Lead Time in months) + Safety Stock (Min), 0)
  This ensures orders are placed before stock runs out during lead time.

IMPORTANT: Includes ALL items from ItemListingReport.csv, even those without
purchase/sale transactions. Items with no movement show 0 for purchased/sold.
"""

import csv
import math
import os
import sys
import tempfile
import shutil
from datetime import datetime
from collections import defaultdict, Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.properties import PageSetupProperties

# ── Paths (relative to project root) ──────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
SOURCE_CSV = os.path.join(PROJECT_ROOT, "2_Source_Data", "ItemMovementReport.csv")
ITEM_LISTING_CSV = os.path.join(PROJECT_ROOT, "2_Source_Data", "ItemListingReport.csv")
PO_BY_ITEM_CSV = os.path.join(PROJECT_ROOT, "2_Source_Data", "OutstandingPOByItemReport.csv")
STOCK_XLSX = os.path.join(PROJECT_ROOT, "2_Source_Data", "Stock Count Update.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "3_Live_Reports")
DEBUG_LOG = os.path.join(PROJECT_ROOT, "5_Archive_and_Debug", "debug_output_utf8.txt")

# ── Formatting constants ──────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def log(msg: str) -> None:
    """Write to both stdout and debug log."""
    try:
        print(msg)
    except UnicodeEncodeError:
        print(msg.encode('ascii', 'replace').decode('ascii'))
    with open(DEBUG_LOG, "a", encoding="utf-8") as f:
        f.write(msg + "\n")


def safe_copy_for_reading(src: str) -> str:
    """Shadow-copy a file so we never read a potentially locked live file."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(src)[1])
    tmp.close()
    shutil.copy2(src, tmp.name)
    return tmp.name


def parse_date(s: str) -> datetime | None:
    """Parse DD/MM/YYYY date strings."""
    s = s.strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def round_amu_half(val: float) -> float:
    """Round AMU up to the next 0.5 increment, 1 decimal place, minimum 1.0."""
    if val <= 0:
        return 1.0
    rounded = math.ceil(val * 2) / 2
    rounded = round(rounded, 1)
    return max(rounded, 1.0)


def load_current_stock_map(filepath: str) -> dict[str, float]:
    if not os.path.isfile(filepath):
        return {}

    tmp_path = safe_copy_for_reading(filepath)
    try:
        wb = load_workbook(tmp_path, data_only=True)
        sheet = "ImportStockCount_TOT" if "ImportStockCount_TOT" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]

        header_row_idx = 0
        item_col_idx = -1
        stock_col_idx = -1

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), 1):
            if not row:
                continue
            row_norm = [str(v).strip().lower() if v is not None else "" for v in row]
            if "item" in row_norm and "current stock" in row_norm:
                header_row_idx = i
                item_col_idx = row_norm.index("item")
                stock_col_idx = row_norm.index("current stock")
                break

        if header_row_idx == 0 or item_col_idx < 0 or stock_col_idx < 0:
            return {}

        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or item_col_idx >= len(row):
                continue
            code = row[item_col_idx]
            if code is None:
                continue
            code = str(code).strip()
            if not code:
                continue
            raw_stock = row[stock_col_idx] if stock_col_idx < len(row) else 0
            try:
                stock = float(raw_stock) if raw_stock is not None else 0.0
            except (TypeError, ValueError):
                stock = 0.0
            result[code] = stock

        return result
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def load_current_stock_raw_map(filepath: str) -> dict[str, float | None]:
    if not os.path.isfile(filepath):
        return {}

    tmp_path = safe_copy_for_reading(filepath)
    try:
        wb = load_workbook(tmp_path, data_only=True)
        sheet = "ImportStockCount_TOT" if "ImportStockCount_TOT" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]

        header_row_idx = 0
        item_col_idx = -1
        stock_col_idx = -1

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), 1):
            if not row:
                continue
            row_norm = [str(v).strip().lower() if v is not None else "" for v in row]
            if "item" in row_norm and "current stock" in row_norm:
                header_row_idx = i
                item_col_idx = row_norm.index("item")
                stock_col_idx = row_norm.index("current stock")
                break

        if header_row_idx == 0 or item_col_idx < 0 or stock_col_idx < 0:
            return {}

        result: dict[str, float | None] = {}
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or item_col_idx >= len(row):
                continue
            code = row[item_col_idx]
            if code is None:
                continue
            code = str(code).strip()
            if not code:
                continue

            raw_stock = row[stock_col_idx] if stock_col_idx < len(row) else None
            if raw_stock is None or (isinstance(raw_stock, str) and not raw_stock.strip()):
                result[code] = None
                continue
            try:
                result[code] = float(raw_stock)
            except (TypeError, ValueError):
                result[code] = None

        return result
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def update_stock_count_update_file(item_listing_csv: str, stock_xlsx_path: str) -> tuple[bool, int]:
    existing_stock = load_current_stock_raw_map(stock_xlsx_path)

    tmp_path = safe_copy_for_reading(item_listing_csv)
    try:
        items: list[tuple[str, str, str, float]] = []
        with open(tmp_path, "r", encoding="utf-8-sig", newline="") as f:
            first_line = f.readline().strip()
            if not first_line.startswith("sep="):
                f.seek(0)

            reader = csv.DictReader(f)
            for row in reader:
                code = (row.get("Code") or "").strip()
                if not code:
                    continue
                desc = (row.get("Description") or "").strip()
                cat = (row.get("Category") or "").strip() or "Uncategorized"
                raw_cost = (row.get("Avg. Cost") or "").strip().replace(",", "").replace("R", "").strip()
                try:
                    cost = float(raw_cost) if raw_cost else 0.0
                except (TypeError, ValueError):
                    cost = 0.0
                items.append((code, desc, cat, cost))
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    items.sort(key=lambda t: ((t[2] or "").strip().lower() == "uncategorized", (t[2] or "").lower(), t[0].lower()))

    tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_out.close()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "ImportStockCount_TOT"

        headers = ["Item", "Description", "Category", "Cost", "Current Stock", "Stock Value"]

        ws.append(["TOTAL STOCK"] + [None] * (len(headers) - 1))
        ws.append(headers)

        for code, desc, cat, cost in items:
            current = existing_stock.get(code)
            ws.append([code, desc, cat, cost, current, None])

        for row_idx in range(3, ws.max_row + 1):
            ws.cell(row=row_idx, column=6).value = f"=IFERROR(E{row_idx}*D{row_idx},0)"

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=2, column=col)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER

        for r in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for c in r:
                c.border = THIN_BORDER
                if c.column in (4, 6):
                    c.number_format = "R #,##0.00"

        ws.freeze_panes = "A3"

        col_widths = {1: 18, 2: 45, 3: 22, 4: 12, 5: 14, 6: 14}
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(tmp_out.name)

        os.makedirs(os.path.dirname(stock_xlsx_path), exist_ok=True)
        os.replace(tmp_out.name, stock_xlsx_path)
        return True, len(items)
    except Exception:
        if os.path.exists(tmp_out.name):
            os.unlink(tmp_out.name)
        raise


# ──────────────────────────────────────────────────────────────────────
# PARSERS (reused from item_movement_report.py)
# ──────────────────────────────────────────────────────────────────────

def parse_csv(filepath: str) -> tuple[list[dict], dict[str, dict]]:
    """Parse the raw ItemMovementReport CSV into item records."""
    items: list[dict] = []
    current_item_code: str = ""
    current_item_desc_raw: str = ""
    item_date_ranges: dict[str, dict] = {}

    with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)
        try:
            next(reader)
        except StopIteration:
            return items, item_date_ranges

        for _ in range(3):
            try:
                next(reader)
            except StopIteration:
                break

        for row in reader:
            if not row or not row[0].strip():
                continue

            first = row[0].strip()

            if first.lower().startswith("total for item:"):
                current_item_code = ""
                current_item_desc_raw = ""
                continue

            dt = parse_date(first)
            if dt and len(row) >= 5:
                ref = row[1].strip()
                desc = row[2].strip()
                cust = row[3].strip()
                try:
                    qty = float(row[4].strip())
                except ValueError:
                    qty = 0.0

                is_purchase = desc == "Supplier Invoice"
                is_sale = desc == "Tax Invoice"

                if current_item_code:
                    if current_item_code not in item_date_ranges:
                        item_date_ranges[current_item_code] = {"min_date": dt, "max_date": dt}
                    else:
                        dr = item_date_ranges[current_item_code]
                        if dt < dr["min_date"]:
                            dr["min_date"] = dt
                        if dt > dr["max_date"]:
                            dr["max_date"] = dt

                items.append({
                    "item_code": current_item_code,
                    "item_desc_raw": current_item_desc_raw,
                    "date": dt,
                    "ref": ref,
                    "desc": desc,
                    "customer_supplier": cust,
                    "qty": qty,
                    "is_purchase": is_purchase,
                    "is_sale": is_sale,
                })
            else:
                current_item_code = first
                current_item_desc_raw = ""
                if " - " in first:
                    parts = first.split(" - ", 1)
                    current_item_code = parts[0].strip()
                    current_item_desc_raw = parts[1].strip()

    return items, item_date_ranges


def load_item_listing_map(filepath: str) -> dict[str, dict]:
    """Load ItemListingReport.csv into a lookup: {Code -> {description, category, avg_cost}}."""
    result: dict[str, dict] = {}
    if not os.path.isfile(filepath):
        return result

    tmp_path = safe_copy_for_reading(filepath)
    try:
        with open(tmp_path, "r", encoding="utf-8-sig") as f:
            first_line = f.readline().strip()
            if not first_line.startswith("sep="):
                f.seek(0)

            reader = csv.DictReader(f)
            for row in reader:
                code = row.get("Code", "").strip()
                if code:
                    raw_cost = row.get("Avg. Cost", "").strip().replace(",", "")
                    try:
                        avg_cost = float(raw_cost)
                    except (ValueError, TypeError):
                        avg_cost = 0.0
                    result[code] = {
                        "description": row.get("Description", "").strip() or "",
                        "category": row.get("Category", "").strip() or "",
                        "avg_cost": avg_cost,
                    }
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
    return result


def parse_po_by_item(filepath: str) -> dict[str, dict]:
    """Parse OutstandingPOByItemReport.csv and extract per-item supplier/lead time/pending qty."""
    if not os.path.isfile(filepath):
        return {}

    tmp_path = safe_copy_for_reading(filepath)
    supplier_counts: dict[str, Counter] = defaultdict(Counter)
    lead_times: dict[str, list[float]] = defaultdict(list)
    pending_qty: dict[str, float] = defaultdict(float)

    try:
        with open(tmp_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            try:
                next(reader)
            except StopIteration:
                return {}
            for _ in range(3):
                try:
                    next(reader)
                except StopIteration:
                    break

            current_item_code = ""

            for row in reader:
                if not row or not row[0].strip():
                    continue

                first = row[0].strip()

                if first.lower().startswith("total for item:"):
                    current_item_code = ""
                    continue

                dt = parse_date(first)
                if dt and len(row) >= 6:
                    supplier = row[3].strip()
                    delivery_date = parse_date(row[4].strip())
                    status = row[5].strip()
                    try:
                        po_qty = float(row[6].strip())
                    except (ValueError, IndexError):
                        po_qty = 0.0

                    if current_item_code and " - " in current_item_code:
                        code_part = current_item_code.split(" - ", 1)[0].strip()
                    else:
                        code_part = current_item_code

                    if supplier:
                        supplier_counts[code_part][supplier] += 1

                    if delivery_date:
                        days_diff = (delivery_date - dt).days
                        weeks = days_diff / 7.0
                        if weeks >= 0:
                            lead_times[code_part].append(weeks)

                    if status == "Pending":
                        pending_qty[code_part] += po_qty
                else:
                    current_item_code = first

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    all_codes = set(supplier_counts.keys()) | set(lead_times.keys())
    result: dict[str, dict] = {}
    for code in all_codes:
        top_supplier = ""
        if code in supplier_counts and supplier_counts[code]:
            top_supplier = supplier_counts[code].most_common(1)[0][0]

        avg_lt = 0.0
        if code in lead_times and lead_times[code]:
            avg_lt = round(sum(lead_times[code]) / len(lead_times[code]), 1)

        pqty = round(pending_qty.get(code, 0.0), 2)

        result[code] = {
            "top_supplier": top_supplier,
            "avg_lead_time_weeks": avg_lt,
            "pending_qty": pqty,
        }

    return result


# ──────────────────────────────────────────────────────────────────────
# AGGREGATION
# ──────────────────────────────────────────────────────────────────────

def aggregate_items(
    items: list[dict],
    item_date_ranges: dict[str, dict],
    item_listing_map: dict[str, dict],
    po_data: dict[str, dict],
    current_stock_map: dict[str, float],
) -> list[dict]:
    """
    Aggregate transactions per item into a rich result dict.
    Includes ALL items from ItemListingReport, even those without transactions.
    """
    # First pass: aggregate transaction data
    agg: dict[str, dict] = defaultdict(lambda: {
        "total_purchased": 0.0,
        "total_sold": 0.0,
        "purchase_txns": 0,
        "sale_txns": 0,
    })

    for rec in items:
        key = rec["item_code"]
        if rec["is_purchase"]:
            agg[key]["total_purchased"] += abs(rec["qty"])
            agg[key]["purchase_txns"] += 1
        if rec["is_sale"]:
            agg[key]["total_sold"] += abs(rec["qty"])
            agg[key]["sale_txns"] += 1

    # Second pass: include ALL items from ItemListingReport
    results = []
    all_item_codes = set(agg.keys()) | set(item_listing_map.keys())
    
    for item_code in sorted(all_item_codes):
        d = agg.get(item_code, {
            "total_purchased": 0.0,
            "total_sold": 0.0,
            "purchase_txns": 0,
            "sale_txns": 0,
        })
        purchased = d["total_purchased"]
        sold = d["total_sold"]

        dr = item_date_ranges.get(item_code)
        if dr and dr["min_date"] and dr["max_date"]:
            delta_years = (dr["max_date"].year - dr["min_date"].year) * 12 + (dr["max_date"].month - dr["min_date"].month)
            period_months = max(delta_years, 1)
        else:
            period_months = 1

        max_qty = max(purchased, sold)
        amu_raw = max_qty / period_months if period_months > 0 else 0.0
        amu = round_amu_half(amu_raw)

        listing_info = item_listing_map.get(item_code, {})
        description = listing_info.get("description") or ""
        category = listing_info.get("category") or "Uncategorized"
        avg_cost = listing_info.get("avg_cost", 0.0)
        if not description:
            raw_descs = [rec["item_desc_raw"] for rec in items if rec["item_code"] == item_code and rec["item_desc_raw"]]
            description = raw_descs[0] if raw_descs else ""

        po_info = po_data.get(item_code, {})
        top_supplier = po_info.get("top_supplier", "")
        lead_time_weeks = po_info.get("avg_lead_time_weeks", 0.0)
        lead_time_months = round(lead_time_weeks / 4.33, 1) if lead_time_weeks > 0 else 0.0
        lead_time_months = max(lead_time_months, 0.5)

        lm = lead_time_months if lead_time_months > 0 else 1
        min_qty = round(amu * lm)
        max_qty_min = min_qty + round(amu * 2)
        if min_qty <= 0:
            min_qty = 1
        if max_qty_min <= min_qty:
            max_qty_min = min_qty * 2

        current_stock = round(current_stock_map.get(item_code, 0.0), 0)
        on_order = round(po_info.get("pending_qty", 0.0), 0)
        if on_order < 0:
            on_order = 0

        results.append({
            "item_code": item_code,
            "description": description,
            "category": category,
            "period_months": period_months,
            "total_purchased": purchased,
            "total_sold": sold,
            "supplier": top_supplier,
            "avg_cost": avg_cost,
            "lead_time_months": lead_time_months,
            "amu": amu,
            "min": min_qty,
            "max": max_qty_min,
            "current_stock": current_stock,
            "on_order": on_order,
        })

    return results


# ──────────────────────────────────────────────────────────────────────
# EXCEL REPORT — CATEGORY-BASED
# ──────────────────────────────────────────────────────────────────────

def sanitize_sheet_name(name: str) -> str:
    """Make sheet name Excel-safe (max 31 chars, no special chars)."""
    invalid = ['\\', '/', '?', '*', '[', ']', ':']
    for ch in invalid:
        name = name.replace(ch, "")
    name = name.strip()[:31]
    if not name:
        name = "Sheet"
    return name


def write_category_sheet(ws, results: list[dict]) -> None:
    headers = [
        "Item Code", "Description", "Category", "Period (month(s))",
        "Total QTY Purchased", "Supplier", "Cost", "Total QTY Sold",
        "Lead Time (months)", "AMU", "Min", "Re-Order Point (ROP)",
        "Current Stock", "On Order", "Max", "Stock Value", "Re-Order Qty", "Re-Order Value"
    ]
    ws.append(headers)

    for row_idx, rec in enumerate(results, 2):
        # ROP Formula: (AMU × Lead Time) + Safety Stock (Min)
        # Column J = AMU, Column I = Lead Time, Column K = Min (Safety Stock)
        # ROP = ROUNDUP((J * I) + K, 0)
        row = [
            rec["item_code"], rec["description"], rec.get("category") or "",
            rec["period_months"], rec["total_purchased"], rec["supplier"],
            rec["avg_cost"] if rec["avg_cost"] > 0 else 0,
            rec["total_sold"], rec["lead_time_months"] if rec["lead_time_months"] > 0 else 0.5,
            rec["amu"], rec["min"], 
            f'=ROUNDUP((J{row_idx}*I{row_idx})+K{row_idx},0)',  # Re-Order Point (ROP)
            rec["current_stock"], rec["on_order"],
            rec["max"], f'=IFERROR(M{row_idx}*G{row_idx},0)',
            f'=IF(O{row_idx}-M{row_idx}-N{row_idx}>0,O{row_idx}-M{row_idx}-N{row_idx},0)',
            f'=IFERROR(Q{row_idx}*G{row_idx},0)'
        ]
        ws.append(row)

    totals = ["TOTAL STOCK"] + [None] * 17
    ws.append(totals)

    for row in ws.iter_rows(min_row=2, max_row=len(results)+1):
        row[6].number_format = 'R #,##0.00'  # Cost
        row[15].number_format = 'R #,##0.00'  # Stock Value
        row[17].number_format = 'R #,##0.00'  # Re-Order Value

    table_name = "cat_" + "".join(filter(str.isalnum, ws.title))
    table_name = table_name[:31]
    tab = Table(displayName=table_name, ref=f"A1:R{len(results)+2}")
    tab.totalsRowCount = 1
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    
    tab._initialise_columns()
    for idx, name in enumerate(headers):
        tab.tableColumns[idx].name = name
        if name in ["Stock Value", "Re-Order Value", "Re-Order Qty", "Min", "Re-Order Point (ROP)", "Max", 
                     "Total QTY Purchased", "Total QTY Sold", "Current Stock", "On Order"]:
            tab.tableColumns[idx].totalsRowFunction = "sum"
            
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    ws.page_margins.left = 1.0
    ws.page_margins.right = 1.0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"

def write_category_summary_sheet(ws, results: list[dict]) -> None:
    headers = ["Category", "Total Items", "Total QTY Purchased", "Total QTY Sold"]
    ws.append(headers)
    
    cat_agg = {}
    for rec in results:
        cat = rec.get("category") or "Uncategorized"
        if cat not in cat_agg:
            cat_agg[cat] = {"item_count": 0, "purchased": 0, "sold": 0}
        cat_agg[cat]["item_count"] += 1
        cat_agg[cat]["purchased"] += rec["total_purchased"]
        cat_agg[cat]["sold"] += rec["total_sold"]
        
    for cat in sorted(cat_agg.keys()):
        d = cat_agg[cat]
        ws.append([cat, d["item_count"], d["purchased"], d["sold"]])

    totals = ["TOTAL STOCK"] + [None] * 3
    ws.append(totals)

    table_name = "CategorySummary"
    tab = Table(displayName=table_name, ref=f"A1:D{len(cat_agg)+2}")
    tab.totalsRowCount = 1
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    tab._initialise_columns()
    for idx, name in enumerate(headers):
        tab.tableColumns[idx].name = name
        if idx in range(1, 4):
            tab.tableColumns[idx].totalsRowFunction = "sum"
        
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    
def main():
    log("=" * 60)
    log("Item Movement By Category Report — Starting")
    log("=" * 60)

    try:
        ok, n = update_stock_count_update_file(ITEM_LISTING_CSV, STOCK_XLSX)
        if ok:
            log(f"Stock Count Update refreshed (sorted by Category): {n} items")
    except Exception as e:
        log(f"WARNING: Could not refresh Stock Count Update file: {e}")

    # Parse source data
    log("Parsing ItemMovementReport.csv ...")
    items, item_date_ranges = parse_csv(SOURCE_CSV)
    log(f"  → {len(items)} transactions loaded")

    log("Loading ItemListingReport.csv ...")
    item_listing_map = load_item_listing_map(ITEM_LISTING_CSV)
    log(f"  → {len(item_listing_map)} listings loaded")

    log("Loading OutstandingPOByItemReport.csv ...")
    po_data = parse_po_by_item(PO_BY_ITEM_CSV)
    log(f"  → {len(po_data)} PO records loaded")

    log("Loading Current Stock reference ...")
    current_stock_map = load_current_stock_map(STOCK_XLSX)
    log(f"  → {len(current_stock_map)} stock records loaded")

    # Aggregate
    log("Aggregating item data ...")
    results = aggregate_items(items, item_date_ranges, item_listing_map, po_data, current_stock_map)
    log(f"  → {len(results)} unique items processed")

    # Group by category
    log("Grouping by category ...")
    by_category: dict[str, list[dict]] = defaultdict(list)
    for rec in results:
        cat = rec.get("category") or "Uncategorized"
        by_category[cat].append(rec)

    log(f"  → {len(by_category)} categories found")

    # Build workbook
    wb = Workbook()

    # Tab 1: Category Summary
    log("Writing Category Summary tab ...")
    ws_summary = wb.active
    ws_summary.title = "Category Summary"
    write_category_summary_sheet(ws_summary, results)

    # Tabs per category
    for cat, cat_items in sorted(by_category.items()):
        sheet_name = sanitize_sheet_name(cat)
        log(f"  Writing category: {sheet_name} ({len(cat_items)} items)")
        ws = wb.create_sheet(title=sheet_name)
        write_category_sheet(ws, cat_items)

    # Save
    today = datetime.now().strftime("%d.%m.%Y")
    filename = f"Item Movement By Cat {today}.xlsx"
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    output_path = os.path.join(OUTPUT_DIR, filename)

    log(f"Saving to: {output_path}")
    wb.save(output_path)
    log("✅ Report generation complete.")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        log(f"❌ ERROR: {e}")
        import traceback
        log(traceback.format_exc())
        sys.exit(1)
