"""Inspect formatting properties of ImportStockFinal.xlsx"""
import tempfile, shutil, os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

src = r"2_Source_Data\ImportStockFinal.xlsx"
with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as tmp:
    temp_path = tmp.name
shutil.copy2(src, temp_path)

wb = load_workbook(temp_path)

with open(r"5_Archive_and_Debug\imp_format_inspection.txt", "w", encoding="utf-8") as f:
    for sname in wb.sheetnames[:2]:  # check first 2 sheets
        ws = wb[sname]
        f.write(f"\n{'='*60}\nSheet: {sname}\n{'='*60}\n")
        f.write(f"Dimensions: {ws.dimensions}\n")
        f.write(f"Frozen panes: {ws.freeze_panes}\n\n")
        
        # Column widths
        f.write("Column widths:\n")
        for col_letter, dim in ws.column_dimensions.items():
            f.write(f"  {col_letter}: width={dim.width}\n")
        
        # Row heights
        f.write("\nRow heights:\n")
        for row_num in range(1, min(4, ws.max_row+1)):
            rh = ws.row_dimensions[row_num].height
            f.write(f"  Row {row_num}: height={rh}\n")
        
        # Header formatting (row 1)
        f.write("\nHeader cells (row 1):\n")
        for col in range(1, min(ws.max_column+1, 20)):
            cell = ws.cell(row=1, column=col)
            f.write(f"  {get_column_letter(col)}1: value='{cell.value}' | "
                    f"font={cell.font.name}, size={cell.font.size}, bold={cell.font.bold}, color={cell.font.color} | "
                    f"fill={cell.fill.start_color.rgb if cell.fill.start_color else 'none'} | "
                    f"alignment=h:{cell.alignment.horizontal}, v:{cell.alignment.vertical}, wrap:{cell.alignment.wrap_text} | "
                    f"border_bottom={cell.border.bottom.style}\n")
        
        # Sample data cell (row 2)
        f.write("\nData cells (row 2):\n")
        for col in range(1, min(ws.max_column+1, 20)):
            cell = ws.cell(row=2, column=col)
            f.write(f"  {get_column_letter(col)}2: value='{cell.value}' | "
                    f"font={cell.font.name}, size={cell.font.size}, bold={cell.font.bold} | "
                    f"fill={cell.fill.start_color.rgb if cell.fill.start_color else 'none'} | "
                    f"alignment=h:{cell.alignment.horizontal}, v:{cell.alignment.vertical}, wrap:{cell.alignment.wrap_text} | "
                    f"number_format={cell.number_format}\n")

os.remove(temp_path)
print("Done - check 5_Archive_and_Debug\\imp_format_inspection.txt")
