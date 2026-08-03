import openpyxl

wb = openpyxl.load_workbook("Sales Order Report - 03.2026.xlsx")
sheet_name = "13.03.2026"

if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print("--- 13.03.2026 TOP 5 ROWS ---")
    for r in range(1, 6):
        row_data = [ws.cell(row=r, column=c).value for c in range(1, 13)]
        print(f"Row {r}: {row_data}")
        
    print(f"Frozen Panes: {ws.freeze_panes}")
else:
    print(f"Sheet {sheet_name} not found.")
