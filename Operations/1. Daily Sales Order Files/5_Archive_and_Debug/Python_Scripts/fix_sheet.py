import openpyxl
import re
import datetime

file_path = "Sales Order Report - 03.2026.xlsx"
wb = openpyxl.load_workbook(file_path)
print("Sheet names:", wb.sheetnames)

# Define Date Pattern
date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
sheets = [s for s in wb.sheetnames if date_pattern.match(s)]
sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))

if len(sheets) < 2:
    print("Not enough sheets to find previous day.")
    prev_sheet = None
else:
    # If the last sheet is today, previous is -2
    if sheets[-1] == "23.03.2026":
        prev_sheet = sheets[-2]
    else:
        prev_sheet = sheets[-1]
print("Previous sheet determined as:", prev_sheet)

today_sheet = "23.03.2026"

if today_sheet in wb.sheetnames:
    ws = wb[today_sheet]
    
    # 1. move contents of cells F26 to F25, F25 to I24, F27 to J25
    val_f26 = ws["F26"].value
    val_f25 = ws["F25"].value
    val_f27 = ws["F27"].value
    
    # Do replacements safely
    ws["F25"].value = val_f26
    ws["I24"].value = val_f25
    ws["J25"].value = val_f27
    
    ws["F26"].value = None
    ws["F27"].value = None
    
    print("Moved cells successfully.")
    
    # 2. Change Column L header
    ws["L1"].value = "Planned Delivery Date"
    
    # Apply formulas for rows
    max_row = 1
    for r in range(2, ws.max_row + 1):
        if ws[f"D{r}"].value is not None:
             max_row = max(max_row, r)

             # Formula for Column L
             ws[f"L{r}"].value = f"=B{r}"
             
             # Formula for Column I (Status)
             if prev_sheet:
                 # IFNA(VLOOKUP(D2, '20.03.2026'!D:I, 6, FALSE), "Pending")
                 ws[f"I{r}"].value = f"=IFNA(VLOOKUP(D{r}, '{prev_sheet}'!D:I, 6, FALSE), \"Pending\")"

    wb.save(file_path)
    print("Saved formulas and header.")
else:
    print(f"Sheet {today_sheet} not found.")
