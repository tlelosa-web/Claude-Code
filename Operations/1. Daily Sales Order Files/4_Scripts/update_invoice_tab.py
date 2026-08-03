"""
update_invoice_tab.py
---------------------
Builds (or replaces) the "Monthly INV - MM.YYYY" sheet inside the live
Ops Sales Order Report workbook.

Value sources
-------------
C3  Cash Sale Invoiced (current month)
        CustomerInvoicesReport.csv – sum of Exclusive where
        Invoice Date month == current month AND Due Date month == current month

C4  Account Invoiced (current month)
        CustomerInvoicesReport.csv – sum of Exclusive where
        Invoice Date month == current month AND Due Date month != current month

C9  Cash Sale Invoiced (next month)
        CustomerInvoicesReport.csv – sum of Exclusive where
        Invoice Date month == next month AND Due Date month == next month

C10 Account Invoiced (next month)
        CustomerInvoicesReport.csv – sum of Exclusive where
        Invoice Date month == next month AND Due Date month != next month

F3  Cash Sale Invoice OS  )
F4  Account Invoice OS    )  values read dynamically from the BREAKDOWN
F9  Cash Sale Invoice OS  )  section of the most recent daily tab.
F10 Account Invoice OS    )  The BREAKDOWN row shifts daily as orders are
                             added/removed, so row numbers are never hardcoded.

All other cells (C5, C7, E7, C11, C13, E13, F5, F11) are kept as
Excel formulas matching the original InvoiceRPT.xlsx layout.
"""

import csv
import datetime
import os
import re
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

# ---------------------------------------------------------------------------
# Paths (relative to the project root – run from project root)
# ---------------------------------------------------------------------------
INVOICES_CSV    = os.path.join("2_Source_Data", "CustomerInvoicesReport.csv")
REPORTS_DIR     = "3_Live_Reports"
INV_SOURCE_XLSX = os.path.join("2_Source_Data", "InvoiceRPT.xlsx")   # for reference / fallback layout


# ---------------------------------------------------------------------------
# Helper: resolve the live report path for the current month
# ---------------------------------------------------------------------------
def _resolve_report_path(today: datetime.datetime) -> str:
    month_tag = today.strftime("%m.%Y")
    preferred = os.path.join(REPORTS_DIR, f"Ops Sales Order Report - {month_tag}.xlsx")
    legacy    = os.path.join(REPORTS_DIR, f"Sales Order Report - {month_tag}.xlsx")

    if os.path.exists(preferred):
        return preferred
    if os.path.exists(legacy):
        # Attempt rename to preferred name
        try:
            os.rename(legacy, preferred)
            print(f"  Renamed legacy file to {preferred}")
            return preferred
        except Exception:
            return legacy

    # Fall back to the most recently modified matching report
    pattern = re.compile(r"^(?:Ops )?Sales Order Report - \d{2}\.\d{4}\.xlsx$")
    candidates = [
        os.path.join(REPORTS_DIR, f)
        for f in os.listdir(REPORTS_DIR)
        if pattern.match(f)
    ]
    if candidates:
        candidates.sort(key=os.path.getmtime, reverse=True)
        print(f"  Warning: current-month report not found; using {candidates[0]}")
        return candidates[0]

    return preferred   # will trigger a clear error later


# ---------------------------------------------------------------------------
# Helper: find the most recent date-tab name in the workbook
# ---------------------------------------------------------------------------
def _latest_date_tab(wb: openpyxl.Workbook):
    date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    date_tabs = [s for s in wb.sheetnames if date_pattern.match(s)]
    if not date_tabs:
        return None
    date_tabs.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
    return date_tabs[-1]


# ---------------------------------------------------------------------------
# Helper: find the BREAKDOWN section rows in a daily tab
#
# Returns a dict with keys:
#   breakdown_row    – Row number where "BREAKDOWN" header appears
#   cash_cur_cell    – Cell reference for Cash Sale Current Month (column G)
#   acct_cur_cell    – Cell reference for Account Current Month (column G)
#   cash_nxt_cell    – Cell reference for Cash Sale Next Month (column J)
#   acct_nxt_cell    – Cell reference for Account Next Month (column J)
# Returns None if BREAKDOWN section cannot be found.
# ---------------------------------------------------------------------------
def _find_breakdown_section(ws):
    # Search for "BREAKDOWN" text in column F (column 6)
    for r in range(1, ws.max_row + 1):
        cell_val = ws.cell(r, 6).value  # Column F
        if cell_val and str(cell_val).strip().upper() == "BREAKDOWN":
            breakdown_row = r
            # Based on update_sales_report.py structure:
            # Row breakdown_row     = "BREAKDOWN" header
            # Row breakdown_row + 1 = Cash Sale labels (G=current month, J=next month)
            # Row breakdown_row + 2 = Account labels (G=current month, J=next month)
            return {
                "breakdown_row": breakdown_row,
                "cash_cur_cell": f"G{breakdown_row + 1}",  # Cash Sale, current month
                "acct_cur_cell": f"G{breakdown_row + 2}",  # Account, current month
                "cash_nxt_cell": f"J{breakdown_row + 1}",  # Cash Sale, next month
                "acct_nxt_cell": f"J{breakdown_row + 2}",  # Account, next month
            }
    return None


# ---------------------------------------------------------------------------
# Helper: compute OS values by directly scanning data rows in a daily tab.
#
# This replicates the SUMIFS formulas written by update_sales_report.py:
#   - Col G  (7)  = Total (R value)
#   - Col I  (9)  = Status  – rows where Status == "Pending" are excluded
#   - Col J  (10) = Payment Status – contains "CASH SALE" or "ACCOUNT"
#   - Col L  (12) = Planned Delivery Date – used for month-bucket filtering
#
# We scan directly because openpyxl loaded via data_only=True returns None
# for formula cells that were saved by Python (no Excel recalc cache).
#
# Returns a dict with keys:
#   cash_cur, acct_cur  – Current-month OS totals
#   cash_nxt, acct_nxt  – Next-month OS totals
# All values default to 0.0 if the section cannot be found.
# ---------------------------------------------------------------------------
def _compute_os_from_data(
    ws,
    cur_m: int, cur_y: int,
    nxt_m: int, nxt_y: int,
) -> dict:
    result = {"cash_cur": 0.0, "acct_cur": 0.0, "cash_nxt": 0.0, "acct_nxt": 0.0}

    # Find the header row (contains "SO Number" or "Status") – normally row 2.
    # Data rows start immediately after.
    data_start = None
    for r in range(1, min(6, ws.max_row + 1)):
        row_vals = [str(ws.cell(r, c).value or "").strip() for c in range(1, 13)]
        if "Status" in row_vals and "Total" in row_vals:
            data_start = r + 1
            break
    if data_start is None:
        data_start = 3  # safe default matching update_sales_report.py

    # Scan every data row until we hit the summary/breakdown section
    # (which starts with a non-SO value in the "Date" / col-A area).
    for r in range(data_start, ws.max_row + 1):
        # Stop when we reach the summary block (col A or col D is no longer a date/SO)
        col_d = ws.cell(r, 4).value  # SO Number column
        if col_d is None:
            continue
        col_d_str = str(col_d).strip()
        if not col_d_str.upper().startswith("SO"):
            continue  # skip non-data rows silently

        status       = str(ws.cell(r, 9).value  or "").strip()   # col I
        pay_status   = str(ws.cell(r, 10).value or "").strip()   # col J
        total_val    = ws.cell(r, 7).value                        # col G
        del_date_val = ws.cell(r, 12).value                       # col L

        # Exclude Pending rows (matches SUMIFS <> "Pending" condition)
        if status.upper() == "PENDING":
            continue

        # Parse total
        try:
            total = float(total_val) if total_val is not None else 0.0
        except (ValueError, TypeError):
            continue

        # Parse delivery date
        if isinstance(del_date_val, datetime.datetime):
            del_dt = del_date_val
        elif isinstance(del_date_val, str) and del_date_val.strip():
            try:
                del_dt = datetime.datetime.strptime(del_date_val.strip(), "%d/%m/%Y")
            except ValueError:
                continue
        else:
            continue

        del_m, del_y = del_dt.month, del_dt.year

        pay_upper = pay_status.upper()
        is_cash    = "CASH SALE" in pay_upper
        is_account = "ACCOUNT"   in pay_upper

        if del_m == cur_m and del_y == cur_y:
            if is_cash:    result["cash_cur"] += total
            if is_account: result["acct_cur"] += total
        elif del_m == nxt_m and del_y == nxt_y:
            if is_cash:    result["cash_nxt"] += total
            if is_account: result["acct_nxt"] += total

    return result


# ---------------------------------------------------------------------------
# Helper: sum Exclusive from the invoices CSV
# ---------------------------------------------------------------------------
def _sum_invoices(
    rows: list[dict],
    inv_month: int,
    inv_year: int,
    due_month: int,
    due_year: int,
    due_match: bool,          # True = Due Date must match due_month/year
                               # False = Due Date must NOT match
) -> float:
    total = 0.0
    for r in rows:
        try:
            inv_dt = datetime.datetime.strptime(r["Date"].strip(), "%d/%m/%Y")
            due_dt = datetime.datetime.strptime(r["Due Date"].strip(), "%d/%m/%Y")
        except ValueError:
            continue

        excl_str = r.get("Exclusive", "").strip()
        excl = float(excl_str) if excl_str else 0.0

        if inv_dt.month != inv_month or inv_dt.year != inv_year:
            continue

        due_matches = (due_dt.month == due_month and due_dt.year == due_year)
        if due_match == due_matches:
            total += excl

    return total


# ---------------------------------------------------------------------------
# Helper: apply a consistent thin border to a cell
# ---------------------------------------------------------------------------
_THIN = Side(style="thin")
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def _style_label(cell, bold: bool = False, size: int = 11):
    cell.font = Font(name="Calibri", bold=bold, size=size)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def _style_value(cell, bold: bool = False, size: int = 11):
    cell.font = Font(name="Calibri", bold=bold, size=size)
    cell.number_format = "#,##0.00"
    cell.alignment = Alignment(horizontal="right", vertical="center")


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------
def build_invoice_tab():
    today = datetime.datetime.now()

    # --- Current month / next month ---
    cur_m, cur_y = today.month, today.year
    nxt_m = (cur_m % 12) + 1
    nxt_y = cur_y if cur_m < 12 else cur_y + 1

    cur_month_name = today.strftime("%B").upper()   # e.g. "JUNE"
    nxt_month_dt   = datetime.datetime(nxt_y, nxt_m, 1)
    nxt_month_name = nxt_month_dt.strftime("%B").upper()  # e.g. "JULY"

    month_tag  = today.strftime("%m.%Y")
    sheet_name = f"Monthly INV - {month_tag}"

    # --- Resolve report path ---
    report_path = _resolve_report_path(today)
    if not os.path.exists(report_path):
        print(f"ERROR: Live report not found at {report_path}")
        return

    # --- Read invoices CSV ---
    if not os.path.exists(INVOICES_CSV):
        print(f"ERROR: {INVOICES_CSV} not found. Cannot build invoice tab.")
        return

    csv_text = open(INVOICES_CSV, "r", encoding="utf-8-sig").read()
    lines = csv_text.splitlines()
    if lines and lines[0].startswith("sep="):
        lines = lines[1:]

    reader = csv.DictReader(lines)
    rows   = list(reader)

    # --- Calculate cell values ---
    c3  = _sum_invoices(rows, cur_m, cur_y, cur_m, cur_y, due_match=True)   # Cash Sale, current month
    c4  = _sum_invoices(rows, cur_m, cur_y, cur_m, cur_y, due_match=False)  # Account,   current month
    c9  = _sum_invoices(rows, nxt_m, nxt_y, nxt_m, nxt_y, due_match=True)  # Cash Sale, next month
    c10 = _sum_invoices(rows, nxt_m, nxt_y, nxt_m, nxt_y, due_match=False) # Account,   next month

    print(f"  Invoice totals from {INVOICES_CSV}:")
    print(f"    C3  Cash Sale Invoiced ({cur_month_name}):  {c3:,.2f}")
    print(f"    C4  Account Invoiced   ({cur_month_name}):  {c4:,.2f}")
    print(f"    C9  Cash Sale Invoiced ({nxt_month_name}): {c9:,.2f}")
    print(f"    C10 Account Invoiced   ({nxt_month_name}): {c10:,.2f}")

    # --- Open the live report ---
    print(f"  Opening {report_path} ...")
    try:
        wb = openpyxl.load_workbook(report_path)
    except PermissionError:
        print("ERROR: File is open in Excel. Please close it and rerun.")
        return

    # --- Find most recent daily tab and locate its BREAKDOWN section ---
    latest_tab = _latest_date_tab(wb)
    if not latest_tab:
        print("ERROR: No date tabs found in the live report. Cannot read OS values.")
        return
    
    breakdown_info = _find_breakdown_section(wb[latest_tab])
    if not breakdown_info:
        print("WARNING: BREAKDOWN section not found in daily tab '" + latest_tab + "'.")
        print("  Falling back to computing OS values from data rows...")
        breakdown = _compute_os_from_data(wb[latest_tab], cur_m, cur_y, nxt_m, nxt_y)
        f3  = breakdown["cash_cur"]   # Cash Sale OS, current month
        f4  = breakdown["acct_cur"]   # Account OS,   current month
        f9  = breakdown["cash_nxt"]   # Cash Sale OS, next month
        f10 = breakdown["acct_nxt"]   # Account OS,   next month
        use_formulas = False
    else:
        print("  Found BREAKDOWN section in daily tab '" + latest_tab + "' at row " + str(breakdown_info["breakdown_row"]))
        # Use formula references to the BREAKDOWN cells
        f3  = f"='{latest_tab}'!{breakdown_info['cash_cur_cell']}"
        f4  = f"='{latest_tab}'!{breakdown_info['acct_cur_cell']}"
        f9  = f"='{latest_tab}'!{breakdown_info['cash_nxt_cell']}"
        f10 = f"='{latest_tab}'!{breakdown_info['acct_nxt_cell']}"
        use_formulas = True

    print("    F3  Cash Sale OS (cur month): " + str(f3))
    print("    F4  Account OS   (cur month): " + str(f4))
    print("    F9  Cash Sale OS (nxt month): " + str(f9))
    print("    F10 Account OS   (nxt month): " + str(f10))

    # --- Remove existing invoice tab if present ---
    if sheet_name in wb.sheetnames:
        del wb[sheet_name]
        print(f"  Removed existing sheet '{sheet_name}'.")

    # --- Create the new sheet (insert after the last date tab) ---
    ws = wb.create_sheet(title=sheet_name)

    # -----------------------------------------------------------------------
    # Layout - mirrors InvoiceRPT.xlsx exactly
    # -----------------------------------------------------------------------

    # Row 1 – Title
    ws["A1"] = "Invoicing"
    ws["A1"].font = Font(name="Calibri", bold=True, size=14)

    # Row 3 – Current month header + Cash Sale row
    ws["A3"] = "Current Month:"
    _style_label(ws["A3"], bold=True)

    ws["B3"] = "Cash Sale Invoiced"
    _style_label(ws["B3"])

    ws["C3"] = c3          # <-- from CustomerInvoicesReport.csv (due in current month)
    _style_value(ws["C3"])

    ws["E3"] = "Cash Sale Invoice OS"
    _style_label(ws["E3"])

    ws["F3"] = f3   # Formula linking to BREAKDOWN section of '{latest_tab}'
    if use_formulas:
        ws["F3"].number_format = '#,##0.00'
    else:
        _style_value(ws["F3"])

    # Row 4 – Account row
    ws["B4"] = "Account Invoiced"
    _style_label(ws["B4"])

    ws["C4"] = c4          # <-- from CustomerInvoicesReport.csv (due in future month)
    _style_value(ws["C4"])

    ws["E4"] = "Account Invoice OS"
    _style_label(ws["E4"])

    ws["F4"] = f4   # Formula linking to BREAKDOWN section of '{latest_tab}'
    if use_formulas:
        ws["F4"].number_format = '#,##0.00'
    else:
        _style_value(ws["F4"])

    # Row 5 – Totals
    ws["B5"] = "Total Excl VAT."
    _style_label(ws["B5"], bold=True)

    ws["C5"] = "=SUM(C3:C4)"
    _style_value(ws["C5"], bold=True)

    ws["E5"] = "Total Excl VAT."
    _style_label(ws["E5"], bold=True)

    ws["F5"] = "=SUM(F3:F4)"
    _style_value(ws["F5"], bold=True)

    # Row 7 – Grand total
    ws["B7"] = "GRAND TOTAL:"
    _style_label(ws["B7"], bold=True)

    ws["C7"] = "=C5+F5"
    _style_value(ws["C7"], bold=True)

    ws["D7"] = "Excl. VAT"
    _style_label(ws["D7"])

    ws["E7"] = "=C7*1.15"
    _style_value(ws["E7"], bold=True)

    ws["F7"] = "Incl. VAT"
    _style_label(ws["F7"])

    # Row 9 – Next month header + Cash Sale row
    ws["A9"] = f"{nxt_month_name}:"
    _style_label(ws["A9"], bold=True)

    ws["B9"] = "Cash Sale Invoiced"
    _style_label(ws["B9"])

    ws["C9"] = c9          # <-- from CustomerInvoicesReport.csv (next month, due same month)
    _style_value(ws["C9"])

    ws["E9"] = "Cash Sale Invoice OS"
    _style_label(ws["E9"])

    ws["F9"] = f9   # Formula linking to BREAKDOWN section of '{latest_tab}'
    if use_formulas:
        ws["F9"].number_format = '#,##0.00'
    else:
        _style_value(ws["F9"])

    # Row 10 – Account row
    ws["B10"] = "Account Invoiced"
    _style_label(ws["B10"])

    ws["C10"] = c10        # <-- from CustomerInvoicesReport.csv (next month, due later)
    _style_value(ws["C10"])

    ws["E10"] = "Account Invoice OS"
    _style_label(ws["E10"])

    ws["F10"] = f10  # Formula linking to BREAKDOWN section of '{latest_tab}'
    if use_formulas:
        ws["F10"].number_format = '#,##0.00'
    else:
        _style_value(ws["F10"])

    # Row 11 – Totals
    ws["B11"] = "Total Excl."
    _style_label(ws["B11"], bold=True)

    ws["C11"] = "=SUM(C9:C10)"
    _style_value(ws["C11"], bold=True)

    ws["E11"] = "Total Excl."
    _style_label(ws["E11"], bold=True)

    ws["F11"] = "=SUM(F9:F10)"
    _style_value(ws["F11"], bold=True)

    # Row 13 – Grand total (next month)
    ws["B13"] = "GRAND TOTAL:"
    _style_label(ws["B13"], bold=True)

    ws["C13"] = "=C11+F11"
    _style_value(ws["C13"], bold=True)

    ws["D13"] = "Excl. VAT"
    _style_label(ws["D13"])

    ws["E13"] = "=C13*1.15"
    _style_value(ws["E13"], bold=True)

    ws["F13"] = "Incl. VAT"
    _style_label(ws["F13"])

    # -----------------------------------------------------------------------
    # Column widths
    # -----------------------------------------------------------------------
    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 22
    ws.column_dimensions["F"].width = 16

    # -----------------------------------------------------------------------
    # Light separator lines between the two blocks (rows 3-7 and 9-13)
    # -----------------------------------------------------------------------
    header_fill  = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    section_fill = PatternFill(start_color="EBF0FA", end_color="EBF0FA", fill_type="solid")
    total_fill   = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")

    # Title row background
    for col in ["A", "B", "C", "D", "E", "F"]:
        ws[f"{col}1"].fill = header_fill

    # Data rows light fill
    for row in [3, 4, 9, 10]:
        for col in ["B", "C", "E", "F"]:
            ws[f"{col}{row}"].fill = section_fill

    # Total rows slightly darker
    for row in [5, 11]:
        for col in ["B", "C", "E", "F"]:
            ws[f"{col}{row}"].fill = total_fill

    # Grand total rows
    grand_fill = PatternFill(start_color="2E75B6", end_color="2E75B6", fill_type="solid")
    white_font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")
    for row in [7, 13]:
        for col in ["B", "C", "D", "E", "F"]:
            ws[f"{col}{row}"].fill = grand_fill
            ws[f"{col}{row}"].font = white_font

    # -----------------------------------------------------------------------
    # Source annotation (row 15)
    # -----------------------------------------------------------------------
    ws["A15"] = (
        "C3/C4/C9/C10: CustomerInvoicesReport.csv (Due Date month determines Cash/Account)  |  "
        "F3/F4/F9/F10: Formulas linking to BREAKDOWN section in daily tab '" + latest_tab + "'  |  "
        "Generated: " + today.strftime("%Y-%m-%d %H:%M")
    )
    ws["A15"].font = Font(name="Calibri", italic=True, size=9, color="808080")

    # -----------------------------------------------------------------------
    # Save
    # -----------------------------------------------------------------------
    print(f"  Saving workbook ...")
    try:
        wb.save(report_path)
        print(f"  Sheet '{sheet_name}' written successfully to {report_path}")
    except PermissionError:
        print("ERROR: Could not save – file may be open in Excel. Please close it and rerun.")
        return


def main():
    print("=" * 60)
    print("UPDATE INVOICE TAB")
    print("=" * 60)
    build_invoice_tab()
    print("=" * 60)


if __name__ == "__main__":
    main()
