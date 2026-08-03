import openpyxl, re

wb = openpyxl.load_workbook(
    r'3_Live_Reports\Ops Sales Order Report - 06.2026.xlsx',
    data_only=True
)
date_pattern = re.compile(r'^\d{2}\.\d{2}\.\d{4}$')
date_tabs = [s for s in wb.sheetnames if date_pattern.match(s)]

for tab in date_tabs:
    ws = wb[tab]
    breakdown_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(r, 6).value   # col F
        if val == 'BREAKDOWN':
            breakdown_row = r
            break
    if breakdown_row:
        print(tab + ': BREAKDOWN at row ' + str(breakdown_row))
        for r in range(breakdown_row, breakdown_row + 4):
            print('  Row ' + str(r) + ': F=' + repr(ws.cell(r, 6).value)
                  + '  G=' + repr(ws.cell(r, 7).value)
                  + '  I=' + repr(ws.cell(r, 9).value)
                  + '  J=' + repr(ws.cell(r, 10).value))
    else:
        print(tab + ': NO BREAKDOWN section found')
