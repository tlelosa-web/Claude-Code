from __future__ import annotations
from pathlib import Path
import shutil
import time
import uuid

import openpyxl

from motor_fla_lookup import lookup_fla_safe
from pathing import resolve_paths


def _norm(s: str) -> str:
    return "".join(str(s).strip().lower().split())

def _speed_to_poles(speed: str) -> str:
    try:
        v = int(float(str(speed).strip()))
    except Exception:
        return ""
    return {2880: "2", 1440: "4", 960: "6", 720: "8"}.get(v, "")

def _speed_code_to_poles(speed: str) -> str:
    if not isinstance(speed, str):
        return ""
    s = speed.strip().upper()
    if s.endswith("P"):
        try:
            p = int(s[:-1])
            return str(p) if p in (2, 4, 6, 8) else ""
        except Exception:
            return ""
    return ""

def _speed_code_to_rpm(speed: str) -> str:
    if not isinstance(speed, str):
        return ""
    return {
        "2P": "2880",
        "4P": "1440",
        "6P": "960",
        "8P": "720",
    }.get(speed.strip().upper(), "")


def _find_labeled_value(ws, label: str):
    target = str(label).strip().lower()
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            v = ws.cell(r, c).value
            if not isinstance(v, str):
                continue
            sv = v.strip().lower()
            if sv == target or sv.startswith(target):
                return ws.cell(r, c + 1).value
    return None

def _fmt_month_year(v) -> str:
    try:
        import datetime
        if isinstance(v, (datetime.date, datetime.datetime)):
            d = v.date() if isinstance(v, datetime.datetime) else v
            return d.strftime("%b.%Y").upper()
    except Exception:
        pass
    return "" if v is None else str(v).strip()

def resolve_default_nameplate_excel_path() -> Path:
    resolved = resolve_paths(Path(__file__).resolve().parent)
    p = resolved.get("nameplate_excel_path")
    if isinstance(p, Path) and p.exists():
        return p
    msg = (resolved.get("errors") or {}).get("NAMEPLATE_EXCEL_PATH") or "Nameplate Excel file not found."
    raise FileNotFoundError(msg)

def _resolve_motor_pdf_path() -> Path | None:
    resolved = resolve_paths(Path(__file__).resolve().parent)
    p = resolved.get("motor_pdf_path")
    if isinstance(p, Path) and p.exists():
        return p
    return None


def _read_block_by_labels(ws, label_col: str, value_col: str) -> dict:
    out = {}
    for r in range(1, 250):
        lab = ws[f"{label_col}{r}"].value
        if lab is None:
            continue
        key = _norm(lab)
        val = ws[f"{value_col}{r}"].value
        out[key] = "" if val is None else str(val).strip()
    return out


def _to_float(s: str):
    try:
        return float(str(s).strip())
    except Exception:
        return None


def _to_int(s: str):
    try:
        return int(float(str(s).strip()))
    except Exception:
        return None


def _safe_copy_excel(src: Path) -> Path:
    tmp = src.parent / f"_tmp_excel_{uuid.uuid4().hex}.xlsx"
    last_err = None
    for _ in range(3):
        try:
            shutil.copy2(src, tmp)
            return tmp
        except Exception as e:
            last_err = e
            time.sleep(0.2)
    raise PermissionError(f"Cannot read Excel file: {last_err}")


def read_nameplate_from_excel(xlsx_path: str) -> dict:
    src = Path(xlsx_path)
    if not src.exists():
        raise FileNotFoundError(f"Excel not found: {src}")

    tmp = _safe_copy_excel(src)

    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
        if "Table 1" in wb.sheetnames:
            ws = wb["Table 1"]

            left = _read_block_by_labels(ws, "G", "H")
            right = _read_block_by_labels(ws, "J", "K")

            def pick(d: dict, *labels: str, default: str = "") -> str:
                for lab in labels:
                    k = _norm(lab)
                    if k in d and d[k] != "":
                        return d[k]
                for lab in labels:
                    k = _norm(lab)
                    if k in d:
                        return d[k]
                return default

            series = pick(left, "Series")
            class_pitch = pick(left, "Class/Pitch", "Class Pitch", "Pitch", "Class")
            motor_kw_str = pick(left, "Motor")
            voltage_str = pick(left, "Voltage")
            pole_str = pick(left, "Pole", "Poles")
            fla_str = pick(left, "F.L.A", "F.L.A.", "FLA")
            op_temp = pick(left, "Op Temp", default="20")
            serial_no = pick(left, "Serial No")

            size = pick(right, "Size")
            op_speed = pick(right, "Op Speed")
            max_speed = op_speed
            phase = pick(right, "Phase")
            frequency = pick(right, "Frequency", default="50")
            connection = pick(right, "Conn", "Connection")
            date_of_manuf = pick(right, "Date of Manuf", "Date of Manufacture")
            customer_name = ""
            relube_interval = "N/A"
        elif "Info+Data Entry Form" in wb.sheetnames:
            ws = wb["Info+Data Entry Form"]
            series = _find_labeled_value(ws, "Fan Series")
            size = _find_labeled_value(ws, "Fan Size")
            class_pitch = _find_labeled_value(ws, "Class/Pitch")
            motor_kw_str = _find_labeled_value(ws, "Power (kW)") or _find_labeled_value(ws, "Motor")
            voltage_str = _find_labeled_value(ws, "Voltage")
            speed_str = _find_labeled_value(ws, "Speed")
            op_speed = _speed_code_to_rpm(speed_str) or str(speed_str).strip() if speed_str is not None else ""
            pole_str = _speed_code_to_poles(speed_str) or _speed_to_poles(op_speed)
            fla_str = _find_labeled_value(ws, "F.L.A") or _find_labeled_value(ws, "Motor Current")
            op_temp = _find_labeled_value(ws, "Op Temp") or "20"
            serial_no = _find_labeled_value(ws, "Contract No")
            phase = _find_labeled_value(ws, "Phase")
            frequency = _find_labeled_value(ws, "Frequency")
            connection = _find_labeled_value(ws, "Connection") or _find_labeled_value(ws, "Conn") or _find_labeled_value(ws, "Connetion")
            date_of_manuf = _find_labeled_value(ws, "Date")
            customer_name = _find_labeled_value(ws, "Customer Name") or ""
            relube_interval = _find_labeled_value(ws, "Re-Lubrication Int") or "N/A"
            max_speed = op_speed
        else:
            ws = wb["Name Plate"]
            series = _find_labeled_value(ws, "Series")
            size = _find_labeled_value(ws, "Size")
            class_pitch = _find_labeled_value(ws, "Class/Pitch")
            motor_kw_str = _find_labeled_value(ws, "Motor")
            voltage_str = _find_labeled_value(ws, "Voltage")
            fla_str = _find_labeled_value(ws, "F.L.A")
            op_speed = _find_labeled_value(ws, "Op Speed")
            max_speed = op_speed
            phase = _find_labeled_value(ws, "Phase")
            frequency = _find_labeled_value(ws, "Frequency")
            connection = _find_labeled_value(ws, "Conn")
            date_of_manuf = _fmt_month_year(_find_labeled_value(ws, "Date of Manuf"))
            serial_no = _find_labeled_value(ws, "Serial No")
            op_temp = "20"
            pole_str = _speed_to_poles(op_speed)
            relube_interval = "N/A"
            customer_name = ""
            if "Info+Data Entry Form" in wb.sheetnames:
                ws_info = wb["Info+Data Entry Form"]
                customer_name = "" if _find_labeled_value(ws_info, "Customer Name:") is None else str(_find_labeled_value(ws_info, "Customer Name:")).strip()

        # ---- SAFE AUTO FLA ----
        fla_error = None
        motor_kw = _to_float(motor_kw_str)
        voltage = _to_float(voltage_str)
        poles = _to_int(pole_str)
        fla_val = _to_float(fla_str)

        if fla_val is None and motor_kw and voltage and poles:
            pdf_path = _resolve_motor_pdf_path()
            if pdf_path:
                fla_str, fla_error = lookup_fla_safe(motor_kw, poles, voltage, str(pdf_path))
            else:
                fla_error = "Motor performance PDF not found."

        return {
            "customer_name": customer_name,
            "series": series,
            "class_pitch": class_pitch,
            "motor": motor_kw_str,
            "voltage": voltage_str,
            "fla": fla_str,
            "op_temp": op_temp,
            "serial_no": serial_no,
            "size": size,
            "op_speed": op_speed,
            "max_speed": max_speed,
            "phase": phase,
            "frequency": frequency,
            "connection": connection,
            "date_of_manuf": date_of_manuf,
            "relube_interval": relube_interval,
            # debug-only
            "pole": pole_str,
            "fla_error": fla_error,
        }

    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass


def read_test_sheet_from_excel(xlsx_path: str, *, series_hint: str | None = None) -> dict:
    src = Path(xlsx_path)
    if not src.exists():
        raise FileNotFoundError(f"Excel not found: {src}")

    tmp = _safe_copy_excel(src)

    def _cell(ws, r: int, c: int):
        v = ws.cell(r, c).value
        if v is None:
            return ""
        if isinstance(v, str):
            return v.strip()
        return v

    def _join_nonempty(*vals) -> str:
        parts = []
        for v in vals:
            if v is None:
                continue
            s = str(v).strip()
            if s:
                parts.append(s)
        return " ".join(parts).strip()

    try:
        wb = openpyxl.load_workbook(tmp, data_only=True)
        sheet_name = "Test Sheet"
        if series_hint and str(series_hint).strip().upper() == "FUMAX" and "Test Sheet FUMAX" in wb.sheetnames:
            sheet_name = "Test Sheet FUMAX"
        ws = wb[sheet_name]

        make = _join_nonempty(_cell(ws, 5, 20), _cell(ws, 5, 22), _cell(ws, 5, 24))

        return {
            "date": _cell(ws, 1, 21),
            "contract_no": _cell(ws, 2, 4),
            "fan_series": _cell(ws, 2, 10),
            "fan_size": _cell(ws, 2, 16),
            "imp_form": _cell(ws, 2, 22),
            "customer_name": _cell(ws, 4, 4),
            "motor_current": _cell(ws, 5, 4),
            "make": make,
            "motor_voltage": _cell(ws, 6, 4),
            "description": _cell(ws, 6, 20),
            "fan_speed": _cell(ws, 7, 4),
            "motor_serial_number": _cell(ws, 12, 4),
            "blade_pitch_deg": _cell(ws, 12, 5),
            "tacho_clamp_serial_no": _cell(ws, 12, 7),
            "speed_actual": _cell(ws, 12, 9),
            "current_ph1": _cell(ws, 12, 11),
            "current_ph2": _cell(ws, 12, 13),
            "current_ph3": _cell(ws, 12, 15),
            "voltage_ph1": _cell(ws, 12, 17),
            "voltage_ph2": _cell(ws, 12, 19),
            "voltage_ph3": _cell(ws, 12, 21),
            "connection": _cell(ws, 12, 24),
        }
    finally:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
