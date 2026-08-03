import re

def process_file(filepath):
    with open(filepath, 'r', encoding='utf-8') as f:
        content = f.read()
    
    if "TableStyleInfo" not in content:
        content = content.replace("from openpyxl.worksheet.properties import PageSetupProperties", 
        "from openpyxl.worksheet.properties import PageSetupProperties\nfrom openpyxl.worksheet.table import Table, TableStyleInfo")
        
        content = content.replace("from openpyxl.utils import get_column_letter", 
        "from openpyxl.utils import get_column_letter\nfrom openpyxl.worksheet.table import Table, TableStyleInfo")

    # For item_movement_report.py
    if "def _write_sheet(ws" in content:
        # replace the function body
        new_func = '''def _write_sheet(ws, results: list[dict], show_all: bool) -> None:
    if show_all:
        headers = [
            "Item Code", "Description", "Category", "Period (month(s))",
            "Total QTY Purchased", "Supplier", "Cost", "Total QTY Sold",
            "Lead Time (months)", "AMU", "Min", "Current Stock", "On Order",
            "Max", "Stock Value", "Re-Order Qty", "Re-Order Value"
        ]
        ws.title = "Item Movement Summary"
        max_col_letter = "Q"
    else:
        headers = [
            "Item", "Description", "Supplier", "Category", "Cost Price",
            "Safety Stock", "Reorder Point (ROP)", "Maximum Stock (MAX)",
            "Current Stock", "On Order", "Order Quantity", "Stock Value", "Order Value"
        ]
        ws.title = "Order Summary"
        max_col_letter = "M"

    ws.append(headers)

    for rec in results:
        if show_all:
            row = [
                rec["item_code"], rec["description"], rec.get("category") or "",
                rec["period_months"], rec["total_purchased"], rec["supplier"],
                rec["avg_cost"] if rec["avg_cost"] > 0 else 0,
                rec["total_sold"], rec["lead_time_months"] if rec["lead_time_months"] > 0 else 0.5,
                rec["amu"], rec["min"], rec["current_stock"], rec["on_order"],
                rec["max"], f'=IFERROR([@[Current Stock]]*[@Cost],0)',
                f'=IF([@Max]-[@[Current Stock]]-[@[On Order]]>0,[@Max]-[@[Current Stock]]-[@[On Order]],0)',
                f'=IFERROR([@[Re-Order Qty]]*[@Cost],0)'
            ]
            ws.append(row)
        else:
            row = [
                rec["item_code"], rec["description"], rec["supplier"],
                rec.get("category") or "", rec["avg_cost"] if rec["avg_cost"] > 0 else 0,
                rec["min"], 0, # Reorder Point (ROP) defaulting to 0
                rec["max"], rec["current_stock"], rec["on_order"],
                f'=IF([@[Maximum Stock (MAX)]]-[@[Current Stock]]-[@[On Order]]>0,[@[Maximum Stock (MAX)]]-[@[Current Stock]]-[@[On Order]],0)',
                f'=IFERROR([@[Current Stock]]*[@[Cost Price]],0)',
                f'=IFERROR([@[Order Quantity]]*[@[Cost Price]],0)'
            ]
            ws.append(row)

    totals = ["TOTAL STOCK"] + [""] * (len(headers) - 1)
    ws.append(totals)

    # Force format cols before adding table
    for row in ws.iter_rows(min_row=2, max_row=len(results)+1):
        if show_all:
            row[6].number_format = 'R #,##0.00'
            row[14].number_format = 'R #,##0.00'
            row[16].number_format = 'R #,##0.00'
        else:
            row[4].number_format = 'R #,##0.00'
            row[11].number_format = 'R #,##0.00'
            row[12].number_format = 'R #,##0.00'

    table_name = ws.title.replace(" ", "")
    tab = Table(displayName=table_name, ref=f"A1:{max_col_letter}{len(results)+2}")
    tab.totalsRowCount = 1
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    
    tab._initialise_columns()
    for idx, name in enumerate(headers):
        if name in ["Stock Value", "Re-Order Value", "Order Value", "Order Quantity", "Re-Order Qty", "Min", "Max", "Maximum Stock (MAX)", "Safety Stock"]:
            tab.tableColumns[idx].totalsRowFunction = "sum"
            
    ws.add_table(tab)

    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    ws.page_margins.left = 1.0; ws.page_margins.right = 1.0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"

'''
        content = re.sub(r'def _write_sheet\(ws, results: list\[dict\], show_all: bool\) -> None:.*?def build_excel_report', new_func + '\ndef build_excel_report', content, flags=re.DOTALL)
        
    # For item_movement_by_cat.py
    if "def write_category_sheet(ws" in content:
        new_cat = '''def write_category_sheet(ws, results: list[dict]) -> None:
    headers = [
        "Item", "Description", "Supplier", "Category", "Cost Price",
        "Safety Stock", "Reorder Point (ROP)", "Maximum Stock (MAX)",
        "Current Stock", "On Order", "Order Quantity", "Stock Value", "Order Value"
    ]
    ws.append(headers)

    for rec in results:
        row = [
            rec["item_code"], rec["description"], rec["supplier"],
            rec.get("category") or "", rec["avg_cost"] if rec["avg_cost"] > 0 else 0,
            rec["min"], 0, # ROP
            rec["max"], rec["current_stock"], rec["on_order"],
            f'=IF([@[Maximum Stock (MAX)]]-[@[Current Stock]]-[@[On Order]]>0,[@[Maximum Stock (MAX)]]-[@[Current Stock]]-[@[On Order]],0)',
            f'=IFERROR([@[Current Stock]]*[@[Cost Price]],0)',
            f'=IFERROR([@[Order Quantity]]*[@[Cost Price]],0)'
        ]
        ws.append(row)

    totals = ["TOTAL STOCK"] + [""] * 12
    ws.append(totals)

    for row in ws.iter_rows(min_row=2, max_row=len(results)+1):
        row[4].number_format = 'R #,##0.00'
        row[11].number_format = 'R #,##0.00'
        row[12].number_format = 'R #,##0.00'

    table_name = "cat_" + "".join(filter(str.isalnum, ws.title))
    table_name = table_name[:31]
    tab = Table(displayName=table_name, ref=f"A1:M{len(results)+2}")
    tab.totalsRowCount = 1
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    
    tab._initialise_columns()
    for idx, name in enumerate(headers):
        if name in ["Safety Stock", "Maximum Stock (MAX)", "Order Quantity", "Stock Value", "Order Value"]:
            tab.tableColumns[idx].totalsRowFunction = "sum"
            
    ws.add_table(tab)
    ws.freeze_panes = "A2"

def write_category_summary_sheet('''
        content = re.sub(r'def write_category_sheet\(ws, results: list\[dict\]\) -> None:.*?def write_category_summary_sheet\(', new_cat, content, flags=re.DOTALL)

    if "def write_category_summary_sheet" in content:
        new_sum = '''write_category_summary_sheet(ws, results: list[dict]) -> None:
    headers = ["Category", "Total Items", "Total QTY Purchased", "Total QTY Sold"]
    ws.append(headers)
    
    cat_agg = {}
    for rec in results:
        cat = rec.get("category") or "Uncategorized"
        if cat not in cat_agg:
            cat_agg[cat] = {"item_count": 0, "purchased": 0, "sold": 0}
        cat_agg[cat]["item_count"] += 1
        cat_agg[cat]["purchased"] += rec["total_purchased"]
        cat_agg[cat]["sold"] += rec["total_sold"]
        
    for cat in sorted(cat_agg.keys()):
        d = cat_agg[cat]
        ws.append([cat, d["item_count"], d["purchased"], d["sold"]])

    totals = ["TOTAL STOCK"] + [""] * 3
    ws.append(totals)

    table_name = "CategorySummary"
    tab = Table(displayName=table_name, ref=f"A1:D{len(cat_agg)+2}")
    tab.totalsRowCount = 1
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    tab._initialise_columns()
    for idx in range(1, 4):
        tab.tableColumns[idx].totalsRowFunction = "sum"
        
    ws.add_table(tab)
    ws.freeze_panes = "A2"
    
def main():'''
        content = re.sub(r'def write_category_summary_sheet.*?def main\(\):', "def " + new_sum, content, flags=re.DOTALL)

    with open(filepath, 'w', encoding='utf-8') as f:
        f.write(content)

process_file('c:/Users/Fan Movement/OneDrive - Fan Movement (Pty) Ltd/Desktop/Operations/2. AvgMovement/4_Scripts/item_movement_report.py')
process_file('c:/Users/Fan Movement/OneDrive - Fan Movement (Pty) Ltd/Desktop/Operations/2. AvgMovement/4_Scripts/item_movement_by_cat.py')
