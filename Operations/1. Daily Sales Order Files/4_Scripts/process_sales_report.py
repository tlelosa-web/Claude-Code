"""
Process Ops Sales Order Report - applies 3 tasks to ALL sheets:
  TASK 1: Replace sales rep full names with initials
  TASK 2: Sort data rows by Delivery Date, Status, Date
  TASK 3: Update totals formulas to exclude Pending rows

Saves with _v1 suffix.
"""
import openpyxl
import os
import re
import datetime
import shutil

# ── Configuration ──────────────────────────────────────────────────────────────
SOURCE_FILE = "3_Live_Reports/Ops Sales Order Report - 05.2026.xlsx"

# Sales rep mapping: full name → initials
SALES_REP_MAP = {
    "Sizwe Tshabalala": "ST",
    "Mashudu Tshifularo": "MT",
    "Simon Mokgola": "SM",
    "Sales Agent Hucra Fans": "S.A.- HUCRA",
}

# Partial name mapping for R.O.A base resolution (first name or short form)
PARTIAL_NAME_MAP = {
    "sizwe": "ST",
    "mashudu": "MT",
    "simon": "SM",
    "sizwe tshabalala": "ST",
    "mashudu tshifularo": "MT",
    "simon mokgola": "SM",
}

# Pattern to detect already-initialised cells (e.g. "ST", "MT", "SM", "S.A.- HUCRA",
# or any R.O.A variant like "ST - R.O.A", "MT-R.O.A", "ST - R.O.A.", "ST-R.O.A.", "MT - R.O.A.")
INITIALS_PATTERN = re.compile(
    r'^(?:[A-Z]{2}|S\.A\.(?:[-–\s]+)?HUCRA)(?:\s*[-–]\s*R\.O\.A\.?)?$',
    re.IGNORECASE
)

# Pattern for "S.A. HUCRA" variants (various dash/space separators)
HUCRA_VARIANT_PATTERN = re.compile(
    r'^S\.?\s*A\.?\s*[-–\s]+HUCRA\s*FANS?\s*$',
    re.IGNORECASE
)

# Summary labels in column F that mark the end of the data block
SUMMARY_LABELS = [
    "Total Value (All):",
    "Cash Sale (May):", "Cash Sale (June):",
    "Account (May):", "Account (June):",
    "Due Today - Total:", "Ready-Dispatch:",
    "BREAKDOWN",
]


def find_data_block_end(ws):
    """Find the last row of the primary data block.
    Scans from row 3 downward for first summary label in column F.
    Returns the row number of the last data row (end row, inclusive).
    If no label found, returns the last non-empty row."""
    for row_idx in range(3, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=6).value  # Column F
        if val is not None:
            s = str(val).strip()
            for label in SUMMARY_LABELS:
                if s.upper() == label.upper() or s.upper().startswith(label.upper()):
                    return row_idx - 1
    # Fallback: last row with any value in column A, B, or C
    last_data = 2
    for row_idx in range(3, ws.max_row + 1):
        a = ws.cell(row=row_idx, column=1).value
        b = ws.cell(row=row_idx, column=2).value
        c = ws.cell(row=row_idx, column=3).value
        if a is not None or b is not None or c is not None:
            last_data = row_idx
    return last_data


def initials_for_name(name):
    """Return the initials for a full sales rep name, or None if no match."""
    if name is None:
        return None
    s = str(name).strip()
    if not s:
        return None

    # ── R.O.A handling ────────────────────────────────────────────────────
    # Detect "Name - R.O.A", "Name-R.O.A", "Name - R.O.A.", "Name-R.O.A." variants
    # where Name is a full name or already an initial
    roa_match = re.match(
        r'^(.+?)\s*[-–]\s*R\.?\s*O\.?\s*A\.?\s*$',
        s,
        re.IGNORECASE
    )
    if roa_match:
        base = roa_match.group(1).strip()
        # Resolve the base part to initials
        initials = initials_for_name(base)  # recursive call
        if initials:
            return f"{initials} - R.O.A"
        return None  # couldn't resolve base

    # ── Direct mapping ────────────────────────────────────────────────────
    if s in SALES_REP_MAP:
        return SALES_REP_MAP[s]

    # ── Already initials? ─────────────────────────────────────────────────
    if INITIALS_PATTERN.match(s):
        return None  # Skip, already done

    # ── Fuzzy matching ────────────────────────────────────────────────────
    # Lowercase, strip whitespace for comparison
    s_lower = s.lower().strip()
    for full_name, initials in SALES_REP_MAP.items():
        if s_lower == full_name.lower().strip():
            return initials

    return None  # No match found


def task1_replace_sales_reps(ws):
    """Replace full sales rep names in column H with initials on ALL rows."""
    skipped = []
    for row_idx in range(1, ws.max_row + 1):
        cell = ws.cell(row=row_idx, column=8)  # Column H
        current = cell.value
        if current is None:
            continue
        initials = initials_for_name(current)
        if initials:
            print(f"  [T1] Row {row_idx}: '{current}' → '{initials}'")
            cell.value = initials
        else:
            # Check if it's already initials - if not, report as skipped
            s = str(current).strip()
            if not INITIALS_PATTERN.match(s) and not any(
                s.lower() == name.lower() for name in SALES_REP_MAP
            ):
                # Might be an unmatched name
                skipped.append((row_idx, s))
    return skipped


def task2_sort_data_rows(ws, start, end):
    """Sort data rows by: 1) Column B (Delivery Date) asc, 2) Column I (Status) asc, 3) Column A (Date) asc.
    Re-anchor column L formulas that reference column B of the same row."""
    if end < start:
        return

    rows = end - start + 1
    if rows <= 1:
        return

    # Extract data
    data_rows = []
    for row_idx in range(start, end + 1):
        row_data = []
        for col_idx in range(1, 13):  # Columns A-L
            row_data.append(ws.cell(row=row_idx, column=col_idx).value)
        data_rows.append((row_idx, row_data))

    # Sort by: B (idx 1), I (idx 8), A (idx 0)
    def sort_key(item):
        _, data = item
        b_val = data[1] if data[1] is not None else datetime.datetime.min
        i_val = str(data[8] or "").strip().lower() if data[8] is not None else ""
        a_val = data[0] if data[0] is not None else datetime.datetime.min
        return (b_val, i_val, a_val)

    sorted_rows = sorted(data_rows, key=sort_key)

    # Write sorted data back
    for new_offset, (old_row_idx, row_data) in enumerate(sorted_rows):
        new_row_idx = start + new_offset
        # Copy all values except column L (will re-anchor below)
        for col_idx in range(1, 12):  # Columns A-K
            ws.cell(row=new_row_idx, column=col_idx).value = row_data[col_idx - 1]
        # Column L will be set below

    # Re-anchor column L formulas: each data row's L should be =B{new_row_number}
    for offset in range(rows):
        row_idx = start + offset
        ws.cell(row=row_idx, column=12).value = f"=B{row_idx}"


def task3_update_formulas(ws, start, end):
    """Update the 5 formula cells in column G to exclude Pending rows."""
    # Locate the 5 target cells by scanning column F for the labels
    targets = {
        "Total Value (All):": None,
        "Cash Sale (May):": None,
        "Cash Sale (June):": None,
        "Account (May):": None,
        "Account (June):": None,
    }

    for row_idx in range(1, ws.max_row + 1):
        val = ws.cell(row=row_idx, column=6).value
        if val is not None:
            s = str(val).strip()
            if s in targets:
                targets[s] = row_idx

    # Build formulas
    formulas = {
        "Total Value (All):": f'=SUMIF(I{start}:I{end},"<>Pending",G{start}:G{end})',
        "Cash Sale (May):": (
            f'=SUMIFS(G{start}:G{end},J{start}:J{end},"*CASH SALE*",'
            f'I{start}:I{end},"<>Pending",'
            f'L{start}:L{end},">="&EOMONTH(TODAY(),-1)+1,'
            f'L{start}:L{end},"<="&EOMONTH(TODAY(),0))'
        ),
        "Cash Sale (June):": (
            f'=SUMIFS(G{start}:G{end},J{start}:J{end},"*CASH SALE*",'
            f'I{start}:I{end},"<>Pending",'
            f'L{start}:L{end},">="&EOMONTH(TODAY(),0)+1,'
            f'L{start}:L{end},"<="&EOMONTH(TODAY(),1))'
        ),
        "Account (May):": (
            f'=SUMIFS(G{start}:G{end},J{start}:J{end},"*ACCOUNT*",'
            f'I{start}:I{end},"<>Pending",'
            f'L{start}:L{end},">="&EOMONTH(TODAY(),-1)+1,'
            f'L{start}:L{end},"<="&EOMONTH(TODAY(),0))'
        ),
        "Account (June):": (
            f'=SUMIFS(G{start}:G{end},J{start}:J{end},"*ACCOUNT*",'
            f'I{start}:I{end},"<>Pending",'
            f'L{start}:L{end},">="&EOMONTH(TODAY(),0)+1,'
            f'L{start}:L{end},"<="&EOMONTH(TODAY(),1))'
        ),
    }

    for label, formula in formulas.items():
        row_idx = targets.get(label)
        if row_idx:
            cell = ws.cell(row=row_idx, column=7)  # Column G
            print(f"  [T3] Row {row_idx} '{label}': updating formula")
            cell.value = formula
        else:
            print(f"  [T3] Warning: '{label}' not found on this sheet - skipping")


def process_workbook(filepath):
    """Main processing function."""
    print(f"\n{'='*60}")
    print(f"Opening: {filepath}")
    
    # Open with formulas
    try:
        wb = openpyxl.load_workbook(filepath)
    except PermissionError:
        print("ERROR: File is open. Close it and try again.")
        return False
    except Exception as e:
        print(f"ERROR: {e}")
        return False

    all_skipped = []

    for sname in wb.sheetnames:
        ws = wb[sname]
        print(f"\n{'─'*50}")
        print(f"Processing sheet: {sname}")

        # ── Find data block boundaries ────────────────────────────────────
        data_end = find_data_block_end(ws)
        data_start = 3
        print(f"  Data block: rows {data_start}–{data_end}")

        # ── TASK 1: Sales Rep initials (entire sheet) ─────────────────────
        print(f"  [T1] Replacing sales rep names with initials...")
        skipped = task1_replace_sales_reps(ws)
        all_skipped.extend([(sname, *s) for s in skipped])

        # ── TASK 2: Sort data rows ───────────────────────────────────────
        print(f"  [T2] Sorting data rows {data_start}–{data_end}...")
        task2_sort_data_rows(ws, data_start, data_end)

        # ── TASK 3: Update formulas to exclude Pending ───────────────────
        print(f"  [T3] Updating formula cells to exclude Pending...")
        task3_update_formulas(ws, data_start, data_end)

    # ── Save with _v1 suffix ─────────────────────────────────────────────
    base, ext = os.path.splitext(filepath)
    outpath = f"{base}_v1{ext}"
    print(f"\n{'='*60}")
    print(f"Saving to: {outpath}")
    try:
        wb.save(outpath)
        print("Saved successfully.")
    except PermissionError:
        print("ERROR: Could not save. File may be open.")
        return False

    # ── Report skipped sales reps ────────────────────────────────────────
    if all_skipped:
        print(f"\n{'─'*50}")
        print(f"WARNING: {len(all_skipped)} sales rep value(s) were not matched and skipped:")
        for sheet, row, val in all_skipped:
            print(f"  Sheet '{sheet}', Row {row}: '{val}'")
    else:
        print("\nAll sales rep names were matched and replaced successfully.")

    return True


if __name__ == "__main__":
    if not os.path.exists(SOURCE_FILE):
        print(f"ERROR: {SOURCE_FILE} not found.")
        exit(1)

    success = process_workbook(SOURCE_FILE)
    if success:
        print("\nDone! All changes applied.")
    else:
        print("\nFailed to process workbook.")
        exit(1)