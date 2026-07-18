"""Excel export builder for the native "Sales Order Report" -- replicates
the standalone `1. Daily Sales Order Files` pipeline's daily-regenerated
report natively inside SOPS, generated on-demand from SOPS's own DB. See
docs/specs/sales-order-report-excel-export-2026-07-17.md "Export design"
section (final) for the exact column list, summary block, and Sheet 2
sourcing rules.

Sheet 1 ("Sales Order Report") -- built from the `sales_orders` list.
`Total` and every summary total use `SalesOrder.balance_due`, not
`total_incl` (Decision 3) -- this is the corrected design replacing the
old pipeline's manual per-job Total correction. `Status` reads
`display_report_status` (computed, includes the On Hold overlay), rendered
as '-' for Closed/Cancelled SOs (only reachable via `?view=all`).

Sheet 2 ("Monthly INV - MM.YYYY") -- built entirely from the `invoices`
list (an `Invoice` query result), independent of `SalesOrder`/`balance_due`.
Cash Sale vs Account bucketing reuses the old pipeline's own heuristic
(due date in the same month as the invoice date => "Cash Sale"-like
same-month settlement; a later due date => "Account"-like credit terms) --
`Invoice` has no explicit Cash/Account field of its own, matching the
source CSV.

No Excel formulas anywhere -- every value is computed directly in Python
from the in-memory lists passed in, since this workbook is generated fresh
on every request (unlike the old pipeline's daily-mutated, formula-linked
file).
"""
from datetime import date

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_THIN = Side(style='thin')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HEADER_FONT = Font(bold=True, color='FFFFFF')
_HEADER_FILL = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
_TITLE_FONT = Font(bold=True, size=14)
_BOLD = Font(bold=True)
_DATE_FMT = 'DD/MM/YYYY'

SHEET1_HEADERS = [
    'Date', 'Delivery Date', 'Job Number', 'SO Number', 'Customer Ref.',
    'Customer', 'Total', 'Sales Rep', 'Status', 'Payment Status', 'Notes',
]
_LEFT_ALIGN_COLS = (5, 6, 11)  # Customer Ref., Customer, Notes


def _next_month(year, month):
    if month == 12:
        return year + 1, 1
    return year, month + 1


def build_sales_order_report_workbook(sales_orders, invoices, currency_symbol='R'):
    """Return an openpyxl Workbook with the 2-sheet Sales Order Report.

    `sales_orders` -- a list/query result of SalesOrder rows, already
    filtered by the caller (route) to the desired `?view=open|all` scope.
    `invoices` -- a list/query result of every Invoice row (Sheet 2's math
    needs the full table regardless of the SO view filter).
    """
    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = 'Sales Order Report'
    _build_sheet1(ws1, sales_orders, currency_symbol)

    today = date.today()
    ws2 = wb.create_sheet(title=f"Monthly INV - {today.strftime('%m.%Y')}")
    _build_sheet2(ws2, invoices, today)

    return wb


def _style_header_row(ws, row, headers):
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.border = _BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _write_so_row(ws, row_idx, so, currency_fmt, include_notes=True):
    status = so.display_report_status or '-'
    values = [
        so.so_date, so.delivery_date, so.job_numbers or '', so.so_number,
        so.reference or '', so.customer_name or '', so.balance_due,
        so.sales_rep or '', status, so.payment_status or '',
    ]
    if include_notes:
        values.append(so.report_notes or '')

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.border = _BORDER
        horizontal = 'left' if col_idx in _LEFT_ALIGN_COLS else 'center'
        cell.alignment = Alignment(horizontal=horizontal, vertical='top', wrap_text=True)

    ws.cell(row=row_idx, column=1).number_format = _DATE_FMT
    ws.cell(row=row_idx, column=2).number_format = _DATE_FMT
    ws.cell(row=row_idx, column=7).number_format = currency_fmt


def _build_sheet1(ws, sales_orders, currency_symbol):
    today = date.today()
    cur_m, cur_y = today.month, today.year
    nxt_y, nxt_m = _next_month(cur_y, cur_m)
    currency_fmt = f'"{currency_symbol}" #,##0.00'

    # Title row + generated date
    ws.cell(row=1, column=1, value='Sales Order Report').font = _TITLE_FONT
    date_cell = ws.cell(row=1, column=len(SHEET1_HEADERS), value=today)
    date_cell.font = _TITLE_FONT
    date_cell.number_format = _DATE_FMT

    _style_header_row(ws, 2, SHEET1_HEADERS)

    row_idx = 3
    for so in sales_orders:
        _write_so_row(ws, row_idx, so, currency_fmt)
        row_idx += 1
    last_data_row = row_idx - 1

    widths = [12, 14, 14, 14, 16, 28, 14, 10, 16, 20, 40]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = 'E3'

    # --- Summary block -- all totals use balance_due (Decision 3), computed
    # directly in Python from `sales_orders`, no Excel formulas. ---
    total_value = sum(so.balance_due for so in sales_orders)
    summary_row = last_data_row + 2
    ws.cell(row=summary_row, column=6, value='Total Value:').font = _BOLD
    tv_cell = ws.cell(row=summary_row, column=7, value=round(total_value, 2))
    tv_cell.font = _BOLD
    tv_cell.number_format = currency_fmt

    # Breakdown: Cash Sale / Account x this month / next month, bucketed by
    # delivery_date, same layout as the old pipeline's BREAKDOWN section.
    breakdown_header_row = summary_row + 2
    ws.cell(row=breakdown_header_row, column=6, value='BREAKDOWN').font = Font(bold=True, underline='single')

    cur_month_name = today.strftime('%B')
    nxt_month_name = date(nxt_y, nxt_m, 1).strftime('%B')

    cash_cur = acct_cur = cash_nxt = acct_nxt = 0.0
    for so in sales_orders:
        d = so.delivery_date
        if not d:
            continue
        pay = (so.payment_status or '').upper()
        is_cash = 'CASH SALE' in pay
        is_account = 'ACCOUNT' in pay
        bal = so.balance_due
        if d.month == cur_m and d.year == cur_y:
            if is_cash:
                cash_cur += bal
            if is_account:
                acct_cur += bal
        elif d.month == nxt_m and d.year == nxt_y:
            if is_cash:
                cash_nxt += bal
            if is_account:
                acct_nxt += bal

    r1 = breakdown_header_row + 1
    r2 = breakdown_header_row + 2
    ws.cell(row=r1, column=6, value=f'Cash Sale ({cur_month_name}):')
    ws.cell(row=r2, column=6, value=f'Account ({cur_month_name}):')
    ws.cell(row=r1, column=9, value=f'Cash Sale ({nxt_month_name}):')
    ws.cell(row=r2, column=9, value=f'Account ({nxt_month_name}):')

    for row, col, val in ((r1, 7, cash_cur), (r2, 7, acct_cur), (r1, 10, cash_nxt), (r2, 10, acct_nxt)):
        cell = ws.cell(row=row, column=col, value=round(val, 2))
        cell.number_format = currency_fmt

    # --- Due-Today detail table ---
    due_today_orders = [so for so in sales_orders if so.delivery_date == today]
    dt_header_row = r2 + 3
    _style_header_row(ws, dt_header_row, SHEET1_HEADERS[:10])

    dt_row = dt_header_row + 1
    for so in due_today_orders:
        _write_so_row(ws, dt_row, so, currency_fmt, include_notes=False)
        dt_row += 1

    due_today_total = sum(so.balance_due for so in due_today_orders)
    ws.cell(row=dt_row, column=6, value='Due Today - Total:').font = _BOLD
    dtt_cell = ws.cell(row=dt_row, column=7, value=round(due_today_total, 2))
    dtt_cell.font = _BOLD
    dtt_cell.number_format = currency_fmt

    # --- Ready-Dispatch total (base report_status, ignoring the On Hold
    # overlay -- On Hold is a payment/dispatch caveat, not a production-
    # readiness state). ---
    rd_row = dt_row + 2
    ready_dispatch_total = sum(
        so.balance_due for so in sales_orders if so.report_status == 'Ready-Dispatch'
    )
    ws.cell(row=rd_row, column=6, value='Ready-Dispatch:').font = _BOLD
    rd_cell = ws.cell(row=rd_row, column=7, value=round(ready_dispatch_total, 2))
    rd_cell.font = _BOLD
    rd_cell.number_format = currency_fmt


def _sum_invoices(invoices, inv_month, inv_year, due_month, due_year, due_match, field):
    """Sum `field` (`exclusive` or `total_selling`) across `invoices` where
    the invoice date falls in `inv_month`/`inv_year`, and the due date
    matching (or not matching) `due_month`/`due_year` is gated by
    `due_match` -- same due-date-based Cash/Account heuristic as the old
    pipeline's `update_invoice_tab.py::_sum_invoices()`."""
    total = 0.0
    for inv in invoices:
        if not inv.invoice_date or not inv.due_date:
            continue
        if inv.invoice_date.month != inv_month or inv.invoice_date.year != inv_year:
            continue
        due_matches = (inv.due_date.month == due_month and inv.due_date.year == due_year)
        if due_match == due_matches:
            total += getattr(inv, field) or 0.0
    return total


def _write_month_block(ws, start_row, month_label, cash_excl, acct_excl, cash_sell, acct_sell, currency_fmt):
    """Write one month's Cash Sale / Account / Total block (3 rows) at
    `start_row`, columns A-F. Returns the row number of the next block's
    label row (2 blank rows below this block's Total row)."""
    ws.cell(row=start_row, column=1, value=f'{month_label}:').font = _BOLD

    ws.cell(row=start_row, column=2, value='Cash Sale Invoiced')
    c = ws.cell(row=start_row, column=3, value=round(cash_excl, 2))
    c.number_format = currency_fmt
    ws.cell(row=start_row, column=5, value='Cash Sale Invoiced (Incl. VAT)')
    c = ws.cell(row=start_row, column=6, value=round(cash_sell, 2))
    c.number_format = currency_fmt

    r2 = start_row + 1
    ws.cell(row=r2, column=2, value='Account Invoiced')
    c = ws.cell(row=r2, column=3, value=round(acct_excl, 2))
    c.number_format = currency_fmt
    ws.cell(row=r2, column=5, value='Account Invoiced (Incl. VAT)')
    c = ws.cell(row=r2, column=6, value=round(acct_sell, 2))
    c.number_format = currency_fmt

    r3 = start_row + 2
    ws.cell(row=r3, column=2, value='Total').font = _BOLD
    c = ws.cell(row=r3, column=3, value=round(cash_excl + acct_excl, 2))
    c.font = _BOLD
    c.number_format = currency_fmt
    ws.cell(row=r3, column=5, value='Total (Incl. VAT)').font = _BOLD
    c = ws.cell(row=r3, column=6, value=round(cash_sell + acct_sell, 2))
    c.font = _BOLD
    c.number_format = currency_fmt

    return r3 + 2


def _build_sheet2(ws, invoices, today):
    cur_m, cur_y = today.month, today.year
    nxt_y, nxt_m = _next_month(cur_y, cur_m)
    cur_month_name = date(cur_y, cur_m, 1).strftime('%B')
    nxt_month_name = date(nxt_y, nxt_m, 1).strftime('%B')
    currency_fmt = '#,##0.00'

    ws.cell(row=1, column=1, value='Invoicing').font = _TITLE_FONT

    cash_cur_excl = _sum_invoices(invoices, cur_m, cur_y, cur_m, cur_y, True, 'exclusive')
    acct_cur_excl = _sum_invoices(invoices, cur_m, cur_y, cur_m, cur_y, False, 'exclusive')
    cash_cur_sell = _sum_invoices(invoices, cur_m, cur_y, cur_m, cur_y, True, 'total_selling')
    acct_cur_sell = _sum_invoices(invoices, cur_m, cur_y, cur_m, cur_y, False, 'total_selling')

    next_start = _write_month_block(
        ws, 3, f'Current Month ({cur_month_name})',
        cash_cur_excl, acct_cur_excl, cash_cur_sell, acct_cur_sell, currency_fmt,
    )

    cash_nxt_excl = _sum_invoices(invoices, nxt_m, nxt_y, nxt_m, nxt_y, True, 'exclusive')
    acct_nxt_excl = _sum_invoices(invoices, nxt_m, nxt_y, nxt_m, nxt_y, False, 'exclusive')
    cash_nxt_sell = _sum_invoices(invoices, nxt_m, nxt_y, nxt_m, nxt_y, True, 'total_selling')
    acct_nxt_sell = _sum_invoices(invoices, nxt_m, nxt_y, nxt_m, nxt_y, False, 'total_selling')

    _write_month_block(
        ws, next_start, f'Next Month ({nxt_month_name})',
        cash_nxt_excl, acct_nxt_excl, cash_nxt_sell, acct_nxt_sell, currency_fmt,
    )

    for col, width in zip('ABCDEF', (24, 26, 16, 4, 30, 16)):
        ws.column_dimensions[col].width = width
