"""
Item Movement Report Generator
Parses raw ItemMovementReport.csv and produces a formatted Excel report with two sheets.

Sheet 1 – "Item Movement Summary" (15 columns):
  Item Code | Description | Category | Period | Total QTY Purchased | Supplier | Cost |
  Total QTY Sold | Lead Time | AMU | Min | Current Stock | Max | Re-Order Qty | Re-Order Value

Sheet 2 – "Order Summary" (11 columns – subset A,B,C,F,G,I,K,L,M,N,O):
  Item Code | Description | Category | Supplier | Cost | Lead Time | Min |
  Current Stock | Max | Re-Order Qty | Re-Order Value

Enriched with:
  - ItemListingReport.csv  (description, category, avg cost)
  - OutstandingPOByItemReport.csv (top supplier, lead time in weeks)
"""

import csv
import math
import os
import sys
import tempfile
import shutil
from datetime import datetime
from collections import defaultdict, Counter
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.worksheet.properties import PageSetupProperties
from openpyxl.worksheet.table import Table, TableStyleInfo

# ── Paths (relative to project root) ──────────────────────────────────
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
SOURCE_CSV = os.path.join(PROJECT_ROOT, "2_Source_Data", "ItemMovementReport.csv")
ITEM_LISTING_CSV = os.path.join(PROJECT_ROOT, "2_Source_Data", "ItemListingReport.csv")
PO_BY_ITEM_CSV = os.path.join(PROJECT_ROOT, "2_Source_Data", "OutstandingPOByItemReport.csv")
STOCK_XLSX = os.path.join(PROJECT_ROOT, "2_Source_Data", "Stock Count Update.xlsx")
OUTPUT_DIR = os.path.join(PROJECT_ROOT, "3_Live_Reports")
DEBUG_LOG = os.path.join(PROJECT_ROOT, "5_Archive_and_Debug", "debug_output_utf8.txt")

# ── Formatting constants ──────────────────────────────────────────────
HEADER_FONT = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
BODY_FONT = Font(name="Calibri", size=11)
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def log(msg: str) -> None:
    """Write to both stdout and debug log."""
    print(msg)
    with open(DEBUG_LOG, "a", encoding="utf-8") as f:
        f.write(msg + "\n")


def safe_copy_for_reading(src: str) -> str:
    """Shadow-copy a file so we never read a potentially locked live file."""
    tmp = tempfile.NamedTemporaryFile(delete=False, suffix=os.path.splitext(src)[1])
    tmp.close()
    shutil.copy2(src, tmp.name)
    return tmp.name


def parse_date(s: str) -> datetime | None:
    """Parse DD/MM/YYYY date strings."""
    s = s.strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


def round_amu_half(val: float) -> float:
    """Round AMU up to the next 0.5 increment, 1 decimal place, minimum 1.0."""
    if val <= 0:
        return 1.0
    # Ceiling to next 0.5
    rounded = math.ceil(val * 2) / 2
    # Round to 1 decimal place
    rounded = round(rounded, 1)
    return max(rounded, 1.0)


# ──────────────────────────────────────────────────────────────────────
# PARSERS
# ──────────────────────────────────────────────────────────────────────

def parse_csv(filepath: str) -> tuple[list[dict], dict[str, dict]]:
    """
    Parse the raw ItemMovementReport CSV into item records.

    Returns:
        items: list of dicts with keys:
               item_code, item_desc_raw, date, ref, desc, customer_supplier, qty, is_purchase, is_sale
        item_date_ranges: dict of item_code -> {min_date, max_date}
    """
    items: list[dict] = []
    current_item_code: str = ""
    current_item_desc_raw: str = ""
    item_date_ranges: dict[str, dict] = {}

    with open(filepath, "r", encoding="utf-8-sig", newline="") as f:
        reader = csv.reader(f)

        # Skip title row
        try:
            next(reader)
        except StopIteration:
            return items, item_date_ranges

        # Skip metadata rows 1-3
        for _ in range(3):
            try:
                next(reader)
            except StopIteration:
                break

        # ── Read data rows ──
        for row in reader:
            if not row or not row[0].strip():
                continue

            first = row[0].strip()

            # "Total for Item:" line -> end of current item block
            if first.lower().startswith("total for item:"):
                current_item_code = ""
                current_item_desc_raw = ""
                continue

            dt = parse_date(first)
            if dt and len(row) >= 5:
                ref = row[1].strip()
                desc = row[2].strip()
                cust = row[3].strip()
                try:
                    qty = float(row[4].strip())
                except ValueError:
                    qty = 0.0

                is_purchase = desc == "Supplier Invoice"
                is_sale = desc == "Tax Invoice"

                # Track per-item date range
                if current_item_code:
                    if current_item_code not in item_date_ranges:
                        item_date_ranges[current_item_code] = {"min_date": dt, "max_date": dt}
                    else:
                        dr = item_date_ranges[current_item_code]
                        if dt < dr["min_date"]:
                            dr["min_date"] = dt
                        if dt > dr["max_date"]:
                            dr["max_date"] = dt

                items.append({
                    "item_code": current_item_code,
                    "item_desc_raw": current_item_desc_raw,
                    "date": dt,
                    "ref": ref,
                    "desc": desc,
                    "customer_supplier": cust,
                    "qty": qty,
                    "is_purchase": is_purchase,
                    "is_sale": is_sale,
                })
            else:
                # Item name row — format is "Code - Description" or just "Code"
                current_item_code = first
                current_item_desc_raw = ""
                if " - " in first:
                    parts = first.split(" - ", 1)
                    current_item_code = parts[0].strip()
                    current_item_desc_raw = parts[1].strip()

    return items, item_date_ranges


def load_item_listing_map(filepath: str) -> dict[str, dict]:
    """
    Load ItemListingReport.csv into a lookup: {Code -> {description, category, avg_cost}}.
    Uses shadow-copy protocol. Handles the 'sep=,' directive line.
    """
    result: dict[str, dict] = {}
    if not os.path.isfile(filepath):
        return result

    tmp_path = safe_copy_for_reading(filepath)
    try:
        with open(tmp_path, "r", encoding="utf-8-sig") as f:
            first_line = f.readline().strip()
            if not first_line.startswith("sep="):
                f.seek(0)

            reader = csv.DictReader(f)
            for row in reader:
                code = row.get("Code", "").strip()
                if code:
                    # Parse Avg. Cost — may be "R 1,248.00" format
                    raw_cost = row.get("Avg. Cost", "").strip().replace(",", "")
                    try:
                        avg_cost = float(raw_cost)
                    except (ValueError, TypeError):
                        avg_cost = 0.0
                    result[code] = {
                        "description": row.get("Description", "").strip() or "",
                        "category": row.get("Category", "").strip() or "",
                        "avg_cost": avg_cost,
                    }
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)
    return result


def parse_po_by_item(filepath: str) -> dict[str, dict]:
    """
    Parse OutstandingPOByItemReport.csv and extract per-item:
      - top_supplier: supplier with the most PO transactions
      - avg_lead_time_weeks: average (delivery_date - order_date) in weeks
      - pending_qty: sum of Qty where Status = "Pending"

    Returns: {item_code -> {top_supplier, avg_lead_time_weeks, pending_qty}}
    """
    if not os.path.isfile(filepath):
        return {}

    tmp_path = safe_copy_for_reading(filepath)
    # Per-item accumulators
    supplier_counts: dict[str, Counter] = defaultdict(Counter)
    lead_times: dict[str, list[float]] = defaultdict(list)
    pending_qty: dict[str, float] = defaultdict(float)

    try:
        with open(tmp_path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)

            # Skip title + metadata rows
            try:
                next(reader)  # title
            except StopIteration:
                return {}
            for _ in range(3):
                try:
                    next(reader)
                except StopIteration:
                    break

            current_item_code = ""

            for row in reader:
                if not row or not row[0].strip():
                    continue

                first = row[0].strip()

                # "Total for Item:" line -> end of block
                if first.lower().startswith("total for item:"):
                    current_item_code = ""
                    continue

                dt = parse_date(first)
                if dt and len(row) >= 6:
                    # Columns: Date, Reference, Order No., Supplier, Delivery Date, Status, Qty, ...
                    supplier = row[3].strip()
                    delivery_date = parse_date(row[4].strip())
                    status = row[5].strip()
                    try:
                        po_qty = float(row[6].strip())
                    except (ValueError, IndexError):
                        po_qty = 0.0

                    # Extract item code from combined name
                    if current_item_code and " - " in current_item_code:
                        code_part = current_item_code.split(" - ", 1)[0].strip()
                    else:
                        code_part = current_item_code

                    if supplier:
                        supplier_counts[code_part][supplier] += 1

                    if delivery_date:
                        days_diff = (delivery_date - dt).days
                        weeks = days_diff / 7.0
                        if weeks >= 0:  # Only valid lead times
                            lead_times[code_part].append(weeks)

                    if status == "Pending":
                        pending_qty[code_part] += po_qty
                else:
                    # Item name row
                    current_item_code = first

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    # Build result
    all_codes = set(supplier_counts.keys()) | set(lead_times.keys())
    result: dict[str, dict] = {}
    for code in all_codes:
        top_supplier = ""
        if code in supplier_counts and supplier_counts[code]:
            top_supplier = supplier_counts[code].most_common(1)[0][0]

        avg_lt = 0.0
        if code in lead_times and lead_times[code]:
            avg_lt = round(sum(lead_times[code]) / len(lead_times[code]), 1)

        pqty = round(pending_qty.get(code, 0.0), 2)

        result[code] = {
            "top_supplier": top_supplier,
            "avg_lead_time_weeks": avg_lt,
            "pending_qty": pqty,
        }

    return result


def load_current_stock_map(filepath: str) -> dict[str, float]:
    if not os.path.isfile(filepath):
        return {}

    tmp_path = safe_copy_for_reading(filepath)
    try:
        wb = load_workbook(tmp_path, data_only=True)
        sheet = "ImportStockCount_TOT" if "ImportStockCount_TOT" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]

        header_row_idx = 0
        item_col_idx = -1
        stock_col_idx = -1

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), 1):
            if not row:
                continue
            row_norm = [str(v).strip().lower() if v is not None else "" for v in row]
            if "item" in row_norm and "current stock" in row_norm:
                header_row_idx = i
                item_col_idx = row_norm.index("item")
                stock_col_idx = row_norm.index("current stock")
                break

        if header_row_idx == 0 or item_col_idx < 0 or stock_col_idx < 0:
            return {}

        result: dict[str, float] = {}
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or item_col_idx >= len(row):
                continue
            code = row[item_col_idx]
            if code is None:
                continue
            code = str(code).strip()
            if not code:
                continue
            raw_stock = row[stock_col_idx] if stock_col_idx < len(row) else 0
            try:
                stock = float(raw_stock) if raw_stock is not None else 0.0
            except (TypeError, ValueError):
                stock = 0.0
            result[code] = stock

        return result
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def load_current_stock_raw_map(filepath: str) -> dict[str, float | None]:
    if not os.path.isfile(filepath):
        return {}

    tmp_path = safe_copy_for_reading(filepath)
    try:
        wb = load_workbook(tmp_path, data_only=True)
        sheet = "ImportStockCount_TOT" if "ImportStockCount_TOT" in wb.sheetnames else wb.sheetnames[0]
        ws = wb[sheet]

        header_row_idx = 0
        item_col_idx = -1
        stock_col_idx = -1

        for i, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), 1):
            if not row:
                continue
            row_norm = [str(v).strip().lower() if v is not None else "" for v in row]
            if "item" in row_norm and "current stock" in row_norm:
                header_row_idx = i
                item_col_idx = row_norm.index("item")
                stock_col_idx = row_norm.index("current stock")
                break

        if header_row_idx == 0 or item_col_idx < 0 or stock_col_idx < 0:
            return {}

        result: dict[str, float | None] = {}
        for row in ws.iter_rows(min_row=header_row_idx + 1, values_only=True):
            if not row or item_col_idx >= len(row):
                continue
            code = row[item_col_idx]
            if code is None:
                continue
            code = str(code).strip()
            if not code:
                continue

            raw_stock = row[stock_col_idx] if stock_col_idx < len(row) else None
            if raw_stock is None or (isinstance(raw_stock, str) and not raw_stock.strip()):
                result[code] = None
                continue
            try:
                result[code] = float(raw_stock)
            except (TypeError, ValueError):
                result[code] = None

        return result
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


def update_stock_count_update_file(item_listing_csv: str, stock_xlsx_path: str) -> tuple[bool, int]:
    existing_stock = load_current_stock_raw_map(stock_xlsx_path)

    tmp_path = safe_copy_for_reading(item_listing_csv)
    try:
        items: list[tuple[str, str, str, float]] = []
        with open(tmp_path, "r", encoding="utf-8-sig", newline="") as f:
            first_line = f.readline().strip()
            if not first_line.startswith("sep="):
                f.seek(0)

            reader = csv.DictReader(f)
            for row in reader:
                code = (row.get("Code") or "").strip()
                if not code:
                    continue
                desc = (row.get("Description") or "").strip()
                cat = (row.get("Category") or "").strip() or "Uncategorized"
                raw_cost = (row.get("Avg. Cost") or "").strip().replace(",", "").replace("R", "").strip()
                try:
                    cost = float(raw_cost) if raw_cost else 0.0
                except (TypeError, ValueError):
                    cost = 0.0
                items.append((code, desc, cat, cost))
    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)

    items.sort(key=lambda t: ((t[2] or "").strip().lower() == "uncategorized", (t[2] or "").lower(), t[0].lower()))

    tmp_out = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
    tmp_out.close()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "ImportStockCount_TOT"

        headers = ["Item", "Description", "Category", "Cost", "Current Stock", "Stock Value"]

        ws.append(["TOTAL STOCK"] + [None] * (len(headers) - 1))
        ws.append(headers)

        for code, desc, cat, cost in items:
            current = existing_stock.get(code)
            ws.append([code, desc, cat, cost, current, None])

        for row_idx in range(3, ws.max_row + 1):
            ws.cell(row=row_idx, column=6).value = f"=IFERROR(E{row_idx}*D{row_idx},0)"

        for col in range(1, len(headers) + 1):
            c = ws.cell(row=2, column=col)
            c.font = HEADER_FONT
            c.fill = HEADER_FILL
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = THIN_BORDER

        for r in ws.iter_rows(min_row=3, max_row=ws.max_row, min_col=1, max_col=len(headers)):
            for c in r:
                c.border = THIN_BORDER
                if c.column in (4, 6):
                    c.number_format = "R #,##0.00"

        ws.freeze_panes = "A3"

        col_widths = {1: 18, 2: 45, 3: 22, 4: 12, 5: 14, 6: 14}
        for col_idx, width in col_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        wb.save(tmp_out.name)

        os.makedirs(os.path.dirname(stock_xlsx_path), exist_ok=True)
        os.replace(tmp_out.name, stock_xlsx_path)
        return True, len(items)
    except Exception:
        if os.path.exists(tmp_out.name):
            os.unlink(tmp_out.name)
        raise


# ──────────────────────────────────────────────────────────────────────
# AGGREGATION
# ──────────────────────────────────────────────────────────────────────

def aggregate_items(
    items: list[dict],
    item_date_ranges: dict[str, dict],
    item_listing_map: dict[str, dict],
    po_data: dict[str, dict],
    current_stock_map: dict[str, float],
) -> list[dict]:
    """
    Aggregate transactions per item into a rich result dict.
    Includes ALL items from ItemListingReport, even those without transactions.
    """
    # First pass: aggregate transaction data
    agg: dict[str, dict] = defaultdict(lambda: {
        "total_purchased": 0.0,
        "total_sold": 0.0,
        "purchase_txns": 0,
        "sale_txns": 0,
    })

    for rec in items:
        key = rec["item_code"]
        if rec["is_purchase"]:
            agg[key]["total_purchased"] += abs(rec["qty"])
            agg[key]["purchase_txns"] += 1
        if rec["is_sale"]:
            agg[key]["total_sold"] += abs(rec["qty"])
            agg[key]["sale_txns"] += 1

    # Second pass: include ALL items from ItemListingReport
    results = []
    all_item_codes = set(agg.keys()) | set(item_listing_map.keys())
    
    for item_code in sorted(all_item_codes):
        d = agg.get(item_code, {
            "total_purchased": 0.0,
            "total_sold": 0.0,
            "purchase_txns": 0,
            "sale_txns": 0,
        })
        purchased = d["total_purchased"]
        sold = d["total_sold"]

        # Per-item period from first to last transaction
        dr = item_date_ranges.get(item_code)
        if dr and dr["min_date"] and dr["max_date"]:
            delta_years = (dr["max_date"].year - dr["min_date"].year) * 12 + (dr["max_date"].month - dr["min_date"].month)
            period_months = max(delta_years, 1)
        else:
            period_months = 1

        # AMU — raw value
        max_qty = max(purchased, sold)
        amu_raw = max_qty / period_months if period_months > 0 else 0.0
        amu = round_amu_half(amu_raw)

        # Enrich from ItemListingReport
        listing_info = item_listing_map.get(item_code, {})
        description = listing_info.get("description") or ""
        category = listing_info.get("category") or ""
        avg_cost = listing_info.get("avg_cost", 0.0)
        if not description:
            raw_descs = [rec["item_desc_raw"] for rec in items if rec["item_code"] == item_code and rec["item_desc_raw"]]
            description = raw_descs[0] if raw_descs else ""

        # Enrich from PO data
        po_info = po_data.get(item_code, {})
        top_supplier = po_info.get("top_supplier", "")
        lead_time_weeks = po_info.get("avg_lead_time_weeks", 0.0)
        # Convert lead time to months (min 0.5)
        lead_time_months = round(lead_time_weeks / 4.33, 1) if lead_time_weeks > 0 else 0.0
        lead_time_months = max(lead_time_months, 0.5)

        # Min / Max calculation
        # Min = AMU * (lead_time_weeks / 4.33)  -- cover lead time demand, 4.33 weeks per month avg
        # Max = Min + (AMU * 2)  -- cover reorder cycle + safety stock
        lm = lead_time_months if lead_time_months > 0 else 1
        min_qty = round(amu * lm)
        max_qty_min = min_qty + round(amu * 2)
        if min_qty <= 0:
            min_qty = 1
        if max_qty_min <= min_qty:
            max_qty_min = min_qty * 2

        current_stock = round(current_stock_map.get(item_code, 0.0), 0)
        # On Order (pending PO qty)
        on_order = round(po_info.get("pending_qty", 0.0), 0)
        if on_order < 0:
            on_order = 0

        results.append({
            "item_code": item_code,
            "description": description,
            "category": category,
            "period_months": period_months,
            "total_purchased": purchased,
            "total_sold": sold,
            "supplier": top_supplier,
            "avg_cost": avg_cost,
            "lead_time_months": lead_time_months,
            "amu": amu,
            "min": min_qty,
            "max": max_qty_min,
            "current_stock": current_stock,
            "on_order": on_order,
        })

    return results


# ──────────────────────────────────────────────────────────────────────
# EXCEL REPORT
# ──────────────────────────────────────────────────────────────────────

def _write_sheet(ws, results: list[dict], show_all: bool) -> None:
    if show_all:
        headers = [
            "Item Code", "Description", "Category", "Period (month(s))",
            "Total QTY Purchased", "Supplier", "Cost", "Total QTY Sold",
            "Lead Time (months)", "AMU", "Min", "Current Stock", "On Order",
            "Max", "Stock Value", "Re-Order Qty", "Re-Order Value"
        ]
        ws.title = "Item Movement Summary"
        max_col_letter = "Q"
    else:
        headers = [
            "Item", "Description", "Supplier", "Category", "Cost Price",
            "Safety Stock", "Reorder Point (ROP)", "Maximum Stock (MAX)",
            "Current Stock", "On Order", "Order Quantity", "Stock Value", "Order Value"
        ]
        ws.title = "Order Summary"
        max_col_letter = "M"

    ws.append(headers)

    for row_idx, rec in enumerate(results, 2):
        if show_all:
            row = [
                rec["item_code"], rec["description"], rec.get("category") or "",
                rec["period_months"], rec["total_purchased"], rec["supplier"],
                rec["avg_cost"] if rec["avg_cost"] > 0 else 0,
                rec["total_sold"], rec["lead_time_months"] if rec["lead_time_months"] > 0 else 0.5,
                rec["amu"], rec["min"], rec["current_stock"], rec["on_order"],
                rec["max"], f'=IFERROR(L{row_idx}*G{row_idx},0)',
                f'=IF(N{row_idx}-L{row_idx}-M{row_idx}>0,N{row_idx}-L{row_idx}-M{row_idx},0)',
                f'=IFERROR(P{row_idx}*G{row_idx},0)'
            ]
            ws.append(row)
        else:
            row = [
                rec["item_code"], rec["description"], rec["supplier"],
                rec.get("category") or "", rec["avg_cost"] if rec["avg_cost"] > 0 else 0,
                rec["min"], 0, # Reorder Point (ROP) defaulting to 0
                rec["max"], rec["current_stock"], rec["on_order"],
                f'=IF(H{row_idx}-I{row_idx}-J{row_idx}>0,H{row_idx}-I{row_idx}-J{row_idx},0)',
                f'=IFERROR(I{row_idx}*E{row_idx},0)',
                f'=IFERROR(K{row_idx}*E{row_idx},0)'
            ]
            ws.append(row)

    totals = ["TOTAL STOCK"] + [None] * (len(headers) - 1)
    ws.append(totals)

    # Force format cols before adding table
    for row in ws.iter_rows(min_row=2, max_row=len(results)+1):
        if show_all:
            row[6].number_format = 'R #,##0.00'
            row[14].number_format = 'R #,##0.00'
            row[16].number_format = 'R #,##0.00'
        else:
            row[4].number_format = 'R #,##0.00'
            row[11].number_format = 'R #,##0.00'
            row[12].number_format = 'R #,##0.00'

    table_name = ws.title.replace(" ", "")
    tab = Table(displayName=table_name, ref=f"A1:{max_col_letter}{len(results)+2}")
    tab.totalsRowCount = 1
    tab.tableStyleInfo = TableStyleInfo(name="TableStyleLight9", showRowStripes=True)
    
    tab._initialise_columns()
    for idx, name in enumerate(headers):
        tab.tableColumns[idx].name = name
        if name in ["Stock Value", "Re-Order Value", "Order Value", "Order Quantity", "Re-Order Qty", "Min", "Max", "Maximum Stock (MAX)", "Safety Stock"]:
            tab.tableColumns[idx].totalsRowFunction = "sum"
            
    ws.add_table(tab)

    ws.freeze_panes = "A2"
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_setup.horizontalCentered = True
    ws.page_margins.left = 1.0; ws.page_margins.right = 1.0
    ws.print_options.horizontalCentered = True
    ws.print_title_rows = "1:1"


def build_excel_report(results: list[dict]) -> str:
    """Create a formatted .xlsx report with two sheets and dynamic date-based filename."""
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    today_str = datetime.now().strftime("%d.%m.%Y")
    output_path = os.path.join(OUTPUT_DIR, f"Item Movement {today_str}.xlsx")

    wb = Workbook()

    # ── Sheet 1: Full 17-column view ──
    ws1 = wb.active
    ws1.title = "Item Movement Summary"
    _write_sheet(ws1, results, show_all=True)

    # ── Sheet 2: 13-column subset ──
    ws2 = wb.create_sheet("Order Summary")
    _write_sheet(ws2, results, show_all=False)

    wb.save(output_path)
    return output_path


# ──────────────────────────────────────────────────────────────────────
# MAIN
# ──────────────────────────────────────────────────────────────────────

def main() -> None:
    if os.path.exists(DEBUG_LOG):
        os.remove(DEBUG_LOG)

    log("=" * 60)
    log("Item Movement Report Pipeline — Starting")
    log("=" * 60)

    # ── Step 1: Validate source file ──
    if not os.path.isfile(SOURCE_CSV):
        log(f"ERROR: Source file not found: {SOURCE_CSV}")
        sys.exit(1)
    log(f"Source file: {SOURCE_CSV}")

    try:
        ok, n = update_stock_count_update_file(ITEM_LISTING_CSV, STOCK_XLSX)
        if ok:
            log(f"Stock Count Update refreshed (sorted by Category): {n} items")
    except Exception as e:
        log(f"WARNING: Could not refresh Stock Count Update file: {e}")

    # ── Step 2: Shadow-copy ──
    tmp_path = safe_copy_for_reading(SOURCE_CSV)
    log(f"Shadow copy created: {tmp_path}")

    try:
        # ── Step 3: Parse Item Movement CSV ──
        log("Parsing Item Movement CSV ...")
        items, item_date_ranges = parse_csv(tmp_path)
        log(f"  Found {len(items)} transaction rows")
        log(f"  Unique items with date ranges: {len(item_date_ranges)}")

        if not items:
            log("ERROR: No transaction data found in CSV.")
            sys.exit(1)

        # ── Step 3b: Load Item Listing reference ──
        log("Loading Item Listing reference ...")
        item_listing_map = load_item_listing_map(ITEM_LISTING_CSV)
        log(f"  Loaded {len(item_listing_map)} item listings")

        # ── Step 3c: Load PO By Item reference ──
        log("Loading Outstanding PO By Item reference ...")
        po_data = parse_po_by_item(PO_BY_ITEM_CSV)
        log(f"  Loaded PO data for {len(po_data)} items")

        log("Loading Current Stock reference ...")
        current_stock_map = load_current_stock_map(STOCK_XLSX)
        log(f"  Loaded current stock for {len(current_stock_map)} items")

        # ── Step 4: Aggregate ──
        log("Aggregating per-item totals with enrichment ...")
        results = aggregate_items(items, item_date_ranges, item_listing_map, po_data, current_stock_map)
        log(f"  Found {len(results)} unique items")

        # ── Step 5: Build report ──
        log("Building formatted Excel report ...")
        output_path = build_excel_report(results)
        log(f"Report saved: {output_path}")

        # ── Summary ──
        log("")
        log("-" * 60)
        log("PIPELINE COMPLETE")
        log(f"  Items processed : {len(results)}")
        log(f"  Output file     : {output_path}")
        log("-" * 60)

    finally:
        if os.path.exists(tmp_path):
            os.unlink(tmp_path)


if __name__ == "__main__":
    main()
