import openpyxl

file_path = "Sales Order Report - 03.2026.xlsx"
try:
    wb = openpyxl.load_workbook(file_path, data_only=False)
    sheet_name = "23.03.2026"
    
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print("====== HEADERS (Row 1) ======")
        headers = [ws.cell(row=1, column=c).value for c in range(1, 15)]
        print(headers)
        
        print("\n====== SAMPLE ROW 2 (Check Formulas/Data) ======")
        row2 = [ws.cell(row=2, column=c).value for c in range(1, 15)]
        print(row2)
        
        print("\n====== SAMPLE ROW ALIGNMENTS (Row 2) ======")
        alignments = [str(ws.cell(row=2, column=c).alignment.horizontal) if ws.cell(row=2, column=c).alignment else "None" for c in range(1, 15)]
        print(alignments)
        
        print(f"\n====== SHEET DIMENSIONS ======\nMax Row: {ws.max_row}, Max Col: {ws.max_column}")
        
        print("\n====== SUMMARY SECTION (Last 20 Rows) ======")
        for r in range(max(1, ws.max_row - 20), ws.max_row + 1):
            row_data = []
            for c in range(1, 13):
                val = ws.cell(row=r, column=c).value
                if val is not None:
                    row_data.append(f"Col {c}: {val}")
            if row_data:
                print(f"Row {r}: {row_data}")
                
    else:
        print(f"Sheet {sheet_name} not found.")
except Exception as e:
    print(f"Error reading file: {e}")
