import pandas as pd
import openpyxl
import re
import datetime

def main():
    comments_file = "Released_Jobs_Comments.xlsx"
    sales_file = "Sales Order Report - 03.2026.xlsx"
    
    print("--- Reading Released_Jobs_Comments.xlsx ---")
    df = pd.read_excel(comments_file)
    # Filter for actually having comments
    df_valid = df[df['Text Box Contents'] != 'NO TEXT BOXES FOUND']
    
    print("Found", len(df_valid), "PDFs with text boxes.")
    
    comments_map = {}
    for _, row in df_valid.iterrows():
        job = str(row['Guessed Order Number']).strip()
        txt = str(row['Text Box Contents']).strip().replace('\n', ' ')
        comments_map[job] = txt
        print(f"[{job}] -> {txt[:50]}...")
        
    print("\n--- Reading Sales Order Report ---")
    wb = openpyxl.load_workbook(sales_file, data_only=True)
    date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    sheets = [s for s in wb.sheetnames if date_pattern.match(s)]
    sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
    latest_sheet = sheets[-1]
    
    ws = wb[latest_sheet]
    print(f"Checking sheet: {latest_sheet}")
    
    for row_idx in range(3, ws.max_row + 1):
        so = str(ws.cell(row=row_idx, column=4).value).strip()
        job = str(ws.cell(row=row_idx, column=3).value).strip()
        pay = str(ws.cell(row=row_idx, column=10).value).strip()
        
        if job in comments_map:
            print(f"Match found Job: {job} | Excel says: '{pay}' | Comment was: '{comments_map[job][:30]}'")
        elif so in comments_map:
            print(f"Match found SO: {so} | Excel says: '{pay}' | Comment was: '{comments_map[so][:30]}'")

if __name__ == '__main__':
    main()
