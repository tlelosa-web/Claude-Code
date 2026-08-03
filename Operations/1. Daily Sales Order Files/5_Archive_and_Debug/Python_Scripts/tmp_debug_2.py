import pandas as pd
import openpyxl
import re
import datetime
import shutil
import tempfile
import os
import sys

# From script logic
def map_comment_to_status(comment):
    if pd.isna(comment) or not str(comment).strip() or comment == "NO TEXT BOXES FOUND" or comment == "ERROR READING PDF":
        return None
    text = str(comment).upper()
    if "ACCOUNT" in text:
        if "OVER LIMIT" in text: return "ACCOUNT - OVER LIMIT"
        if "UP TO DATE" in text: return "ACCOUNT - UP TO DATE"
        if "PENDING" in text: return "ACCOUNT - PENDING"
        return "ACCOUNT - PENDING"
    if "CASH" in text or "SALE" in text:
        if "UNPAID" in text or "DUE" in text or "OUTSTANDING" in text: return "CASH SALE - UNPAID"
        if "PARTIAL" in text: return "CASH SALE – PARTIAL"
        if "PAID" in text: return "CASH SALE - PAID"
    if "OVER LIMIT" in text: return "ACCOUNT - OVER LIMIT"
    if "UP TO DATE" in text: return "ACCOUNT - UP TO DATE"
    if "PARTIAL" in text: return "CASH SALE – PARTIAL"
    if "PAID" in text: return "CASH SALE - PAID"
    if "UNPAID" in text or "DUE" in text: return "CASH SALE - UNPAID"
    return None

def main():
    comments_file = "Released_Jobs_Comments.xlsx"
    sales_file = "Sales Order Report - 03.2026.xlsx"
    
    print("--- Reading Released_Jobs_Comments.xlsx ---")
    df = pd.read_excel(comments_file)
    df_valid = df[df['Text Box Contents'] != 'NO TEXT BOXES FOUND']
    
    comments_map = {}
    for _, row in df_valid.iterrows():
        job = str(row['Guessed Order Number']).strip()
        txt = str(row['Text Box Contents']).strip().replace('\n', ' ')
        comments_map[job] = txt
        
    print(f"\nFound {len(comments_map)} Jobs/SOs with text boxes. Simulating Mapping:")
    for job, txt in list(comments_map.items())[:15]:
        mapped = map_comment_to_status(txt)
        print(f"[{job}] -> '{txt[:35]}...' -> Mapped to: '{mapped}'")
        
    print("\n--- Reading Sales Order Report ---")
    
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
    shutil.copy2(sales_file, temp_file)
    
    wb = openpyxl.load_workbook(temp_file, data_only=True)
    date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    sheets = [s for s in wb.sheetnames if date_pattern.match(s)]
    sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
    latest_sheet = sheets[-1]
    
    ws = wb[latest_sheet]
    print(f"Checking sheet: {latest_sheet}")
    
    matches_found = 0
    mismatches = 0
    for row_idx in range(3, ws.max_row + 1):
        so = str(ws.cell(row=row_idx, column=4).value).strip()
        job = str(ws.cell(row=row_idx, column=3).value).strip()
        pay = str(ws.cell(row=row_idx, column=10).value).strip()
        
        target_txt = None
        if job in comments_map: target_txt = comments_map[job]
        elif so in comments_map: target_txt = comments_map[so]
        
        if target_txt:
            expected = map_comment_to_status(target_txt)
            if pay != "None" and pay != expected:
                print(f"WARNING: [Row {row_idx}] Job '{job}' / SO '{so}' | Excel says: '{pay}' | Rule expects: '{expected}' | Raw Comment: '{target_txt[:40]}'")
                mismatches += 1
            else:
                matches_found += 1
                
    print(f"\nDone. Processed {matches_found} matching rows. Found {mismatches} mismatches.")
    
    wb.close()
    os.remove(temp_file)

if __name__ == '__main__':
    main()
