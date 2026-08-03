import openpyxl
import os

files_to_clean = ["Contract register 2026.xlsx", "Contract register 2025.xlsx"]

for file in files_to_clean:
    if not os.path.exists(file):
        continue
    print(f"Cleaning {file}...", flush=True)
    try:
        wb = openpyxl.load_workbook(file)
        for sheet in wb.sheetnames:
            ws = wb[sheet]
            print(f"  Sheet: {sheet}, Max Row: {ws.max_row}, Max Col: {ws.max_column}", flush=True)
            
            # Find the actual last row with data
            actual_max_row = 1
            for row in range(ws.max_row, 0, -1):
                has_data = False
                for col in range(1, min(ws.max_column + 1, 100)): # Check first 100 cols max
                    if ws.cell(row=row, column=col).value is not None:
                        has_data = True
                        break
                if has_data:
                    actual_max_row = row
                    break
            
            print(f"  Actual data ends at row {actual_max_row}", flush=True)
            
            if ws.max_row > actual_max_row:
                diff = ws.max_row - actual_max_row
                print(f"  Deleting {diff} empty rows...", flush=True)
                ws.delete_rows(actual_max_row + 1, diff)
            
        print(f"Saving cleaned {file}...", flush=True)
        wb.save(file)
        print(f"Done cleaning {file}.", flush=True)
    except Exception as e:
        print(f"Error cleaning {file}: {e}", flush=True)

print("Cleanup script finished.", flush=True)
