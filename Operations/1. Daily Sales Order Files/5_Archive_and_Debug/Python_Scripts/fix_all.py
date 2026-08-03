import openpyxl

file_path = "Sales Order Report - 03.2026.xlsx"
try:
    wb = openpyxl.load_workbook(file_path)
    sheet_name = "23.03.2026"
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        # Pull values (+1 shift applied since data starts at Row 3 instead of 2)
        val_f26 = ws["F26"].value
        val_f27 = ws["F27"].value
        val_f28 = ws["F28"].value
        val_g27 = ws["G27"].value
        val_g28 = ws["G28"].value
        
        # Inject values 
        ws["I25"].value = val_f26
        ws["F26"].value = val_f27
        ws["J25"].value = val_g27
        ws["I26"].value = val_f28
        ws["J26"].value = val_g28
        
        # Clear out remaining values that shifted down
        ws["F27"].value = None
        ws["G27"].value = None
        ws["F28"].value = None
        ws["G28"].value = None
        
        wb.save(file_path)
        print("Successfully applied comprehensive data sweep fixes to anomalous split rows (Offset +1).")
    else:
        print("Sheet not found.")
except Exception as e:
    print(e)
