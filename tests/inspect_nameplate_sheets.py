from pathlib import Path
import openpyxl

p = Path('2_Source_Data/raw_sources/NAME PLATE PROCEDURE.xlsx')
wb = openpyxl.load_workbook(p, data_only=True)
for name in ['Info+Data Entry Form', 'Name Plate', 'Test Sheet']:
    if name in wb.sheetnames:
        ws = wb[name]
        print('\n=== Sheet:', name, 'rows', ws.max_row, 'cols', ws.max_column)
        for r in range(1, min(40, ws.max_row) + 1):
            row = [ws.cell(r, c).value for c in range(1, min(40, ws.max_column) + 1)]
            if any(cell is not None for cell in row):
                print(r, row)
    else:
        print('Sheet not found:', name)
