from pathlib import Path
import sys
sys.path.insert(0, str(Path('4_Scripts/backend').resolve()))
from connection_lookup import suggest_connection
import openpyxl

p = Path('2_Source_Data/raw_sources/NAME PLATE PROCEDURE.xlsx')
wb = openpyxl.load_workbook(p, data_only=True)
ws = wb['NamePlateProc']
expected = []
for r in range(15, 19):
    v = ws.cell(r, 1).value
    s = ws.cell(r, 2).value
    kw = ws.cell(r, 3).value
    if v and s and kw:
        expected.append((v, s, kw, 'STAR'))
for r in range(15, 19):
    v = ws.cell(r, 5).value
    s = ws.cell(r, 6).value
    kw = ws.cell(r, 7).value
    if v and s and kw:
        expected.append((v, s, kw, 'DELTA 525 MAX'))
for r in range(22, 26):
    v = ws.cell(r, 1).value
    s = ws.cell(r, 2).value
    kw = ws.cell(r, 3).value
    if v and s and kw:
        expected.append((v, s, kw, 'DELTA 380 MIN'))
for r in range(22, 26):
    v = ws.cell(r, 5).value
    s = ws.cell(r, 6).value
    kw = ws.cell(r, 7).value
    if v and s and kw:
        expected.append((v, s, kw, 'DELTA 220 MAX'))
print('Testing rules...')
for v, s, kw, kind in expected:
    poles = int(str(s).split()[0])
    voltage = int(str(v).replace('V', ''))
    motor_kw = float(str(kw).replace('kW', ''))
    conn, err = suggest_connection(voltage, poles, motor_kw)
    print(f'{v} {s} {kw} => {conn} ({kind}) err={err}')
