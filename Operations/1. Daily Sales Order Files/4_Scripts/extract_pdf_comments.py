import os
import re
import pdfplumber
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Side, Font, PatternFill

# --- Configuration ---
RELEASED_JOBS_DIR = r"C:\Users\Fan Movement\OneDrive - Fan Movement (Pty) Ltd\Zinziso Xhallie's files - Sales\Sales Contracts & Register\Contract Register\Released Jobs"
OUTPUT_EXCEL = "5_Archive_and_Debug\\Released_Jobs_Comments.xlsx"

def extract_comments_from_pdf(pdf_path):
    """
    Extracts text contents from all annotations in a PDF.
    """
    comments = []
    try:
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                if page.annots:
                    for annot in page.annots:
                        # We are looking for text boxes / sticky notes (FreeText, Text, etc.)
                        # As long as it has 'contents', it contains user text.
                        contents = annot.get('contents')
                        if contents:
                            # Clean up the text by removing carriage returns and excessive whitespace
                            clean_text = str(contents).strip()
                            if clean_text:
                                comments.append(clean_text)
    except Exception as e:
        print(f"Error reading PDF {os.path.basename(pdf_path)}: {e}")
        return ["ERROR READING PDF"]
    
    return comments

def format_excel_file(file_path):
    """
    Applies standard formatting to the generated Excel file.
    """
    try:
        wb = load_workbook(file_path)
        ws = wb.active
        ws.title = "Extracted Comments"
        
        # Header Style
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        # Format Headers
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
        # Format Data Rows
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
        # Column Widths
        ws.column_dimensions['A'].width = 40 # Filename
        ws.column_dimensions['B'].width = 25 # Order/Job Number
        ws.column_dimensions['C'].width = 60 # Extracted Text
        
        # View setup
        ws.freeze_panes = "A2"
        ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
        ws.page_setup.fitToWidth = 1
        ws.page_setup.fitToHeight = 0
        
        wb.save(file_path)
    except Exception as e:
        print(f"Error formatting Excel file: {e}")

def main():
    print(f"Scanning directory: {RELEASED_JOBS_DIR}")
    
    if not os.path.exists(RELEASED_JOBS_DIR):
        print(f"Error: Directory does not exist -> {RELEASED_JOBS_DIR}")
        return
        
    data_rows = []
    pdf_files = [f for f in os.listdir(RELEASED_JOBS_DIR) if f.lower().endswith('.pdf')]
    
    print(f"Found {len(pdf_files)} PDF files.")
    
    for filename in pdf_files:
        pdf_path = os.path.join(RELEASED_JOBS_DIR, filename)
        
        # Guess job/SO number from filename (everything before first underscore/space)
        guessed_job = re.split(r'[_ -]', filename)[0]
        
        comments = extract_comments_from_pdf(pdf_path)
        
        # Combine comments into a single string if multiple exist
        if comments:
            combined_comments = "\n---\n".join(comments)
        else:
            combined_comments = "NO TEXT BOXES FOUND"
            
        data_rows.append({
            "Filename": filename,
            "Guessed Order Number": guessed_job,
            "Text Box Contents": combined_comments
        })
        
    if not data_rows:
        print("No data extracted. Exiting.")
        return
        
    # Export to Excel
    print(f"Generating spreadsheet: {OUTPUT_EXCEL}")
    df = pd.DataFrame(data_rows)
    
    # We sort by filename to make it neat
    df.sort_values(by="Filename", inplace=True)
    
    df.to_excel(OUTPUT_EXCEL, index=False)
    
    # Apply formatting
    format_excel_file(OUTPUT_EXCEL)
    
    print("Done! Spreadsheet created successfully.")

if __name__ == "__main__":
    main()
