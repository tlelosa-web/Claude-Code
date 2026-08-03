import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import datetime
import re
import csv


def format_sales_report():
    """
    Formats the sales report by highlighting pending jobs in yellow and setting column widths
    """
    # --- Configuration ---
    TODAY_DT = datetime.datetime.now()
    TODAY = TODAY_DT.strftime("%d.%m.%Y")
    
    # Resolve the sales report file
    reports_dir = "3_Live_Reports"
    month_tag = TODAY_DT.strftime("%m.%Y")
    preferred = os.path.join(reports_dir, f"Ops Sales Order Report - {month_tag}.xlsx")
    legacy = os.path.join(reports_dir, f"Sales Order Report - {month_tag}.xlsx")

    if os.path.exists(preferred):
        SALES_REPORT_FILE = preferred
    elif os.path.exists(legacy):
        SALES_REPORT_FILE = legacy
    else:
        # Always target the current month report file.
        SALES_REPORT_FILE = preferred

    if not os.path.exists(SALES_REPORT_FILE):
        print(f"ERROR: {SALES_REPORT_FILE} not found.")
        return

    # Open the workbook
    print(f"Opening {SALES_REPORT_FILE}...")
    try:
        wb = openpyxl.load_workbook(SALES_REPORT_FILE)
    except PermissionError:
        print("ERROR: File is open. Please close it and try again.")
        return

    # Find the sheet with today's date
    date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    date_sheets = [s for s in wb.sheetnames if date_pattern.match(s)]

    if not date_sheets:
        print("No valid date sheets found.")
        return

    # Use the most recent date sheet
    date_sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
    ws = wb[date_sheets[-1]]
    print(f"Formatting sheet: {ws.title}")

    # Define yellow fill for pending jobs
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    # Define borders
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))

    # Locate the Status column (column I)
    status_col = None
    for col_idx in range(1, 20):  # Check first 20 columns
        val = ws.cell(row=2, column=col_idx).value
        if val == "Status":
            status_col = col_idx
            break

    if not status_col:
        print("Could not find 'Status' column.")
        return

    # Locate all data rows and highlight pending jobs
    for row_idx in range(3, ws.max_row + 1):  # Start from row 3 (data rows)
        status_cell = ws.cell(row=row_idx, column=status_col)
        if status_cell.value and str(status_cell.value).strip().upper() == "PENDING":
            # Highlight entire row in yellow
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = yellow_fill
                cell.border = thin_border

    # Set column J width to 23.3
    ws.column_dimensions['J'].width = 23.3
    
    # Adjust other column widths if needed
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    # J is set to 23.3 above
    ws.column_dimensions['K'].width = 50
    ws.column_dimensions['L'].width = 12

    # Add footnote with recommended print range and suggested date for next update
    last_row = ws.max_row
    footnote_row = last_row + 3
    
    # Add footnote text
    footnote_cell = ws.cell(row=footnote_row, column=1, value="Recommended Print Range: A1:L{}".format(footnote_row+2))
    footnote_cell.font = Font(italic=True, size=10)
    
    # Suggested Date for Next Update:
    # Find the earliest order date (col A) among open (non-Pending) orders,
    # then subtract 2 days.  Falls back to tomorrow if no qualifying date is found.
    earliest_order = None
    for data_row_idx in range(3, last_row + 1):
        status_val = ws.cell(row=data_row_idx, column=9).value  # Column I - Status
        if status_val and str(status_val).strip().lower() == "pending":
            continue
        date_val = ws.cell(row=data_row_idx, column=1).value   # Column A - Order Date
        if isinstance(date_val, datetime.datetime):
            if earliest_order is None or date_val < earliest_order:
                earliest_order = date_val

    if earliest_order is not None:
        next_update_date = (earliest_order - datetime.timedelta(days=2)).strftime("%d.%m.%Y")
    else:
        # Fallback: suggest tomorrow if no open orders with an order date exist
        next_update_date = (TODAY_DT + datetime.timedelta(days=1)).strftime("%d.%m.%Y")

    next_update_cell = ws.cell(row=footnote_row+1, column=1, value="Suggested Date for Next Update: {}".format(next_update_date))
    next_update_cell.font = Font(italic=True, size=10)
    
    # Set print area
    ws.print_area = f"A1:L{footnote_row+2}"

    # Save the workbook
    print("Saving workbook...")
    try:
        wb.save(SALES_REPORT_FILE)
        print("Sales report formatting completed successfully!")
    except PermissionError:
        print("ERROR: Could not save file. It might be open.")


def format_csv_files():
    """
    Creates formatted versions of the CSV files with recommended print ranges and footnotes
    """
    import shutil
    
    # Define file paths
    source_report = "2_Source_Data\\CustomerSalesOrdersReport.csv"
    source_by_customer = "2_Source_Data\\CustomerSalesOrdersByCustomer.csv"
    
    # Define output paths for formatted Excel versions
    output_report = "5_Archive_and_Debug\\Formatted_CustomerSalesOrdersReport.xlsx"
    output_by_customer = "5_Archive_and_Debug\\Formatted_CustomerSalesOrdersByCustomer.xlsx"
    
    # Process CustomerSalesOrdersReport.csv
    if os.path.exists(source_report):
        # Read CSV
        df_report = pd.read_csv(source_report)
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Sales Orders Report"
        
        # Add headers
        for col_idx, col_name in enumerate(df_report.columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Add data
        for row_num, (idx, row) in enumerate(df_report.iterrows(), 2):
            for col_idx, value in enumerate(row.values, 1):
                ws.cell(row=row_num, column=col_idx, value=value)
        
        # Apply borders
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=1, max_row=len(df_report)+1, max_col=len(df_report.columns)):
            for cell in row:
                cell.border = thin_border
        
        # Highlight pending jobs in yellow
        status_col_idx = None
        for col_idx in range(1, len(df_report.columns) + 1):
            if ws.cell(row=1, column=col_idx).value == 'Status':
                status_col_idx = col_idx
                break
        
        if status_col_idx:
            for row_idx in range(2, len(df_report) + 2):  # Start from row 2 (data rows)
                status_cell = ws.cell(row=row_idx, column=status_col_idx)
                if status_cell.value and str(status_cell.value).strip().upper() == "PENDING":
                    # Highlight entire row in yellow
                    for col_idx in range(1, len(df_report.columns) + 1):
                        cell = ws.cell(row=row_idx, column=col_idx)
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
        
        # Add footnote with recommended print range and suggested lookback date
        last_row = len(df_report) + 1
        footnote_row = last_row + 3
        
        # Add footnote text
        footnote_cell = ws.cell(row=footnote_row, column=1, value="Recommended Print Range: A1:{}{}".format(
            openpyxl.utils.get_column_letter(len(df_report.columns)), footnote_row+2))
        footnote_cell.font = Font(italic=True, size=10)
        
        # Add suggested lookback date based on earliest date in the Ops Sales Order Report
        TODAY_DT = datetime.datetime.now()
        reports_dir = "3_Live_Reports"
        month_tag = TODAY_DT.strftime("%m.%Y")
        ops_report_path = os.path.join(reports_dir, f"Ops Sales Order Report - {month_tag}.xlsx")
        if os.path.exists(ops_report_path):
            try:
                df_ops = pd.read_excel(ops_report_path, header=1)
                # Get column A (Date) and find the earliest date
                dates = pd.to_datetime(df_ops.iloc[:, 0], errors='coerce').dropna()
                
                # Filter out future dates (anything after today) to avoid footer/summary dates
                today = datetime.datetime.now()
                past_dates = dates[dates <= today]
                
                if len(past_dates) > 0:
                    earliest_date = past_dates.min()
                    # Safe lookback date is the first day of the month of the earliest order
                    safe_lookback_date = earliest_date.replace(day=1)
                    suggested_lookback_date = safe_lookback_date.strftime("%d.%m.%Y")
                    next_update_cell = ws.cell(row=footnote_row+1, column=1, value="Suggested Lookback Date: {}".format(suggested_lookback_date))
                    next_update_cell.font = Font(italic=True, size=10)
            except Exception as e:
                print(f"Warning: Could not determine lookback date from Ops report: {e}")
        
        # Set print area
        ws.print_area = f"A1:{openpyxl.utils.get_column_letter(len(df_report.columns))}{footnote_row+2}"
        
        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limit max width to 50
            ws.column_dimensions[column].width = adjusted_width
        
        # Save the workbook
        wb.save(output_report)
        print("Formatted CustomerSalesOrdersReport created successfully!")
    
    # Process CustomerSalesOrdersByCustomer.csv - This is a special format that needs special handling
    if os.path.exists(source_by_customer):
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Customer Sales Orders By Customer"
        
        # Read the special CSV format line by line
        with open(source_by_customer, 'r', encoding='utf-8') as f:
            lines = f.readlines()
        
        # Process the file to extract data in a structured way
        # This file has a complex structure with customer groups, orders, and line items
        current_row = 1
        headers_written = False
        
        # First, write the header row (which is on line 4)
        header_line = lines[3].strip()  # The header is on line 4 (0-indexed as 3)
        # Clean and parse the header
        header_fields = []
        for field in header_line.split(','):
            field = field.strip().strip('"')
            if field:
                header_fields.append(field)
        
        # Write headers
        for col_idx, header in enumerate(header_fields, 1):
            cell = ws.cell(row=current_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        current_row += 1
        
        # Now process the rest of the file
        for i in range(4, len(lines)):  # Start from line 5 (0-indexed as 4)
            line = lines[i].strip()
            if not line:
                continue
                
            # Split the line by commas, accounting for quoted fields
            fields = []
            field = ''
            in_quotes = False
            for char in line:
                if char == '"' and not in_quotes:
                    in_quotes = True
                elif char == '"' and in_quotes:
                    in_quotes = False
                elif char == ',' and not in_quotes:
                    fields.append(field.strip().strip('"'))
                    field = ''
                else:
                    field += char
            if field:
                fields.append(field.strip().strip('"'))
            
            # Ensure fields list has same length as header for proper column alignment
            while len(fields) < len(header_fields):
                fields.append("")
            
            # Write the fields to the worksheet
            for col_idx, field in enumerate(fields, 1):
                ws.cell(row=current_row, column=col_idx, value=field)
            
            # Check if this row has a Status column with "Pending"
            # Find status column index
            status_col_idx = None
            for col_idx, header in enumerate(header_fields, 1):
                if header == "Status":
                    status_col_idx = col_idx
                    break
            
            if status_col_idx and status_col_idx <= len(fields):
                status_value = fields[status_col_idx - 1] if status_col_idx <= len(fields) else ""
                if status_value and status_value.strip().upper() == "PENDING":
                    # Highlight entire row in yellow
                    for col_idx in range(1, len(fields) + 1):
                        cell = ws.cell(row=current_row, column=col_idx)
                        cell.fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
            
            current_row += 1
        
        # Apply borders to all cells
        thin_border = Border(left=Side(style='thin'), 
                             right=Side(style='thin'), 
                             top=Side(style='thin'), 
                             bottom=Side(style='thin'))
        
        for row in ws.iter_rows(min_row=1, max_row=current_row-1, max_col=len(header_fields)):
            for cell in row:
                cell.border = thin_border
        
        # Add footnote with recommended print range and suggested date for next update
        footnote_row = current_row + 2
        
        # Add footnote text
        footnote_cell = ws.cell(row=footnote_row, column=1, value="Recommended Print Range: A1:{}{}".format(
            openpyxl.utils.get_column_letter(len(header_fields)), footnote_row+2))
        footnote_cell.font = Font(italic=True, size=10)
        
        # Add suggested date for next update (tomorrow)
        next_update_date = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%d.%m.%Y")
        next_update_cell = ws.cell(row=footnote_row+1, column=1, value="Suggested Date for Next Update: {}".format(next_update_date))
        next_update_cell.font = Font(italic=True, size=10)
        
        # Set print area
        ws.print_area = f"A1:{openpyxl.utils.get_column_letter(len(header_fields))}{footnote_row+2}"
        
        # Set column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Limit max width to 50
            ws.column_dimensions[column].width = adjusted_width
        
        # Save the workbook
        wb.save(output_by_customer)
        print("Formatted CustomerSalesOrdersByCustomer created successfully!")


def main():
    print("Formatting reports...")
    format_sales_report()
    print("All formatting completed!")


if __name__ == "__main__":
    main()