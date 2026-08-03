import pandas as pd
import pdfplumber
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
import os
import datetime
import re
import csv
import sys
import time
import tempfile
import shutil

import calendar
import glob

def _next_month_name(dt):
    year = dt.year + (1 if dt.month == 12 else 0)
    month = 1 if dt.month == 12 else (dt.month + 1)
    return datetime.datetime(year, month, 1).strftime("%B")

def resolve_sales_report_file(dt):
    reports_dir = "3_Live_Reports"
    month_tag = dt.strftime("%m.%Y")
    preferred = os.path.join(reports_dir, f"Ops Sales Order Report - {month_tag}.xlsx")
    legacy = os.path.join(reports_dir, f"Sales Order Report - {month_tag}.xlsx")

    if os.path.exists(preferred):
        return preferred

    if os.path.exists(legacy):
        try:
            os.rename(legacy, preferred)
            return preferred
        except Exception:
            return legacy

    # Always use the current month output file when starting a new month.
    return preferred


def archive_old_sales_reports(dt):
    reports_dir = "3_Live_Reports"
    archive_dir = "5_Archive_and_Debug"
    os.makedirs(archive_dir, exist_ok=True)
    current_month_tag = dt.strftime("%m.%Y")

    for pattern in (
        os.path.join(reports_dir, "Ops Sales Order Report - *.xlsx"),
        os.path.join(reports_dir, "Sales Order Report - *.xlsx"),
    ):
        for old_path in glob.glob(pattern):
            if current_month_tag not in os.path.basename(old_path):
                basename = os.path.basename(old_path)
                target = os.path.join(archive_dir, basename)
                if os.path.exists(target):
                    base, ext = os.path.splitext(basename)
                    count = 1
                    while os.path.exists(os.path.join(archive_dir, f"{base}_{count}{ext}")):
                        count += 1
                    target = os.path.join(archive_dir, f"{base}_{count}{ext}")
                try:
                    shutil.move(old_path, target)
                    print(f"Archived old report: {basename}", flush=True)
                except Exception as e:
                    print(f"Warning: Could not archive {basename}: {e}", flush=True)


# --- Configuration ---
TODAY_DT = datetime.datetime.now()
TODAY = TODAY_DT.strftime("%d.%m.%Y")
THIS_MONTH_NAME = TODAY_DT.strftime("%B")
NEXT_MONTH_NAME = _next_month_name(TODAY_DT)
archive_old_sales_reports(TODAY_DT)
SALES_REPORT_FILE = resolve_sales_report_file(TODAY_DT)
CSV_SOURCE_FILE = "2_Source_Data\\CustomerSalesOrdersReport.csv"
CSV_DETAILS_FILE = "2_Source_Data\\CustomerSalesOrdersByCustomer.csv"
CONTRACT_REGISTER_SHEET = "Contracts 2026"
RELEASED_JOBS_DIR = r"C:\Users\Fan Movement\OneDrive - Fan Movement (Pty) Ltd\Zinziso Xhallie's files - Sales\Sales Contracts & Register\Contract Register\Released Jobs"

# --- Helper Functions ---

def clean_currency(val):
    """Converts currency string (e.g., 'R 1 234.56') to float."""
    if pd.isna(val) or val == "":
        return 0.0
    val_str = str(val).replace("R", "").replace(" ", "").replace(",", "")
    try:
        return float(val_str)
    except ValueError:
        return 0.0

def parse_date(date_str):
    """Parses DD/MM/YYYY string to datetime object."""
    if pd.isna(date_str) or date_str == "":
        return None
    try:
        return datetime.datetime.strptime(str(date_str).strip(), "%d/%m/%Y")
    except ValueError:
        return None

def to_initials(name):
    """Convert a full name to uppercase initials. e.g. 'Simon Mokgola' -> 'SM'"""
    if not name or name.strip() == "":
        return ""
    return "".join(word[0].upper() for word in name.strip().split() if word)

def load_contract_mapping():
    """
    Loads Contract Registers and returns a mapping:
    SO Number (Document No.) -> Contract Number (Job Number)
    
    IMPORTANT: This function requires CONTRACT_REGISTER_PATH environment variable to be set.
    This ensures the file is only accessed during pipeline execution.
    """
    # Get the contract register path from environment variable only (no fallback)
    contract_register_path = os.getenv("CONTRACT_REGISTER_PATH")
    
    if not contract_register_path:
        raise EnvironmentError(
            "CONTRACT_REGISTER_PATH environment variable is not set. "
            "The contract register can only be accessed when the pipeline is run via RUN_DAILY_PIPELINE.bat"
        )
    
    contract_registers = [contract_register_path]
    
    mapping = {}
    print("Loading Contract Registers...")
    for reg_file in contract_registers:
        if not os.path.exists(reg_file):
            print(f"Warning: {reg_file} not found.", flush=True)
            continue
        
        try:
            # Create a temp copy to bypass file lock
            temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
            shutil.copy2(reg_file, temp_file)
            
            # Read the Contracts 2026 tab with header starting on row 4.
            try:
                df = pd.read_excel(temp_file, sheet_name=CONTRACT_REGISTER_SHEET, header=3, nrows=5000)
            except Exception as e:
                print(f"Warning: could not read sheet {CONTRACT_REGISTER_SHEET} from {reg_file}: {e}", flush=True)
                df = pd.read_excel(temp_file, header=3, nrows=5000)
            
            # Ensure columns exist
            if 'Document No.' in df.columns and 'Contract Number' in df.columns:
                # Iterate rows
                for _, row in df.iterrows():
                    so_num = str(row['Document No.']).strip()
                    job_num = str(row['Contract Number']).strip()
                    
                    if so_num and so_num != 'nan' and job_num and job_num != 'nan':
                        mapping[so_num] = job_num
                        
            # Also check 'Order Number' just in case some SOs are there
            if 'Order Number' in df.columns and 'Contract Number' in df.columns:
                 for _, row in df.iterrows():
                    order_num = str(row['Order Number']).strip()
                    job_num = str(row['Contract Number']).strip()
                    
                    if order_num and order_num != 'nan' and job_num and job_num != 'nan':
                        # Avoid overwriting if SO key already exists from Document No?
                        # Maybe safer to add both if different keys
                        if order_num not in mapping:
                            mapping[order_num] = job_num
            
            # Cleanup temp file
            try:
                os.remove(temp_file)
            except OSError:
                pass
                            
        except Exception as e:
            print(f"Error loading {reg_file}: {e}", flush=True)
            
    print(f"Loaded {len(mapping)} contract mappings.", flush=True)
    return mapping

def load_contract_details_mapping():
    """Returns SO/job number -> job description from the contract register."""
    contract_register_path = os.getenv("CONTRACT_REGISTER_PATH")

    if not contract_register_path:
        return {}

    mapping = {}
    if not os.path.exists(contract_register_path):
        return mapping

    try:
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx").name
        shutil.copy2(contract_register_path, temp_file)

        try:
            df = pd.read_excel(temp_file, sheet_name=CONTRACT_REGISTER_SHEET, header=3, nrows=5000)
        except Exception as e:
            print(f"Warning: could not read sheet {CONTRACT_REGISTER_SHEET} from {contract_register_path}: {e}", flush=True)
            df = pd.read_excel(temp_file, header=3, nrows=5000)

        required = {"Contract Number", "Job Description", "Qty"}
        if required.issubset(df.columns):
            key_columns = [col for col in ("Document No.", "Order Number", "Contract Number") if col in df.columns]
            for _, row in df.iterrows():
                description = str(row["Job Description"]).strip()
                if not description or description == "nan":
                    continue

                qty = str(row["Qty"]).strip()
                details = description
                if qty and qty != "nan":
                    details = f"{description} (Qty: {qty})"

                for key_col in key_columns:
                    key = str(row[key_col]).strip()
                    if key and key != "nan" and key not in mapping:
                        mapping[key] = details

        try:
            os.remove(temp_file)
        except OSError:
            pass

    except Exception as e:
        print(f"Warning: could not load contract details: {e}", flush=True)

    print(f"Loaded contract details for {len(mapping)} keys.", flush=True)
    return mapping

def load_csv_details_mapping():
    """
    Parses CustomerSalesOrdersByCustomer.csv to extract details per SO.
    Returns dict: SO Number -> Details String
    """
    mapping = {}
    print("Loading CSV Details...", flush=True)
    if not os.path.exists(CSV_DETAILS_FILE):
        print(f"Warning: {CSV_DETAILS_FILE} not found.", flush=True)
        return mapping

    try:
        with open(CSV_DETAILS_FILE, 'r', encoding='utf-8') as f:
            # Skip until header found
            while True:
                line = f.readline()
                if not line:
                    break
                if "Reference" in line and "Description" in line:
                    break
            
            reader = csv.reader(f)
            current_so = None
            current_items = []
            
            for row in reader:
                if not row or len(row) < 4:
                    continue
                
                # Check column indices based on inspection
                # Col 0: Name/Total, Col 1: Date/Ref, Col 2: Status, Col 3: Description, Col 4: Qty
                col0 = row[0].strip()
                col1 = row[1].strip() # Reference (SO number)
                col3 = row[3].strip() # Description
                col4 = row[4].strip() if len(row) > 4 else "" # Qty
                
                # Detect new SO line
                if col1.startswith("SO") and not col0.startswith("Total"):
                    # Save previous SO
                    if current_so and current_items:
                        mapping[current_so] = ", ".join(current_items)
                    
                    current_so = col1
                    current_items = []
                
                # Detect Item line (empty Reference, has Description)
                elif not col1 and col3 and current_so:
                    if col3.lower() != "nan" and col3 != "":
                        qty_str = f" (Qty: {col4})" if (col4 and col4 != "0") else ""
                        current_items.append(f"{col3}{qty_str}")
                
                # Detect End of SO
                elif col0.startswith("Total for SO"):
                    if current_so and current_items:
                        mapping[current_so] = ", ".join(current_items)
                    current_so = None
                    current_items = []

            # Capture last one
            if current_so and current_items:
                mapping[current_so] = ", ".join(current_items)

    except Exception as e:
        print(f"Error loading CSV details: {e}", flush=True)
        
    print(f"Loaded details for {len(mapping)} SOs.", flush=True)
    return mapping

RELEASED_JOBS_CACHE = None

def get_released_jobs():
    """Caches the list of PDFs in the Released Jobs directory for faster execution."""
    global RELEASED_JOBS_CACHE
    if RELEASED_JOBS_CACHE is None:
        RELEASED_JOBS_CACHE = []
        try:
            RELEASED_JOBS_CACHE = [f for f in os.listdir(RELEASED_JOBS_DIR) if f.lower().endswith('.pdf')]
        except OSError:
            pass
    return RELEASED_JOBS_CACHE

def check_released_job_status(job_number):
    """
    Checks if a PDF exists for the Job Number in Released Jobs.
    Returns (status_text, payment_status_text)
    """
    if not job_number or job_number == "Pending":
        return "Pending", ""

    found_pdf = None
    for filename in get_released_jobs():
        if filename.startswith(job_number):
            found_pdf = os.path.join(RELEASED_JOBS_DIR, filename)
            break

    if found_pdf:
        payment_status = extract_payment_status(found_pdf)
        return "In Progress", payment_status
    
    return "Pending", ""

def check_released_job_status_by_so(so_num):
    """
    Secondary check: Searches Released Jobs folder by SO number.
    Returns (status_text, payment_status_text, guessed_job_number)
    """
    if not so_num or so_num == "Pending":
        return "Pending", "", "Pending"
        
    for filename in get_released_jobs():
        if so_num in filename:
            found_pdf = os.path.join(RELEASED_JOBS_DIR, filename)
            payment_status = extract_payment_status(found_pdf)
            # Try to guess job number from filename (assuming it's before the first underscore or space)
            guessed_job = re.split(r'[_ -]', filename)[0]
            return "In Progress", payment_status, guessed_job
            
    return "Pending", "", "Pending"

def extract_payment_status(pdf_path):
    """
    Scans PDF for payment status using a hybrid approach.
    1. Checks for explicit PDF comments (annotations).
    2. Checks the PDF filename to determine payment type (Cash Sale vs. Account).
    3. Scans the visible text for the 'BALANCE DUE' amount to determine the status.
    """
    try:
        # --- Step 1: Check for comments/annotations first ---
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                if page.annots:
                    for annot in page.annots:
                        if annot.get('subtype') == 'Text' and annot.get('contents'):
                            comment_text = annot.get('contents', '').upper()
                            if "CASH SALE" in comment_text or "ACCOUNT" in comment_text:
                                return comment_text.splitlines()[0].strip() # Return comment as highest truth source

            # --- Step 2: Determine payment type from filename ---
            filename_upper = os.path.basename(pdf_path).upper()
            is_cash_sale_from_filename = "CASH SALE" in filename_upper

            # --- Step 3: Extract text and find balance due ---
            text = ""
            for page in pdf.pages:
                text += page.extract_text(x_tolerance=1) or ""
        
            text_upper = text.upper()

            # Regex to find balance due. Handles amounts like "R 1,234.56"
            balance_match = re.search(r"BALANCE DUE\s*R\s*([\d\s,]+\.\d{2})", text_upper)
            balance_due = 0.0

            if balance_match:
                # Clean up the matched string (remove spaces and commas) and convert to float
                balance_str = balance_match.group(1).replace(" ", "").replace(",", "")
                balance_due = float(balance_str)

            # --- Step 4: Combine filename type and balance to determine status ---
            if is_cash_sale_from_filename:
                # It's a Cash Sale according to the filename
                if balance_due > 0:
                    return "CASH SALE - UNPAID"
                else:
                    return "CASH SALE - PAID" # Balance is 0 or not found, assume paid
            else:
                # It's an Account sale by default
                if balance_due > 0:
                    return "ACCOUNT - OUTSTANDING"
                else:
                    return "ACCOUNT - UP TO DATE" # Balance is 0 or not found, assume up-to-date
                    
    except Exception as e:
        print(f"Error reading PDF {os.path.basename(pdf_path)}: {e}", flush=True)
        return "PDF READ ERROR"

def get_previous_day_statuses():
    """
    Reads the most recent sheet in the Sales Report to map SO Number -> {Status, Payment Status}.
    This preserves manual status updates (like 'Ready - Dispatch' or 'Manual Paid') across days.
    """
    mapping = {}
    if not os.path.exists(SALES_REPORT_FILE):
        return mapping
        
    try:
        wb = openpyxl.load_workbook(SALES_REPORT_FILE, data_only=True)
        # Find the latest date sheet
        date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
        sheets = [s for s in wb.sheetnames if date_pattern.match(s)]
        
        if not sheets:
            return mapping
            
        sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
        latest_sheet = sheets[-1]
        
        # Don't read today's sheet if we are rerunning the script on the same day
        if latest_sheet == TODAY and len(sheets) > 1:
             latest_sheet = sheets[-2]
        elif latest_sheet == TODAY:
             return mapping # Only one sheet and it's today's

        ws = wb[latest_sheet]
        
        # Find column indices (1-based) by scanning top 3 rows
        so_col = None
        status_col = None
        pay_col = None
        data_start_row = 2
        
        for r_idx in range(1, 4):
            headers = {cell.value: cell.column for cell in ws[r_idx] if cell.value}
            if 'SO Number' in headers and 'Status' in headers:
                so_col = headers['SO Number']
                status_col = headers['Status']
                pay_col = headers.get('Payment Status')
                data_start_row = r_idx + 1
                break
        
        if so_col and status_col:
            for row in ws.iter_rows(min_row=data_start_row):
                so_val = row[so_col - 1].value
                status_val = row[status_col - 1].value
                pay_val = row[pay_col - 1].value if pay_col else ""
                if so_val and str(so_val).strip():
                    mapping[str(so_val).strip()] = {
                        'status': str(status_val).strip() if status_val else "Pending",
                        'payment_status': str(pay_val).strip() if pay_val else ""
                    }
                    
        return mapping
    except Exception as e:
        print(f"Error reading previous statuses: {e}", flush=True)
        return mapping

# --- Main Logic ---

def generate_report_data():
    """
    Builds the report data rows by merging sources.
    """
    print("Reading Primary Source: CustomerSalesOrdersReport.csv", flush=True)
    try:
        # Check if first row is a header or data
        with open(CSV_SOURCE_FILE, 'r') as f:
            first_line = f.readline()
            if "Document No." in first_line or "Date" in first_line:
                main_df = pd.read_csv(CSV_SOURCE_FILE)
            else:
                main_df = pd.read_csv(CSV_SOURCE_FILE, skiprows=1)
    except Exception as e:
        print(f"Error reading {CSV_SOURCE_FILE}: {e}", flush=True)
        return None

    # Load mappings
    contract_map = load_contract_mapping()
    details_map = load_csv_details_mapping()
    contract_details_map = load_contract_details_mapping()
    previous_statuses = get_previous_day_statuses()

    report_rows = []
    
    print("Processing rows...", flush=True)
    for _, row in main_df.iterrows():
        # Filter out Invoiced orders
        csv_status = str(row.get('Status', '')).strip()
        if csv_status.lower() == 'invoiced':
            continue

        # Extract base data
        date_val = parse_date(row.get('Date'))
        del_date = parse_date(row.get('Delivery Date'))
        
        # If delivery date has passed, move it to today
        if del_date and del_date < TODAY_DT:
            del_date = TODAY_DT

        so_num = str(row.get('Document No.', '')).strip()
        cust_ref = str(row.get('Customer Ref.', '')).strip()
        customer = str(row.get('Customer', '')).strip()
        total = clean_currency(row.get('Total'))
        sales_rep = to_initials(str(row.get('Sales Rep', '')).strip())
        
        # Derived Data
        job_num = contract_map.get(so_num, "Pending")
        # If not found by SO, try Customer Ref?
        if job_num == "Pending" and cust_ref in contract_map:
            job_num = contract_map[cust_ref]

        details = (
            details_map.get(so_num, "")
            or contract_details_map.get(so_num, "")
            or contract_details_map.get(job_num, "")
        )
        
        # Status & Payment from PDF
        csv_status = str(row.get('Status', '')).strip()
        status = "Pending"
        payment_status = ""
        
        if so_num in previous_statuses:
            prev_data = previous_statuses[so_num]
            prev_status = prev_data.get('status', 'Pending')
            if prev_status != "Pending" and prev_status != "None":
                status = prev_status
            prev_payment_status = prev_data.get('payment_status', '')
            if prev_payment_status and prev_payment_status != "None":
                payment_status = prev_payment_status

        if job_num != "Pending":
            pdf_status, payment_status_val = check_released_job_status(job_num)
            if pdf_status == "In Progress":
                # Only overwrite with In Progress if we don't already have a more advanced status
                if status == "Pending":
                    status = "In Progress"
                # Only overwrite payment status if PDF provides a value
                if payment_status_val:
                    payment_status = payment_status_val
        else:
            # Secondary check
            pdf_status, payment_status_val, guessed_job = check_released_job_status_by_so(so_num)
            if pdf_status == "In Progress":
                if status == "Pending":
                    status = "In Progress"
                if payment_status_val:
                    payment_status = payment_status_val
                job_num = guessed_job

        # Preserve "Ready - Dispatch" from CSV if it exists (takes highest precedence from source)
        if "Ready" in csv_status and "Dispatch" in csv_status:
            status = "Ready - Dispatch"
        
        # Construct Row (Columns A-L)
        # A: Date
        # B: Delivery Date
        # C: Job Number
        # D: SO Number
        # E: Customer Ref
        # F: Customer
        # G: Total
        # H: Sales Rep
        # I: Status
        # J: Payment Status
        # K: Details
        # L: Delivery Date (Copy of B)
        
        report_row = [
            date_val,       # A
            del_date,       # B
            job_num,        # C
            so_num,         # D
            cust_ref,       # E
            customer,       # F
            total,          # G
            sales_rep,      # H
            status,         # I
            payment_status, # J
            details,        # K
            del_date        # L
        ]
        report_rows.append(report_row)
    
    def get_sort_key(row):
        """Sort by: Column B (Delivery Date) -> Column I (Status) -> Column A (Order Date).
        Null delivery dates are sorted to the bottom.
        """
        del_date   = row[1]  # Column B
        status     = str(row[8]).lower()  # Column I
        order_date = row[0]  # Column A

        is_none = (del_date is None)
        safe_del_date   = del_date   if del_date   is not None else datetime.datetime.max
        safe_order_date = order_date if order_date is not None else datetime.datetime.min

        return (is_none, safe_del_date, status, safe_order_date)

    report_rows.sort(key=get_sort_key)
        
    return report_rows

def get_previous_comments(wb, prev_sheet_name):
    comments_map = {}
    if not prev_sheet_name or prev_sheet_name not in wb.sheetnames:
        return comments_map
    ws = wb[prev_sheet_name]
    start_row = None
    for r in range(ws.max_row, 0, -1):
        if str(ws.cell(row=r, column=11).value).strip() == "Comments":
            start_row = r + 1
            break
    if start_row:
        for r in range(start_row, ws.max_row + 1):
            so_val = ws.cell(row=r, column=4).value
            comment_val = ws.cell(row=r, column=11).value
            if so_val and comment_val:
                comments_map[str(so_val).strip()] = str(comment_val).strip()
    return comments_map

def get_previous_details(wb, prev_sheet_name):
    details_map = {}
    if not prev_sheet_name or prev_sheet_name not in wb.sheetnames:
        return details_map
    ws = wb[prev_sheet_name]
    for r in range(3, ws.max_row + 1):
        so_val = ws.cell(row=r, column=4).value
        details_val = ws.cell(row=r, column=11).value
        if not so_val or not details_val:
            continue

        so_text = str(so_val).strip()
        details_text = str(details_val).strip()
        if so_text.startswith("SO") and so_text != "SO Number" and details_text != "Comments":
            details_map[so_text] = details_text
    return details_map

def write_to_excel(data_rows):
    """Writes the data to the Excel file."""
    if not os.path.exists(SALES_REPORT_FILE):
        print(f"Creating new workbook: {SALES_REPORT_FILE}", flush=True)
        wb = openpyxl.Workbook()
        wb.remove(wb.active) # Remove default sheet
    else:
        print(f"Opening workbook: {SALES_REPORT_FILE}", flush=True)
        try:
            wb = openpyxl.load_workbook(SALES_REPORT_FILE)
        except PermissionError:
            print("ERROR: File is open. Please close it and try again.", flush=True)
            return

    # Find previous sheet for carrying forward statuses, comments, and details
    prev_sheet = None
    date_pattern = re.compile(r"^\d{2}\.\d{2}\.\d{4}$")
    sheets = [s for s in wb.sheetnames if date_pattern.match(s)]
    if sheets:
        sheets.sort(key=lambda s: datetime.datetime.strptime(s, "%d.%m.%Y"))
        if sheets[-1] == TODAY and len(sheets) > 1:
            prev_sheet = sheets[-2]
        elif sheets[-1] != TODAY:
            prev_sheet = sheets[-1]

    prev_comments = get_previous_comments(wb, prev_sheet)
    prev_details = get_previous_details(wb, prev_sheet)
    
    # Create new sheet
    sheet_name = TODAY
    if sheet_name in wb.sheetnames:
        print(f"Sheet {sheet_name} already exists. Overwriting...", flush=True)
        del wb[sheet_name]
    
    ws = wb.create_sheet(sheet_name)
    
    # Title Row (Row 1)
    ws.cell(row=1, column=1, value="SALES ORDER REPORT").font = Font(bold=True, size=14)
    k1 = ws.cell(row=1, column=11, value="=TODAY()")
    k1.font = Font(bold=True, size=14)
    k1.number_format = "DD/MM/YYYY"

    # Headers (Row 2)
    headers = [
        "Date", "Delivery Date", "Job Number", "SO Number", "Customer Ref.", 
        "Customer", "Total", "Sales Rep", "Status", "Payment Status", 
        "Details", "Planned Delivery Date"
    ]
    for col_idx, h_val in enumerate(headers, start=1):
        ws.cell(row=2, column=col_idx, value=h_val)
    
    # Write Data starting at Row 3
    for row_idx, row in enumerate(data_rows, start=3):
        so_num = str(row[3]).strip()
        if so_num in prev_details:
            row[10] = prev_details[so_num]

        for col_idx, val in enumerate(row, start=1):
             ws.cell(row=row_idx, column=col_idx, value=val)
        
        # Override Column L with formula
        ws.cell(row=row_idx, column=12).value = f"=B{row_idx}"

    # Formatting
    # Header Style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    # Define a standard border
    thin_border = Border(left=Side(style='thin'), 
                         right=Side(style='thin'), 
                         top=Side(style='thin'), 
                         bottom=Side(style='thin'))
                         
    # Define Light Green Fill for orders due today
    light_green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    # Define Yellow Fill for pending jobs
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    for cell in ws[2]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Freeze panes
    ws.freeze_panes = "E3"
        
    # Column Widths (Approximation)
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 8
    ws.column_dimensions['I'].width = 15
    ws.column_dimensions['J'].width = 23.3
    ws.column_dimensions['K'].width = 50 # Details
    ws.column_dimensions['L'].width = 12

    # Date Format and Data Row Borders
    date_style = "DD/MM/YYYY"
    today_formatted = TODAY_DT.strftime("%Y-%m-%d")
    
    for row in ws.iter_rows(min_row=3, max_row=len(data_rows)+2, max_col=12):
        # Check if Delivery Date (Column B / index 1) matches today
        # Note: openpyxl cells with dates store datetime objects
        is_due_today = False
        cell_date = row[1].value
        if isinstance(cell_date, datetime.datetime):
             if cell_date.date() == TODAY_DT.date():
                 is_due_today = True
        
        for cell in row:
            cell.border = thin_border
            
            # User defined Column Alignment
            col_letter = cell.column_letter
            if col_letter in ['A', 'B', 'C', 'D', 'H', 'I', 'J', 'L']:
                cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
            elif col_letter in ['E', 'F', 'K']:
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            else:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            
            if is_due_today:
                cell.fill = light_green_fill
            
            # Check if Status (Column I) is 'Pending' and highlight entire row in yellow
            if cell.column_letter == 'I':  # Status column
                if str(cell.value).strip().upper() == 'PENDING':
                    # Highlight entire row in yellow
                    for col_letter in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L']:
                        ws[col_letter + str(cell.row)].fill = yellow_fill

        # A (1) and B (2) and L (12) are dates
        row[0].number_format = date_style
        row[1].number_format = date_style
        row[11].number_format = date_style
        
        # G (7) is Currency
        row[6].number_format = '"R" #,##0.00'

    # Print Setup
    ws.page_setup.orientation = ws.ORIENTATION_LANDSCAPE
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0  # 0 means don't force fit to height (can span multiple pages vertically)
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:L{len(data_rows)+10}" # Include summary in print area
    
    # Set page margins (narrow)
    ws.page_margins.left = 0.25
    ws.page_margins.right = 0.25
    ws.page_margins.top = 0.75
    ws.page_margins.bottom = 0.75
    ws.page_margins.header = 0.3
    ws.page_margins.footer = 0.3

    # Summary Section
    last_row = len(data_rows) + 2
    
    # Total Value (All) - Row below last order, Title in Col F, Value in Col G
    # Excludes Pending orders (Status column I != "Pending")
    total_row = last_row + 1
    ws.cell(row=total_row, column=6, value="Total Value (All):").font = Font(bold=True)
    ws.cell(row=total_row, column=7, value=f'=SUMIF(I3:I{last_row},"<>Pending",G3:G{last_row})')
    ws.cell(row=total_row, column=7).number_format = '"R" #,##0.00'
    ws.cell(row=total_row, column=7).font = Font(bold=True)
    
    # Breakdown Section
    summary_start = total_row + 2 # Leave one empty row
    
    ws.cell(row=summary_start, column=6, value="BREAKDOWN").font = Font(bold=True, underline="single")
    
    # Label Layout (Side-by-Side)
    ws.cell(row=summary_start+1, column=6, value=f"Cash Sale ({THIS_MONTH_NAME}):")
    ws.cell(row=summary_start+2, column=6, value=f"Account ({THIS_MONTH_NAME}):")
    ws.cell(row=summary_start+1, column=9, value=f"Cash Sale ({NEXT_MONTH_NAME}):")
    ws.cell(row=summary_start+2, column=9, value=f"Account ({NEXT_MONTH_NAME}):")
    
    # Formulas Helper (Ensure data links to Row 3)
    # All breakdown formulas also exclude Pending orders via I3:I{last_row},"<>Pending"
    this_m_cond = f'L3:L{last_row}, ">="&EOMONTH(TODAY(),-1)+1, L3:L{last_row}, "<="&EOMONTH(TODAY(),0)'
    next_m_cond = f'L3:L{last_row}, ">="&EOMONTH(TODAY(),0)+1, L3:L{last_row}, "<="&EOMONTH(TODAY(),1)'
    excl_pending = f'I3:I{last_row}, "<>Pending"'

    # Cash Sale (This Month) -> Col 7  [excludes Pending]
    ws.cell(row=summary_start+1, column=7, value=f'=SUMIFS(G3:G{last_row}, J3:J{last_row}, "*CASH SALE*", {this_m_cond}, {excl_pending})')
    # Cash Sale (Next Month) -> Col 10  [excludes Pending]
    ws.cell(row=summary_start+1, column=10, value=f'=SUMIFS(G3:G{last_row}, J3:J{last_row}, "*CASH SALE*", {next_m_cond}, {excl_pending})')

    # Account (This Month) -> Col 7  [excludes Pending]
    ws.cell(row=summary_start+2, column=7, value=f'=SUMIFS(G3:G{last_row}, J3:J{last_row}, "*ACCOUNT*", {this_m_cond}, {excl_pending})')
    # Account (Next Month) -> Col 10  [excludes Pending]
    ws.cell(row=summary_start+2, column=10, value=f'=SUMIFS(G3:G{last_row}, J3:J{last_row}, "*ACCOUNT*", {next_m_cond}, {excl_pending})')

    # Formatting
    for r_os in (1, 2):
        ws.cell(row=summary_start+r_os, column=7).number_format = '"R" #,##0.00'
        ws.cell(row=summary_start+r_os, column=10).number_format = '"R" #,##0.00'

    # --- Append Due Today Detail Table ---
    rd_header_row = summary_start + 5
    
    # Re-write Headers (only columns A-J)
    for col_idx, header_val in enumerate(headers[:10], start=1):
        cell = ws.cell(row=rd_header_row, column=col_idx, value=header_val)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
    # Append Comments header in Column K (11)
    cell_k_header = ws.cell(row=rd_header_row, column=11, value="Comments")
    cell_k_header.font = header_font
    cell_k_header.fill = header_fill
    cell_k_header.border = thin_border
    cell_k_header.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    current_row = rd_header_row + 1
    rd_start_data_row = current_row
    
    for row_data in data_rows:
        del_date = row_data[1]  # Delivery Date
        is_due_today = False
        if isinstance(del_date, datetime.datetime):
             if del_date.date() == TODAY_DT.date():
                 is_due_today = True

        if is_due_today:
            for col_idx in range(10): # A to J
                cell = ws.cell(row=current_row, column=col_idx+1, value=row_data[col_idx])
                cell.border = thin_border
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                
                # Format Dates
                if col_idx in (0, 1):
                    cell.number_format = date_style
                # Format Currency
                elif col_idx == 6:
                    cell.number_format = '"R" #,##0.00'
                    
            # Fetch previous comment if exists
            so_num = str(row_data[3]).strip()
            comment_text = prev_comments.get(so_num, "")
                    
            # Apply border, left alignment, and text to Comments column (K)
            cell_k = ws.cell(row=current_row, column=11, value=comment_text)
            cell_k.border = thin_border
            cell_k.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            
            current_row += 1

    # Due Today Total Row (at bottom of new table)
    ws.cell(row=current_row, column=6, value="Due Today - Total:").font = Font(bold=True)
    if current_row > rd_start_data_row:
        ws.cell(row=current_row, column=7, value=f'=SUM(G{rd_start_data_row}:G{current_row-1})')
    else:
        ws.cell(row=current_row, column=7, value=0)
    ws.cell(row=current_row, column=7).number_format = '"R" #,##0.00'
    ws.cell(row=current_row, column=7).font = Font(bold=True)
    
    current_row += 2
    
    # Ready-Dispatch Value exclusively below the table
    ws.cell(row=current_row, column=6, value="Ready-Dispatch:").font = Font(bold=True)
    ws.cell(row=current_row, column=7, value=f'=SUMIFS(G3:G{last_row}, I3:I{last_row}, "*Ready*Dispatch*")')
    ws.cell(row=current_row, column=7).number_format = '"R" #,##0.00'
    ws.cell(row=current_row, column=7).font = Font(bold=True)

    ws.print_area = f"A1:L{current_row+2}"

    print("Saving workbook...", flush=True)
    try:
        wb.save(SALES_REPORT_FILE)
        print("Done!", flush=True)
    except PermissionError:
        print("ERROR: Could not save file. It might be open.", flush=True)

def main():
    data = generate_report_data()
    if data:
        write_to_excel(data)

if __name__ == "__main__":
    main()
