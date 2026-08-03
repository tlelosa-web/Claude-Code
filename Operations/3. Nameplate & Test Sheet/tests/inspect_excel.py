from pathlib import Path
import openpyxl

p = Path('2_Source_Data/raw_sources/NAME PLATE PROCEDURE.xlsx')
print('exists', p.exists())
wb = openpyxl.load_workbook(p, data_only=True)
print('sheets', wb.sheetnames)
if 'NamePlateProc' in wb.sheetnames:
    ws = wb['NamePlateProc']
    print('max_row', ws.max_row, 'max_column', ws.max_column)
    for r in range(1, min(40, ws.max_row) + 1):
        row = [ws.cell(r, c).value for c in range(1, min(100, ws.max_column) + 1)]
        if any(cell is not None for cell in row):
            print(r, row)
else:
    print('NamePlateProc not found')
