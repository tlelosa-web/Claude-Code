import codecs
import openpyxl

file_path = "Sales Order Report - 03.2026.xlsx"
wb = openpyxl.load_workbook(file_path, data_only=False)
sheet_name = "13.03.2026"

with codecs.open('read13_out.txt', 'w', 'utf-8') as f:
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        f.write("--- 13.03.2026 TOP 6 ROWS ---\n")
        for r in range(1, 7):
            row_data = [str(ws.cell(row=r, column=c).value) for c in range(1, 13)]
            f.write(f"Row {r}: {row_data}\n")
        
        f.write(f"Frozen Panes: {ws.freeze_panes}\n")
    else:
        f.write(f"Sheet {sheet_name} not found.\n")
